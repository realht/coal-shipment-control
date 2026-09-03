import datetime as _dt
from unittest.mock import patch as _patch

import pytest
from django import forms as _forms
from django.test import Client
from django.urls import reverse
from django.contrib.auth.models import Group, Permission


@pytest.mark.django_db
def test_index_redirects_anonymous():
    client = Client()
    response = client.get(reverse("index"))
    assert response.status_code == 302
    assert "/accounts/login/" in response["Location"]


@pytest.mark.django_db
def test_index_accessible_for_authenticated(django_user_model):
    user = django_user_model.objects.create_user(username="testuser", password="testpass")
    client = Client()
    client.login(username="testuser", password="testpass")
    response = client.get(reverse("index"))
    assert response.status_code == 200
    assert "testuser" in response.content.decode()


def test_settings_secret_key_present():
    from django.conf import settings
    assert settings.SECRET_KEY
    assert len(settings.SECRET_KEY) > 0


def test_pagination_goto_form_preserves_multivalue_filters():
    """V17-MED-8: форма «Перейти» сохраняет все повторяющиеся значения фильтра."""
    from django.core.paginator import Paginator
    from django.template.loader import render_to_string
    from django.test import RequestFactory

    request = RequestFactory().get("/?f_coal_grade=%D0%94%D0%93&f_coal_grade=%D0%94%D0%A0")
    paginator = Paginator(list(range(10)), 3)
    page_obj = paginator.page(1)

    html = render_to_string(
        "partials/pagination.html",
        {"request": request, "page_obj": page_obj, "pages": []},
    )

    assert 'value="ДГ"' in html
    assert 'value="ДР"' in html


def test_page_url_preserves_multivalue_filters():
    """V17-MED-8: номерные ссылки пагинации сохраняют все значения мультифильтра."""
    from django.test import RequestFactory
    from core.templatetags.coal_tags import page_url

    request = RequestFactory().get("/?f_customer_object=A&f_customer_object=B")
    url = page_url({"request": request}, 2)
    assert "f_customer_object=A" in url
    assert "f_customer_object=B" in url
    assert "page=2" in url


def test_querystring_preserves_multivalue_filters():
    """V17-MED-8: тег querystring (экспорт-ссылки) не теряет значения мультифильтра."""
    from django.test import RequestFactory
    from core.templatetags.coal_tags import querystring

    request = RequestFactory().get("/?f_customer_object=A&f_customer_object=B")
    out = querystring({"request": request}, page=None)
    assert "f_customer_object=A" in out
    assert "f_customer_object=B" in out


def test_settings_database_is_sqlite():
    from django.conf import settings
    assert settings.DATABASES["default"]["ENGINE"] == "django.db.backends.sqlite3"


@pytest.mark.django_db
def test_index_can_add_export_for_rail_only_user(client):
    """User with only rail add/export sees can_add=True and can_export=True."""
    from django.contrib.auth import get_user_model
    User = get_user_model()
    user = User.objects.create_user("rail_only_L5", password="x")
    user.user_permissions.add(
        Permission.objects.get(codename="view_railshipment"),
        Permission.objects.get(codename="add_railshipment"),
        Permission.objects.get(codename="export_excel", content_type__app_label="shipments_rail"),
    )
    client.force_login(user)
    response = client.get("/")
    assert response.status_code == 200
    assert response.context["can_add"] is True
    assert response.context["can_export"] is True


@pytest.mark.django_db
class TestDashboard:
    @pytest.fixture
    def viewer_auto(self, django_user_model):
        user = django_user_model.objects.create_user(username="dash_viewer", password="pass")
        group, _ = Group.objects.get_or_create(name="dash_viewer_group")
        for app, codename in [
            ("shipments_auto", "view_autoshipment"),
            ("shipments_rail", "view_railshipment"),
        ]:
            perm = Permission.objects.get(codename=codename, content_type__app_label=app)
            group.permissions.add(perm)
        user.groups.add(group)
        return user

    def test_dashboard_context_has_stats(self, client, viewer_auto):
        from shipments_auto.models import AutoShipment
        from shipments_rail.models import RailShipment
        AutoShipment.objects.create(
            shipment_date="2026-05-01", customer_object="А", coal_grade="ДГ", quantity="500",
            created_by=viewer_auto, updated_by=viewer_auto,
        )
        RailShipment.objects.create(
            departure_date="2026-05-02", wagon_number="12345", cargo="Уголь Д",
            receiver="Завод", volume="1000",
            created_by=viewer_auto, updated_by=viewer_auto,
        )
        client.login(username="dash_viewer", password="pass")
        response = client.get(reverse("index"))
        assert response.status_code == 200
        ctx = response.context
        assert "auto_count_month" in ctx
        assert "rail_count_month" in ctx
        assert "auto_by_grade" in ctx
        assert "rail_by_grade" in ctx

    def test_dashboard_no_stats_for_no_perm(self, django_user_model, client):
        user = django_user_model.objects.create_user(username="noperm", password="pass")
        client.login(username="noperm", password="pass")
        response = client.get(reverse("index"))
        assert response.status_code == 200
        assert "auto_count_month" not in response.context

    def test_auto_no_docs_widget(self, client, viewer_auto):
        from shipments_auto.models import AutoShipment
        AutoShipment.objects.create(
            shipment_date="2026-05-01", customer_object="Без_доков", coal_grade="ДГ", quantity="10",
            created_by=viewer_auto, updated_by=viewer_auto,
        )
        client.login(username="dash_viewer", password="pass")
        response = client.get(reverse("index"))
        assert response.context["auto_no_docs"] >= 1

    def test_auto_ttn_no_file_widget(self, client, viewer_auto):
        from shipments_auto.models import AutoShipment
        AutoShipment.objects.create(
            shipment_date="2026-05-01", customer_object="С_ТТН", coal_grade="ДГ",
            quantity="10", ttn_number="ТТН-999",
            created_by=viewer_auto, updated_by=viewer_auto,
        )
        client.login(username="dash_viewer", password="pass")
        response = client.get(reverse("index"))
        assert response.context["auto_ttn_no_file"] >= 1


@pytest.mark.django_db
class TestDuplicatesPage:
    @pytest.fixture
    def viewer_user(self, django_user_model):
        user = django_user_model.objects.create_user(username="dup_viewer", password="pass")
        group, _ = Group.objects.get_or_create(name="dup_viewer_group")
        for app, codename in [
            ("shipments_auto", "view_autoshipment"),
            ("shipments_rail", "view_railshipment"),
        ]:
            perm = Permission.objects.get(codename=codename, content_type__app_label=app)
            group.permissions.add(perm)
        user.groups.add(group)
        return user

    @pytest.fixture
    def operator_user(self, django_user_model):
        user = django_user_model.objects.create_user(username="dup_operator", password="secret")
        group, _ = Group.objects.get_or_create(name="dup_operator_group")
        for app, codename in [
            ("shipments_auto", "view_autoshipment"),
            ("shipments_auto", "delete_autoshipment"),
            ("shipments_rail", "view_railshipment"),
            ("shipments_rail", "delete_railshipment"),
        ]:
            perm = Permission.objects.get(codename=codename, content_type__app_label=app)
            group.permissions.add(perm)
        user.groups.add(group)
        return user

    def test_anonymous_redirects(self, client):
        response = client.get(reverse("duplicates"))
        assert response.status_code == 302
        assert "/accounts/login/" in response["Location"]

    def test_no_view_perm_redirects_to_index(self, client, django_user_model):
        user = django_user_model.objects.create_user(username="dup_noperm", password="pass")
        client.login(username="dup_noperm", password="pass")
        response = client.get(reverse("duplicates"))
        assert response.status_code == 302
        assert response["Location"] == reverse("index")

    def test_viewer_can_access(self, client, viewer_user):
        client.login(username="dup_viewer", password="pass")
        response = client.get(reverse("duplicates"))
        assert response.status_code == 200

    def test_invalid_get_shipment_type_returns_400(self, client, viewer_user):
        client.login(username="dup_viewer", password="pass")
        response = client.get(reverse("duplicates") + "?type=bad")
        assert response.status_code == 400

    def test_no_results_without_search(self, client, viewer_user):
        client.login(username="dup_viewer", password="pass")
        response = client.get(reverse("duplicates"))
        assert response.context["searched"] is False
        assert response.context["clusters"] == []

    def test_search_returns_clusters(self, client, viewer_user):
        from shipments_auto.models import AutoShipment
        for _ in range(2):
            AutoShipment.objects.create(
                shipment_date="2026-03-01",
                customer_object="Завод",
                coal_grade="ДГ",
                quantity="500",
                vehicle_number="А001АА",
                driver_name="Иванов",
            )
        client.login(username="dup_viewer", password="pass")
        response = client.get(reverse("duplicates") + "?type=auto")
        assert response.context["searched"] is True
        assert len(response.context["clusters"]) >= 1

    def test_decimal_scale_does_not_split_duplicate_cluster(self):
        from decimal import Decimal
        from core.views import _find_duplicates

        rows = [
            {
                "id": 1,
                "customer_object": "Завод",
                "coal_grade": "ДГ",
                "quantity": Decimal("5.0"),
            },
            {
                "id": 2,
                "customer_object": "Завод",
                "coal_grade": "ДГ",
                "quantity": Decimal("5.00"),
            },
        ]

        class FakeObject:
            def __init__(self, pk):
                self.pk = pk

        class FakeQuerySet:
            def __init__(self, values):
                self.values_list = values

            def filter(self, **kwargs):
                return self

            def values(self, *fields):
                return self

            def order_by(self, *fields):
                return self

            def __getitem__(self, key):
                return FakeQuerySet(self.values_list[key])

            def __iter__(self):
                return iter(self.values_list)

        class FakeManager:
            def filter(self, **kwargs):
                if "pk__in" in kwargs:
                    return [FakeObject(pk) for pk in kwargs["pk__in"]]
                return FakeQuerySet(rows)

        class FakeModel:
            objects = FakeManager()

        clusters = _find_duplicates(
            FakeModel,
            ["customer_object", "coal_grade", "quantity"],
            "id",
            "",
            "",
        )

        assert [[obj.pk for obj in cluster] for cluster in clusters] == [[1, 2]]

    def test_case_and_whitespace_do_not_split_duplicate_cluster(self):
        from decimal import Decimal
        from core.views import _find_duplicates

        rows = [
            {
                "id": 1,
                "customer_object": "Завод ",
                "coal_grade": "ДГ",
                "quantity": Decimal("5.0"),
            },
            {
                "id": 2,
                "customer_object": "завод",
                "coal_grade": "дг",
                "quantity": Decimal("5.0"),
            },
        ]

        class FakeObject:
            def __init__(self, pk):
                self.pk = pk

        class FakeQuerySet:
            def __init__(self, values):
                self.values_list = values

            def filter(self, **kwargs):
                return self

            def values(self, *fields):
                return self

            def order_by(self, *fields):
                return self

            def __getitem__(self, key):
                return FakeQuerySet(self.values_list[key])

            def __iter__(self):
                return iter(self.values_list)

        class FakeManager:
            def filter(self, **kwargs):
                if "pk__in" in kwargs:
                    return [FakeObject(pk) for pk in kwargs["pk__in"]]
                return FakeQuerySet(rows)

        class FakeModel:
            objects = FakeManager()

        clusters = _find_duplicates(
            FakeModel,
            ["customer_object", "coal_grade", "quantity"],
            "id",
            "",
            "",
        )

        assert [[obj.pk for obj in cluster] for cluster in clusters] == [[1, 2]]

    def test_date_range_filters_results(self, client, viewer_user):
        from shipments_auto.models import AutoShipment
        for _ in range(2):
            AutoShipment.objects.create(
                shipment_date="2026-01-10",
                customer_object="Завод",
                coal_grade="ДГ",
                quantity="500",
                vehicle_number="А001АА",
                driver_name="Иванов",
            )
        client.login(username="dup_viewer", password="pass")
        response = client.get(
            reverse("duplicates") + "?type=auto&date_from=2026-02-01&date_to=2026-02-28"
        )
        assert response.context["clusters"] == []

    def test_invalid_date_range_is_ignored_without_500(self, client, viewer_user):
        client.login(username="dup_viewer", password="pass")
        response = client.get(reverse("duplicates") + "?type=auto&date_from=abc&date_to=zzzz")

        assert response.status_code == 200
        assert response.context["searched"] is True

    def test_no_duplicates_fewer_than_3_matches(self, client, viewer_user):
        from shipments_auto.models import AutoShipment
        AutoShipment.objects.create(
            shipment_date="2026-03-01", customer_object="Завод-А",
            coal_grade="ДГ", quantity="100", vehicle_number="А001АА",
        )
        AutoShipment.objects.create(
            shipment_date="2026-03-02", customer_object="Завод-Б",
            coal_grade="Т", quantity="200", vehicle_number="А001АА",
        )
        client.login(username="dup_viewer", password="pass")
        response = client.get(reverse("duplicates") + "?type=auto")
        assert response.context["clusters"] == []

    def test_viewer_cannot_delete(self, client, viewer_user):
        from shipments_auto.models import AutoShipment
        obj = AutoShipment.objects.create(
            shipment_date="2026-03-01", customer_object="Завод",
            coal_grade="ДГ", quantity="500",
        )
        client.login(username="dup_viewer", password="pass")
        client.post(reverse("duplicates"), {"pk": str(obj.pk), "shipment_type": "auto"})
        obj.refresh_from_db()
        assert obj.is_deleted is False

    def test_delete_invalid_post_shipment_type_returns_400(self, client, operator_user):
        client.login(username="dup_operator", password="secret")
        response = client.post(reverse("duplicates"), {"pk": "1", "shipment_type": "bad"})
        assert response.status_code == 400

    @pytest.mark.parametrize("bad_pk", ["abc", "", "0"])
    def test_delete_invalid_pk_returns_400(self, client, operator_user, bad_pk):
        client.login(username="dup_operator", password="secret")
        response = client.post(reverse("duplicates"), {"pk": bad_pk, "shipment_type": "auto"})
        assert response.status_code == 400

    def test_delete_missing_numeric_pk_redirects_without_500(self, client, operator_user):
        client.login(username="dup_operator", password="secret")
        response = client.post(reverse("duplicates"), {"pk": "999999", "shipment_type": "auto"})
        assert response.status_code == 302
        assert response["Location"].startswith("/duplicates/")

    def test_operator_can_delete_auto(self, client, operator_user):
        from shipments_auto.models import AutoShipment
        obj = AutoShipment.objects.create(
            shipment_date="2026-03-01", customer_object="Завод",
            coal_grade="ДГ", quantity="500",
        )
        client.login(username="dup_operator", password="secret")
        response = client.post(
            reverse("duplicates"),
            {"pk": str(obj.pk), "shipment_type": "auto"},
        )
        assert response.status_code == 302
        obj.refresh_from_db()
        assert obj.is_deleted is True

    def test_delete_auto_writes_audit(self, client, operator_user):
        from audit.models import AuditLog
        from shipments_auto.models import AutoShipment
        obj = AutoShipment.objects.create(
            shipment_date="2026-03-01", customer_object="Аудит",
            coal_grade="ДГ", quantity="100",
        )
        client.login(username="dup_operator", password="secret")
        client.post(reverse("duplicates"), {"pk": str(obj.pk), "shipment_type": "auto"})
        assert AuditLog.objects.filter(
            entity_type=AuditLog.ENTITY_AUTO,
            entity_id=obj.pk,
            action=AuditLog.ACTION_DELETE,
        ).exists()

    def test_operator_can_delete_rail(self, client, operator_user):
        from shipments_rail.models import RailShipment
        obj = RailShipment.objects.create(
            departure_date="2026-03-10", wagon_number="11223344",
            cargo="Уголь ДГ", receiver="ООО Завод", volume="1000",
        )
        client.login(username="dup_operator", password="secret")
        response = client.post(
            reverse("duplicates"),
            {"pk": str(obj.pk), "shipment_type": "rail"},
        )
        assert response.status_code == 302
        obj.refresh_from_db()
        assert obj.is_deleted is True

    def test_rail_duplicates_found(self, client, viewer_user):
        from shipments_rail.models import RailShipment
        for _ in range(2):
            RailShipment.objects.create(
                departure_date="2026-03-10",
                wagon_number="11223344",
                cargo="Уголь ДГ",
                receiver="ООО Завод",
                volume="1000",
                document_number="ТД-001",
            )
        client.login(username="dup_viewer", password="pass")
        response = client.get(reverse("duplicates") + "?type=rail")
        assert len(response.context["clusters"]) >= 1

    def test_duplicates_post_skips_search(self, client, operator_user):
        """POST to /duplicates/ must not invoke _find_duplicates (avoids heavy query)."""
        from unittest.mock import patch
        client.login(username="dup_operator", password="secret")
        with patch("core.views._find_duplicates") as spy:
            client.post("/duplicates/?type=auto", {"pk": "999999", "shipment_type": "auto"})
            spy.assert_not_called()

    def test_duplicates_post_redirect_uses_urlencode(self, client, operator_user):
        """Redirect after POST must URL-encode date parameters (pk 999999 = not found → 302 with msg)."""
        client.login(username="dup_operator", password="secret")
        response = client.post(
            "/duplicates/",
            {"pk": "999999", "shipment_type": "auto", "date_from": "2026-01-01", "date_to": "2026-12-31"},
        )
        assert response.status_code == 302
        location = response["Location"]
        assert "date_from=2026-01-01" in location
        assert "date_to=2026-12-31" in location


class TestDuplicateSearch:
    """Unit tests for _find_duplicates() internals (V15-M1)."""

    def test_find_duplicates_slices_queryset_before_materialize(self):
        """V15-M1: queryset нарезается [:MAX+1] до list(), а не после."""
        from unittest.mock import MagicMock, patch
        from core.views import _find_duplicates
        import core.views as cv

        sliced = MagicMock()
        sliced.__iter__ = lambda s: iter([{"id": i, "f": "v"} for i in range(3)])
        ordered = MagicMock()
        ordered.__getitem__ = MagicMock(return_value=sliced)
        values_qs = MagicMock()
        values_qs.order_by.return_value = ordered

        class FakeManager:
            def filter(self, **kw): return self
            def values(self, *a): return values_qs

        class FakeModel:
            objects = FakeManager()

        with patch.object(cv, "MAX_DUPLICATE_RECORDS", 2):
            with pytest.raises(ValueError, match="Слишком много записей"):
                _find_duplicates(FakeModel, ["f"], "f", None, None)

        # Ключевая проверка: __getitem__ вызван с slice, не просто __iter__
        ordered.__getitem__.assert_called_once_with(slice(None, 3, None))


class TestRequestBodySizeLimitMiddleware:
    """V18-MED-5: app-level guard rejects oversized bodies via Content-Length."""

    def _make_middleware(self):
        from core.middleware import RequestBodySizeLimitMiddleware

        sentinel = object()

        def get_response(request):
            get_response.called = True
            return sentinel

        get_response.called = False
        get_response.sentinel = sentinel
        return RequestBodySizeLimitMiddleware(get_response), get_response

    def test_oversized_post_returns_413_and_skips_view(self, settings):
        from django.test import RequestFactory

        settings.MAX_REQUEST_BODY_SIZE_BYTES = 1024
        middleware, get_response = self._make_middleware()
        request = RequestFactory().post("/import/")
        request.META["CONTENT_LENGTH"] = str(1024 * 10)

        response = middleware(request)

        assert response.status_code == 413
        assert get_response.called is False

    def test_within_limit_passes_through(self, settings):
        from django.test import RequestFactory

        settings.MAX_REQUEST_BODY_SIZE_BYTES = 1024
        middleware, get_response = self._make_middleware()
        request = RequestFactory().post("/import/")
        request.META["CONTENT_LENGTH"] = str(512)

        response = middleware(request)

        assert response is get_response.sentinel
        assert get_response.called is True

    def test_missing_content_length_passes_through(self, settings):
        from django.test import RequestFactory

        settings.MAX_REQUEST_BODY_SIZE_BYTES = 1024
        middleware, get_response = self._make_middleware()
        request = RequestFactory().post("/import/")
        request.META.pop("CONTENT_LENGTH", None)

        response = middleware(request)

        assert response is get_response.sentinel
        assert get_response.called is True

    def test_non_numeric_content_length_passes_through(self, settings):
        from django.test import RequestFactory

        settings.MAX_REQUEST_BODY_SIZE_BYTES = 1024
        middleware, get_response = self._make_middleware()
        request = RequestFactory().post("/import/")
        request.META["CONTENT_LENGTH"] = "chunked"

        response = middleware(request)

        assert response is get_response.sentinel
        assert get_response.called is True

    def test_get_request_not_limited(self, settings):
        from django.test import RequestFactory

        settings.MAX_REQUEST_BODY_SIZE_BYTES = 1024
        middleware, get_response = self._make_middleware()
        request = RequestFactory().get("/")
        request.META["CONTENT_LENGTH"] = str(1024 * 10)

        response = middleware(request)

        assert response is get_response.sentinel
        assert get_response.called is True


@pytest.mark.django_db
class TestHealthEndpoint:
    def test_health_requires_login(self, client):
        response = client.get(reverse("core:health"))
        assert response.status_code == 302
        assert "/accounts/login/" in response["Location"]

    def test_health_returns_200_when_authenticated(self, client, django_user_model):
        django_user_model.objects.create_user(username="health_user", password="pass")
        client.login(username="health_user", password="pass")
        response = client.get(reverse("core:health"))
        assert response.status_code == 200

    def test_health_returns_json_with_db(self, client, django_user_model):
        import json
        django_user_model.objects.create_user(username="health_user2", password="pass")
        client.login(username="health_user2", password="pass")
        response = client.get(reverse("core:health"))
        data = json.loads(response.content)
        assert data == {"status": "ok", "db": True}

    def test_health_returns_503_and_logs_when_database_check_fails(self, client, django_user_model):
        import json
        from unittest.mock import patch

        django_user_model.objects.create_user(username="health_user3", password="pass")
        client.login(username="health_user3", password="pass")

        with (
            patch("core.views.database_health", side_effect=RuntimeError("db down")),
            patch("core.views.logger.exception") as mock_exception,
        ):
            response = client.get(reverse("core:health"))

        data = json.loads(response.content)
        assert response.status_code == 503
        assert data == {"status": "error", "db": False}
        mock_exception.assert_called_once_with("Database health check failed")


@pytest.mark.django_db
class TestHealthzEndpoint:
    def test_healthz_returns_200(self, client):
        response = client.get(reverse("core:healthz"))
        assert response.status_code == 200

    def test_healthz_returns_json_ok(self, client):
        import json
        response = client.get(reverse("core:healthz"))
        data = json.loads(response.content)
        assert data == {"status": "ok"}

    def test_healthz_accessible_without_login(self, client):
        response = client.get(reverse("core:healthz"))
        assert response.status_code == 200


@pytest.mark.django_db
class TestReadyzEndpoint:
    def _make_dirs(self, settings, tmp_path, *, media_ok=True, backup_ok=True):
        media_dir = tmp_path / "uploads"
        backup_dir = tmp_path / "backups"
        if media_ok:
            media_dir.mkdir()
        if backup_ok:
            backup_dir.mkdir()
        settings.MEDIA_ROOT = str(media_dir)
        settings.BACKUP_DIR = str(backup_dir)

    def _touch_heartbeat(self):
        from core.system_ops import get_system_state
        from django.utils import timezone

        state = get_system_state()
        state.scheduler_heartbeat_at = timezone.now()
        state.save(update_fields=["scheduler_heartbeat_at"])

    def test_readyz_accessible_without_login(self, client):
        response = client.get(reverse("core:readyz"))
        assert response.status_code != 302

    def test_readyz_returns_ok_when_healthy(self, client, settings, tmp_path):
        import json

        self._make_dirs(settings, tmp_path)
        self._touch_heartbeat()

        response = client.get(reverse("core:readyz"))
        data = json.loads(response.content)

        assert response.status_code == 200
        assert data["status"] == "ok"
        assert data["checks"]["database"]["ok"] is True
        assert data["checks"]["media_root"]["ok"] is True
        assert data["checks"]["backup_dir"]["ok"] is True
        assert data["checks"]["scheduler"]["ok"] is True
        assert data["checks"]["system_mode"] == {"ok": True, "mode": "normal"}

    def test_readyz_returns_503_and_error_when_database_check_fails(self, client, settings, tmp_path):
        import json
        from unittest.mock import patch

        self._make_dirs(settings, tmp_path)
        self._touch_heartbeat()

        with patch("core.system_ops.database_health", side_effect=RuntimeError("db down")):
            response = client.get(reverse("core:readyz"))
        data = json.loads(response.content)

        assert response.status_code == 503
        assert data["status"] == "error"
        assert data["checks"]["database"]["ok"] is False

    def test_readyz_degraded_when_backup_dir_missing(self, client, settings, tmp_path):
        import json

        self._make_dirs(settings, tmp_path, backup_ok=False)
        self._touch_heartbeat()

        response = client.get(reverse("core:readyz"))
        data = json.loads(response.content)

        assert response.status_code == 200
        assert data["status"] == "degraded"
        assert data["checks"]["backup_dir"]["ok"] is False
        assert data["checks"]["database"]["ok"] is True
        assert data["checks"]["media_root"]["ok"] is True

    def test_readyz_degraded_when_scheduler_heartbeat_missing(self, client, settings, tmp_path):
        import json

        self._make_dirs(settings, tmp_path)

        response = client.get(reverse("core:readyz"))
        data = json.loads(response.content)

        assert response.status_code == 200
        assert data["status"] == "degraded"
        assert data["checks"]["scheduler"]["ok"] is False

    def test_readyz_degraded_when_system_mode_not_normal(self, client, settings, tmp_path):
        import json
        from core.models import SystemState
        from core.system_ops import set_system_mode

        self._make_dirs(settings, tmp_path)
        self._touch_heartbeat()
        set_system_mode(SystemState.MODE_ADMIN_ONLY, reason="maintenance")

        response = client.get(reverse("core:readyz"))
        data = json.loads(response.content)

        assert response.status_code == 200
        assert data["status"] == "degraded"
        assert data["checks"]["system_mode"] == {"ok": False, "mode": "admin_only"}


# ── Tests for CatalogHybridFormMixin ─────────────────────────────────────────

def _make_fake_form_class(entity, all_fields, labels, date_field):
    """Concrete test form using the mixin without a real Django model."""
    from core.forms import CatalogHybridFormMixin

    class FakeBaseForm(_forms.Form):
        def __init__(self, *args, instance=None, **kwargs):
            self.instance = instance or type("FakeInstance", (), {"pk": None})()
            super().__init__(*args, **kwargs)

    class FakeForm(CatalogHybridFormMixin, FakeBaseForm):
        pass

    FakeForm.entity = entity
    FakeForm.all_fields = all_fields
    FakeForm.labels = labels
    FakeForm.date_field = date_field
    for field_name in all_fields:
        FakeForm.base_fields[field_name] = _forms.CharField(required=False)
    return FakeForm


_MIXIN_FIELDS = ["test_date", "category", "comment"]
_MIXIN_LABELS = {"test_date": "Дата теста", "category": "Категория", "comment": "Комментарий"}

_MIXIN_CFG_BASE = {
    "test_date": {
        "visible": True, "required": True, "use_catalog": False,
        "section": "main", "sort_order": 0, "label": "Дата теста",
        "is_system": True, "show_in_list": True,
        "allow_filter": False, "allow_sort": False,
        "filter_type": "none", "sticky_col": False, "preset_membership": "",
    },
    "category": {
        "visible": True, "required": False, "use_catalog": False,
        "section": "main", "sort_order": 1, "label": "Категория",
        "is_system": False, "show_in_list": True,
        "allow_filter": False, "allow_sort": False,
        "filter_type": "none", "sticky_col": False, "preset_membership": "",
    },
    "comment": {
        "visible": True, "required": False, "use_catalog": False,
        "section": "advanced", "sort_order": 2, "label": "Комментарий",
        "is_system": False, "show_in_list": True,
        "allow_filter": False, "allow_sort": False,
        "filter_type": "none", "sticky_col": False, "preset_membership": "",
    },
}
_MIXIN_CFG_WITH_CATALOG = {
    **_MIXIN_CFG_BASE,
    "category": {**_MIXIN_CFG_BASE["category"], "use_catalog": True, "required": True},
}


class TestCatalogHybridFormMixinConstants:
    def test_catalog_sentinel_value(self):
        from core.forms import CatalogHybridFormMixin
        assert CatalogHybridFormMixin.CATALOG_SENTINEL == "__other__"


class TestCatalogHybridFormMixinCatalogChoices:
    def test_blank_label_first(self):
        from core.forms import CatalogHybridFormMixin

        class FakeObj:
            def __init__(self, name):
                self.name = name

        choices = CatalogHybridFormMixin._lazy_catalog_choices([FakeObj("Alpha")])()
        assert choices[0] == ("", "—")

    def test_sentinel_last(self):
        from core.forms import CatalogHybridFormMixin
        choices = CatalogHybridFormMixin._lazy_catalog_choices([])()
        assert choices[-1] == ("__other__", "Другое…")

    def test_custom_blank_label(self):
        from core.forms import CatalogHybridFormMixin
        choices = CatalogHybridFormMixin._lazy_catalog_choices([], blank_label="выберите")()
        assert choices[0] == ("", "выберите")


class TestCatalogHybridFormMixinBuildFieldLists:
    def test_main_and_advanced_split(self):
        FakeForm = _make_fake_form_class("fake", _MIXIN_FIELDS, _MIXIN_LABELS, "test_date")
        with _patch("core.forms.get_entity_config", return_value=_MIXIN_CFG_BASE):
            main, advanced = FakeForm._build_field_lists()
        assert "test_date" in main
        assert "category" in main
        assert "comment" in advanced

    def test_invisible_field_excluded(self):
        cfg = {
            **_MIXIN_CFG_BASE,
            "category": {**_MIXIN_CFG_BASE["category"], "visible": False},
        }
        FakeForm = _make_fake_form_class("fake", _MIXIN_FIELDS, _MIXIN_LABELS, "test_date")
        with _patch("core.forms.get_entity_config", return_value=cfg):
            main, advanced = FakeForm._build_field_lists()
        assert "category" not in main
        assert "category" not in advanced

    def test_catalog_field_gets_select_suffix(self):
        FakeForm = _make_fake_form_class("fake", _MIXIN_FIELDS, _MIXIN_LABELS, "test_date")
        with _patch("core.forms.get_entity_config", return_value=_MIXIN_CFG_WITH_CATALOG):
            main, _ = FakeForm._build_field_lists()
        assert "category_select" in main
        assert "category" not in main


class TestCatalogHybridFormMixinClean:
    _BASE = {"test_date": "2026-01-15", "comment": "x"}

    def _cfg(self, required):
        return {
            "category": {
                "visible": True, "required": required, "use_catalog": True,
                "section": "main", "sort_order": 1, "label": "Категория",
                "is_system": False, "show_in_list": True,
                "allow_filter": False, "allow_sort": False,
                "filter_type": "none", "sticky_col": False, "preset_membership": "",
            }
        }

    def test_sentinel_without_other_raises_error(self):
        FakeForm = _make_fake_form_class("fake", _MIXIN_FIELDS, _MIXIN_LABELS, "test_date")
        data = {**self._BASE, "category_select": "__other__", "category_other": ""}
        with _patch("core.forms.get_entity_config", return_value=self._cfg(required=True)):
            form = FakeForm(data=data)
            form._catalog_fields = {"category"}
            assert not form.is_valid()
            assert "category_other" in form.errors

    def test_sentinel_with_other_sets_field_value(self):
        FakeForm = _make_fake_form_class("fake", _MIXIN_FIELDS, _MIXIN_LABELS, "test_date")
        data = {**self._BASE, "category_select": "__other__", "category_other": "CustomVal"}
        with _patch("core.forms.get_entity_config", return_value=self._cfg(required=True)):
            form = FakeForm(data=data)
            form._catalog_fields = {"category"}
            form.is_valid()
            assert form.cleaned_data.get("category") == "CustomVal"

    def test_required_catalog_empty_select_raises_error(self):
        FakeForm = _make_fake_form_class("fake", _MIXIN_FIELDS, _MIXIN_LABELS, "test_date")
        data = {**self._BASE, "category_select": "", "category_other": ""}
        with _patch("core.forms.get_entity_config", return_value=self._cfg(required=True)):
            form = FakeForm(data=data)
            form._catalog_fields = {"category"}
            assert not form.is_valid()
            assert "category_select" in form.errors

    def test_optional_catalog_empty_select_valid(self):
        FakeForm = _make_fake_form_class("fake", _MIXIN_FIELDS, _MIXIN_LABELS, "test_date")
        data = {**self._BASE, "category_select": "", "category_other": ""}
        with _patch("core.forms.get_entity_config", return_value=self._cfg(required=False)):
            form = FakeForm(data=data)
            form._catalog_fields = {"category"}
            assert form.is_valid()
            assert form.cleaned_data.get("category") == ""


def test_build_shipment_xlsx():
    from unittest.mock import MagicMock
    from django.http import StreamingHttpResponse
    from core.shipment_views import build_shipment_xlsx

    mock_qs = MagicMock()
    mock_qs.iterator.return_value = iter([MagicMock(), MagicMock()])
    headers = ["Дата", "Объект", "Кол-во"]

    def row_getter(s):
        return ["2026-01-01", "Завод А", "100"]

    response = build_shipment_xlsx(mock_qs, "test.xlsx", "Отгрузки", headers, row_getter)

    assert isinstance(response, StreamingHttpResponse), (
        f"Ожидался StreamingHttpResponse, получен {type(response).__name__}"
    )
    content = b"".join(response.streaming_content)
    mock_qs.iterator.assert_called_once()
    assert content[:4] == b"PK\x03\x04", "Ответ не содержит XLSX magic bytes"
    assert response["Content-Disposition"] == 'attachment; filename="test.xlsx"'
    assert "spreadsheetml" in response["Content-Type"]


def test_build_shipment_xlsx_decimal_precision():
    from decimal import Decimal
    from io import BytesIO
    from unittest.mock import MagicMock
    import openpyxl
    from core.shipment_views import build_shipment_xlsx

    mock_qs = MagicMock()
    mock_qs.iterator.return_value = iter([MagicMock()])
    expected = Decimal("1234.567")

    response = build_shipment_xlsx(
        mock_qs, "test.xlsx", "Тест", ["Значение"],
        lambda s: [expected],
    )
    content = b"".join(response.streaming_content)
    wb = openpyxl.load_workbook(BytesIO(content))
    ws = wb.active
    cell_value = ws.cell(2, 1).value
    assert cell_value is not None, "Ячейка пуста"
    assert isinstance(cell_value, (int, float)), (
        f"Ожидался числовой тип, получен {type(cell_value).__name__}"
    )
    assert abs(cell_value - float(expected)) < 1e-9


def test_build_shipment_xlsx_neutralizes_formula_like_strings():
    from decimal import Decimal
    from io import BytesIO
    from unittest.mock import MagicMock
    import openpyxl
    from core.shipment_views import build_shipment_xlsx

    mock_qs = MagicMock()
    mock_qs.iterator.return_value = iter([MagicMock()])
    dangerous = ["=1+1", "+SUM(A1:A2)", "-2+3", "@HYPERLINK(\"x\")"]

    response = build_shipment_xlsx(
        mock_qs,
        "test.xlsx",
        "Тест",
        ["A", "B", "C", "D", "Negative number"],
        lambda s: [*dangerous, Decimal("-12.5")],
    )

    content = b"".join(response.streaming_content)
    wb = openpyxl.load_workbook(BytesIO(content), data_only=False)
    ws = wb.active
    values = [ws.cell(2, idx).value for idx in range(1, 5)]
    data_types = [ws.cell(2, idx).data_type for idx in range(1, 5)]
    assert values == dangerous
    assert data_types == ["s", "s", "s", "s"]
    assert ws.cell(2, 5).data_type == "n"


@pytest.mark.django_db
class TestFieldSettingsBulkUpdate:
    """Regression tests for V14-L12: bulk_update replaces per-row save()."""

    @pytest.fixture
    def editor_user(self, django_user_model):
        user = django_user_model.objects.create_user(username="fs_editor", password="pass")
        perm = Permission.objects.get(codename="change_fieldsettings", content_type__app_label="core")
        user.user_permissions.add(perm)
        return user

    @pytest.fixture
    def fs_records(self):
        from core.models import FieldSettings
        FieldSettings.objects.update_or_create(
            entity="auto_shipment", field_name="customer_object",
            defaults=dict(
                sort_order=10, visible=True, required=True,
                show_in_list=True, section=FieldSettings.SECTION_MAIN,
                allow_filter=False, allow_sort=False, filter_type="none",
                sticky_col=False, preset_membership="",
                is_system=False,
            ),
        )
        FieldSettings.objects.update_or_create(
            entity="auto_shipment", field_name="shipment_date",
            defaults=dict(
                sort_order=20, visible=True, required=True,
                show_in_list=True, section=FieldSettings.SECTION_MAIN,
                allow_filter=False, allow_sort=False, filter_type="none",
                sticky_col=False, preset_membership="",
                is_system=True,
            ),
        )

    def test_post_saves_sort_order_visible_allow_filter(self, client, editor_user, fs_records):
        """bulk_update persists sort_order, visible and allow_filter correctly."""
        from core.models import FieldSettings
        client.force_login(editor_user)
        response = client.post(
            "/settings/fields/",
            {
                "entity": "auto_shipment",
                "field_order": "shipment_date,customer_object",
                # customer_object: visible, required (model-required), allow_filter with type
                "visible_customer_object": "on",
                "required_customer_object": "on",
                "show_in_list_customer_object": "on",
                "allow_filter_customer_object": "on",
                "filter_type_customer_object": "value",
                # shipment_date (system): allow_filter=False (not in POST)
            },
        )
        assert response.status_code == 302

        fs_customer = FieldSettings.objects.get(entity="auto_shipment", field_name="customer_object")
        fs_date = FieldSettings.objects.get(entity="auto_shipment", field_name="shipment_date")

        # sort_order reflects order from field_order param
        assert fs_date.sort_order == 0
        assert fs_customer.sort_order == 1

        # visible toggled via POST for non-system field
        assert fs_customer.visible is True
        assert fs_customer.allow_filter is True

        # system field: allow_filter NOT in POST → False
        assert fs_date.allow_filter is False

    def test_system_field_visible_unchanged_after_bulk_update(self, client, editor_user, fs_records):
        """System field visible=True is not overwritten by bulk_update (value is re-written as-is)."""
        from core.models import FieldSettings
        # Set system field visible=True in DB explicitly
        FieldSettings.objects.filter(
            entity="auto_shipment", field_name="shipment_date"
        ).update(visible=True)

        client.force_login(editor_user)
        client.post(
            "/settings/fields/",
            {
                "entity": "auto_shipment",
                "field_order": "shipment_date,customer_object",
                "visible_customer_object": "on",
                "required_customer_object": "on",
                "show_in_list_customer_object": "on",
            },
        )

        fs_date = FieldSettings.objects.get(entity="auto_shipment", field_name="shipment_date")
        # System field: loop does NOT assign fs.visible, so Python attribute == DB value == True
        # bulk_update writes it back unchanged → still True
        assert fs_date.visible is True


# ── Tests for ShipmentExportBaseMixin (V14-L3) ───────────────────────────────

class TestShipmentExportBaseMixin:
    """Unit tests for the extracted ShipmentExportBaseMixin."""

    def _make_request(self, query_string=""):
        from django.test import RequestFactory
        factory = RequestFactory()
        url = f"/fake/{f'?{query_string}' if query_string else ''}"
        return factory.get(url)

    def _make_mixin(self, list_url_name):
        from core.shipment_views import ShipmentExportBaseMixin

        class ConcreteMixin(ShipmentExportBaseMixin):
            pass

        obj = ConcreteMixin()
        obj.list_url_name = list_url_name
        return obj

    def test_rejection_redirect_url_without_query_string(self):
        mixin = self._make_mixin("auto:list")
        request = self._make_request()
        url = mixin._rejection_redirect_url(request)
        assert url == reverse("auto:list")
        assert "?" not in url

    def test_rejection_redirect_url_with_query_string(self):
        mixin = self._make_mixin("auto:list")
        request = self._make_request("q=test&sort=customer_object")
        url = mixin._rejection_redirect_url(request)
        assert url == reverse("auto:list") + "?q=test&sort=customer_object"

    def test_rejection_redirect_url_with_complex_query_string(self):
        mixin = self._make_mixin("auto:list")
        request = self._make_request("date_from=2026-01-01&sort=customer_object&dir=asc")
        url = mixin._rejection_redirect_url(request)
        assert url == reverse("auto:list") + "?date_from=2026-01-01&sort=customer_object&dir=asc"

    def test_export_selected_mixin_inherits_base(self):
        from core.shipment_views import ShipmentExportBaseMixin, ShipmentExportSelectedMixin
        assert issubclass(ShipmentExportSelectedMixin, ShipmentExportBaseMixin)

    def test_export_mixin_inherits_base(self):
        from core.shipment_views import ShipmentExportBaseMixin, ShipmentExportMixin
        assert issubclass(ShipmentExportMixin, ShipmentExportBaseMixin)

    def test_export_selected_mixin_has_no_own_list_url_name(self):
        """list_url_name must not be redefined on ShipmentExportSelectedMixin itself."""
        from core.shipment_views import ShipmentExportSelectedMixin
        assert "list_url_name" not in ShipmentExportSelectedMixin.__dict__

    def test_export_mixin_has_no_own_list_url_name(self):
        """list_url_name must not be redefined on ShipmentExportMixin itself."""
        from core.shipment_views import ShipmentExportMixin
        assert "list_url_name" not in ShipmentExportMixin.__dict__

    def test_export_selected_mixin_has_no_own_rejection_redirect_url(self):
        """_rejection_redirect_url must not be redefined on ShipmentExportSelectedMixin."""
        from core.shipment_views import ShipmentExportSelectedMixin
        assert "_rejection_redirect_url" not in ShipmentExportSelectedMixin.__dict__

    def test_export_mixin_has_no_own_rejection_redirect_url(self):
        """_rejection_redirect_url must not be redefined on ShipmentExportMixin."""
        from core.shipment_views import ShipmentExportMixin
        assert "_rejection_redirect_url" not in ShipmentExportMixin.__dict__


@pytest.mark.django_db
class TestShipmentExportBaseMixinIntegration:
    """Integration: POST to export_selected with rejection error preserves query string."""

    @pytest.fixture
    def exporter_user(self, django_user_model):
        user = django_user_model.objects.create_user(username="base_mixin_exporter", password="pass")
        user.user_permissions.add(
            Permission.objects.get(codename="view_autoshipment", content_type__app_label="shipments_auto"),
            Permission.objects.get(codename="export_excel", content_type__app_label="shipments_auto"),
        )
        return user

    def test_rejection_preserves_query_string(self, client, exporter_user):
        """POST with invalid ids → redirect to list URL with query string intact."""
        client.login(username="base_mixin_exporter", password="pass")
        query = "?q=test&sort=customer_object"
        response = client.post(
            reverse("auto:export_selected") + query,
            {"ids": ["abc"]},
        )
        assert response.status_code == 302
        assert response["Location"] == reverse("auto:list") + query

    def test_rejection_without_query_string(self, client, exporter_user):
        """POST with invalid ids and no GET params → redirect without '?'."""
        client.login(username="base_mixin_exporter", password="pass")
        response = client.post(
            reverse("auto:export_selected"),
            {"ids": ["abc"]},
        )
        assert response.status_code == 302
        assert response["Location"] == reverse("auto:list")
        assert "?" not in response["Location"]
