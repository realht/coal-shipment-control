import io
import os
import datetime
import zipfile
from pathlib import Path
import pytest
import openpyxl
from django.core.management import call_command
from django.test import Client
from django.urls import reverse
from django.contrib.auth.models import Group, Permission

from catalogs.models import AutoBase, AutoCoalGrade, RailCoalGrade
from imports.models import ImportLog, ImportRowResult


@pytest.fixture
def client():
    return Client()


@pytest.fixture
def admin_user(django_user_model):
    user = django_user_model.objects.create_user(username="import_admin", password="pass", is_staff=True)
    group, _ = Group.objects.get_or_create(name="import_admin_perms")
    for codename in ("import_shipments", "view_importlog"):
        group.permissions.add(Permission.objects.get(content_type__app_label="imports", codename=codename))
    user.groups.add(group)
    return user


@pytest.fixture
def viewer_user(django_user_model):
    user = django_user_model.objects.create_user(username="import_viewer", password="pass")
    group, _ = Group.objects.get_or_create(name="viewer")
    user.groups.add(group)
    return user


def _latest_token(client):
    # V12-17: задания импорта в session изолированы по токену вкладки.
    jobs = client.session.get("import_jobs") or {}
    return next(reversed(jobs), None)


def _latest_path(client):
    token = _latest_token(client)
    if not token:
        return None
    return client.session["import_jobs"][token]["tmp_path"]


def _set_import_job(client, token="tok-test", **fields):
    """Записать job напрямую в session (для тестов с ручным состоянием)."""
    job = {"tmp_path": None, "type": "auto", "filename": "", "year": None}
    job.update(fields)
    session = client.session
    jobs = session.get("import_jobs") or {}
    jobs[token] = job
    session["import_jobs"] = jobs
    session.save()
    return token


def _post_preview(client, data=None):
    payload = dict(data or {})
    payload.setdefault("t", _latest_token(client) or "")
    return client.post(reverse("imports:preview"), payload)


def _get_preview(client, token=None):
    return client.get(reverse("imports:preview"), {"t": token or _latest_token(client) or ""})


def _make_auto_excel(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ам"
    ws.append(["месяц", "число", "объект", "подобъект", "машины", "водители", "номер ттн", "марка", "кол-во", "база", "упд", "перевозчик", "остатки"])
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    buf.name = "auto_test.xlsx"
    return buf


def _make_auto_excel_with_prefix(prefix_rows, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ам"
    for row in prefix_rows:
        ws.append(row)
    ws.append(["месяц", "число", "объект", "марка", "кол-во"])
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    buf.name = "auto_prefixed.xlsx"
    return buf


def _make_rail_excel(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "жд новый"
    ws.append(["Дата отправления", "Номер вагона", "Номер документа", "Груз", "Область отправления", "Станция отправления РФ", "Грузоотправитель", "Область назначения", "Станция назначения РФ", "Грузополучатель", "Объем"])
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    buf.name = "rail_test.xlsx"
    return buf


def _make_minimal_xlsx_zip(extra_files):
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        zf.writestr("[Content_Types].xml", "<Types></Types>")
        zf.writestr("xl/workbook.xml", "<workbook></workbook>")
        for name, content in extra_files.items():
            zf.writestr(name, content)
    buf.seek(0)
    buf.name = "security.xlsx"
    return buf


def test_serialize_row_handles_non_json_excel_scalars():
    from decimal import Decimal
    from imports.views import _serialize_row

    row = {
        "date": datetime.date(2025, 1, 2),
        "datetime": datetime.datetime(2025, 1, 2, 3, 4, 5),
        "time": datetime.time(7, 8, 9),
        "duration": datetime.timedelta(hours=1, minutes=2, seconds=3),
        "decimal": Decimal("12.30"),
        "none": None,
        "text": "plain",
        "number": 5,
    }

    assert _serialize_row(row) == {
        "date": "2025-01-02",
        "datetime": "2025-01-02T03:04:05",
        "time": "07:08:09",
        "duration": "1:02:03",
        "decimal": "12.30",
        "none": None,
        "text": "plain",
        "number": 5,
    }


def _make_excel_without_known_headers(sheet_title):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.append(["служебная строка", "без известных заголовков"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    buf.name = "missing_headers.xlsx"
    return buf


class _CloseTrackingWorkbook:
    def __init__(self, workbook):
        self._workbook = workbook
        self.closed = False

    def __getattr__(self, name):
        return getattr(self._workbook, name)

    def __getitem__(self, key):
        return self._workbook[key]

    def close(self):
        self.closed = True
        return self._workbook.close()


@pytest.mark.django_db
class TestCatalogs:
    def test_auto_base_created(self):
        AutoBase.objects.create(name="Озеры")
        assert AutoBase.objects.filter(name="Озеры").exists()

    def test_auto_coal_grade_created(self):
        AutoCoalGrade.objects.create(name="ДПК")
        assert AutoCoalGrade.objects.filter(name="ДПК").exists()

    def test_rail_coal_grade_created(self):
        RailCoalGrade.objects.create(name="Уголь Д")
        assert RailCoalGrade.objects.filter(name="Уголь Д").exists()

    def test_inactive_filtered(self):
        AutoBase.objects.create(name="Старая база", is_active=False)
        assert not AutoBase.objects.filter(name="Старая база", is_active=True).exists()

    def test_str_representation(self):
        obj = AutoBase(name="Тучково")
        assert str(obj) == "Тучково"


@pytest.mark.django_db
class TestImportAccessControl:
    def test_anonymous_redirects(self, client):
        r = client.get(reverse("imports:index"))
        assert r.status_code == 302

    def test_viewer_gets_403(self, client, viewer_user):
        client.login(username="import_viewer", password="pass")
        r = client.get(reverse("imports:index"))
        assert r.status_code == 403

    def test_staff_without_permissions_gets_403(self, client, django_user_model):
        staff = django_user_model.objects.create_user(
            username="import_staff", password="pass", is_staff=True
        )
        client.login(username="import_staff", password="pass")
        r = client.get(reverse("imports:index"))
        assert r.status_code == 403

    def test_admin_gets_200(self, client, admin_user):
        client.login(username="import_admin", password="pass")
        r = client.get(reverse("imports:index"))
        assert r.status_code == 200

    def test_upload_page_accessible(self, client, admin_user):
        client.login(username="import_admin", password="pass")
        r = client.get(reverse("imports:upload") + "?type=auto")
        assert r.status_code == 200

    def test_upload_rail_page_accessible(self, client, admin_user):
        client.login(username="import_admin", password="pass")
        r = client.get(reverse("imports:upload") + "?type=rail")
        assert r.status_code == 200

    def test_log_page_accessible(self, client, admin_user):
        client.login(username="import_admin", password="pass")
        r = client.get(reverse("imports:log"))
        assert r.status_code == 200

    def test_preview_without_session_redirects(self, client, admin_user):
        client.login(username="import_admin", password="pass")
        r = client.get(reverse("imports:preview"))
        assert r.status_code == 302


@pytest.mark.django_db
class TestExcelParser:
    def test_parse_auto_valid(self):
        from imports.excel_parser import parse_auto_excel
        buf = _make_auto_excel([
            ["январь", 5, "ООО Тест", None, "А123БВ", "Иванов", "1001", "ДПК", 15, "т", "", "Озеры", None],
            ["январь", 6, "ООО Другой", None, None, None, None, "Т", 20, "п", "", "самовывоз", None],
        ])
        rows, errors = parse_auto_excel(buf, 2025)
        assert not errors
        assert len(rows) == 2
        valid = [r for r in rows if not r["errors"]]
        assert len(valid) == 2
        assert valid[0]["data"]["shipment_date"] == datetime.date(2025, 1, 5)
        assert valid[0]["data"]["coal_grade"] == "ДПК"
        assert valid[0]["data"]["quantity"] == 15.0

    def test_parse_auto_missing_date(self):
        from imports.excel_parser import parse_auto_excel
        buf = _make_auto_excel([
            [None, None, "ООО Тест", None, None, None, None, "ДПК", 15, None, None, None, None],
        ])
        rows, errors = parse_auto_excel(buf, 2025)
        assert len(rows) == 1
        assert rows[0]["errors"]

    def test_parse_auto_zero_quantity_is_not_empty_required_field(self):
        from imports.excel_parser import parse_auto_excel
        buf = _make_auto_excel([
            ["январь", 5, "ООО Тест", None, None, None, None, "ДПК", 0, None, None, None, None],
        ])
        rows, errors = parse_auto_excel(buf, 2025)
        assert not errors
        assert rows[0]["data"]["quantity"] == 0.0
        # quantity=0 теперь вызывает ошибку «больше нуля», а не «пустое поле»
        assert rows[0]["errors"]
        assert any("Количество" in e for e in rows[0]["errors"])
        assert not any("пустое" in e for e in rows[0]["errors"])

    def test_parse_auto_unknown_month(self):
        from imports.excel_parser import parse_auto_excel
        buf = _make_auto_excel([
            ["julye", 5, "ООО Тест", None, None, None, None, "ДПК", 15, None, None, None, None],
        ])
        rows, errors = parse_auto_excel(buf, 2025)
        assert rows[0]["errors"]

    def test_parse_auto_non_numeric_day_returns_row_error(self):
        from imports.excel_parser import parse_auto_excel
        buf = _make_auto_excel([
            ["январь", "пятое", "ООО Тест", None, None, None, None, "ДПК", 15, None, None, None, None],
        ])
        rows, errors = parse_auto_excel(buf, 2025)
        assert not errors
        assert len(rows) == 1
        assert rows[0]["row_num"] == 2
        assert rows[0]["data"]["source_day_number"] is None
        assert any("некорректное число" in err for err in rows[0]["errors"])

    def test_parse_auto_uses_actual_row_numbers_when_header_is_offset(self):
        from imports.excel_parser import parse_auto_excel
        buf = _make_auto_excel_with_prefix(
            [["служебная строка"], ["ещё одна строка"]],
            [["март", 7, "ООО Тест", "ДПК", 15]],
        )
        rows, errors = parse_auto_excel(buf, 2025)
        assert not errors
        assert len(rows) == 1
        assert rows[0]["row_num"] == 4
        assert not rows[0]["errors"]

    def test_parse_auto_row_cap_ignores_empty_sparse_rows(self):
        from imports.excel_parser import parse_auto_excel

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ам"
        ws.append(["месяц", "число", "объект", "марка", "кол-во"])
        ws.cell(row=10002, column=1, value="январь")
        ws.cell(row=10002, column=2, value=5)
        ws.cell(row=10002, column=3, value="ООО Тест")
        ws.cell(row=10002, column=4, value="ДПК")
        ws.cell(row=10002, column=5, value=15)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = "sparse_auto.xlsx"

        rows, errors = parse_auto_excel(buf, 2025)

        assert errors == []
        assert len(rows) == 1
        assert rows[0]["row_num"] == 10002
        assert rows[0]["errors"] == []

    def test_parse_auto_closes_workbook_when_headers_are_missing(self):
        from unittest.mock import patch
        from imports import excel_parser

        buf = _make_excel_without_known_headers("ам")
        original_load_workbook = excel_parser.openpyxl.load_workbook
        tracked = None

        def load_tracking_workbook(*args, **kwargs):
            nonlocal tracked
            tracked = _CloseTrackingWorkbook(original_load_workbook(*args, **kwargs))
            return tracked

        with patch("imports.excel_parser.openpyxl.load_workbook", side_effect=load_tracking_workbook):
            rows, errors = excel_parser.parse_auto_excel(buf, 2025)

        assert rows == []
        assert errors == ["Не удалось определить заголовки столбцов."]
        assert tracked is not None
        assert tracked.closed is True

    def test_parse_rail_valid(self):
        from imports.excel_parser import parse_rail_excel
        buf = _make_rail_excel([
            [datetime.datetime(2025, 3, 1), 65579682, "ЭС519244", "УГОЛЬ Д", "Респ. Хакасия", "ЧЕРНОГ.КОПИ", "АО РАЗРЕЗ", "Московская", "ОЗЕРЫ", "ООО Озерское", 68],
        ])
        rows, errors = parse_rail_excel(buf)
        assert not errors
        assert len(rows) == 1
        assert not rows[0]["errors"]
        assert rows[0]["data"]["departure_date"] == datetime.date(2025, 3, 1)
        assert rows[0]["data"]["volume"] == 68.0

    def test_parse_rail_missing_wagon(self):
        from imports.excel_parser import parse_rail_excel
        buf = _make_rail_excel([
            [datetime.datetime(2025, 3, 1), None, "", "УГОЛЬ Д", "", "", "", "", "", "ООО Озерское", 68],
        ])
        rows, errors = parse_rail_excel(buf)
        assert rows[0]["errors"]

    def test_parse_rail_zero_volume_is_not_empty_required_field(self):
        from imports.excel_parser import parse_rail_excel
        buf = _make_rail_excel([
            [datetime.datetime(2025, 3, 1), 65579682, "ЭС519244", "УГОЛЬ Д", "", "", "", "", "ОЗЕРЫ", "ООО Озерское", 0],
        ])
        rows, errors = parse_rail_excel(buf)
        assert not errors
        assert rows[0]["data"]["volume"] == 0.0
        # volume=0 теперь вызывает ошибку «больше нуля», а не «пустое поле»
        assert rows[0]["errors"]
        assert any("Объём" in e for e in rows[0]["errors"])
        assert not any("пустое" in e for e in rows[0]["errors"])

    def test_parse_rail_row_cap_ignores_empty_sparse_rows(self):
        from imports.excel_parser import parse_rail_excel

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "жд новый"
        ws.append([
            "Дата отправления", "Номер вагона", "Номер документа", "Груз",
            "Область отправления", "Станция отправления РФ", "Грузоотправитель",
            "Область назначения", "Станция назначения РФ", "Грузополучатель", "Объем",
        ])
        ws.cell(row=10002, column=1, value=datetime.datetime(2025, 3, 1))
        ws.cell(row=10002, column=2, value="65579682")
        ws.cell(row=10002, column=4, value="УГОЛЬ Д")
        ws.cell(row=10002, column=10, value="ООО Озерское")
        ws.cell(row=10002, column=11, value=68)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = "sparse_rail.xlsx"

        rows, errors = parse_rail_excel(buf)

        assert errors == []
        assert len(rows) == 1
        assert rows[0]["row_num"] == 10002
        assert rows[0]["errors"] == []

    def test_parse_rail_closes_workbook_when_headers_are_missing(self):
        from unittest.mock import patch
        from imports import excel_parser

        buf = _make_excel_without_known_headers("жд новый")
        original_load_workbook = excel_parser.openpyxl.load_workbook
        tracked = None

        def load_tracking_workbook(*args, **kwargs):
            nonlocal tracked
            tracked = _CloseTrackingWorkbook(original_load_workbook(*args, **kwargs))
            return tracked

        with patch("imports.excel_parser.openpyxl.load_workbook", side_effect=load_tracking_workbook):
            rows, errors = excel_parser.parse_rail_excel(buf)

        assert rows == []
        assert errors == ["Не удалось определить заголовки столбцов."]
        assert tracked is not None
        assert tracked.closed is True

    def test_parse_auto_fractional_quantity(self):
        from decimal import Decimal
        from imports.excel_parser import parse_auto_excel
        buf = _make_auto_excel([
            ["январь", 5, "ООО Тест", None, "А123БВ", "Иванов", "1001", "ДПК", 1.3, "т", "", "Озеры", None],
        ])
        rows, errors = parse_auto_excel(buf, 2025)
        assert not errors
        assert rows[0]["data"]["quantity"] == Decimal("1.3")
        assert isinstance(rows[0]["data"]["quantity"], Decimal)

    def test_parse_rail_fractional_volume(self):
        from decimal import Decimal
        from imports.excel_parser import parse_rail_excel
        buf = _make_rail_excel([
            [datetime.datetime(2025, 3, 1), 65579682, "ЭС519244", "УГОЛЬ Д",
             "Респ. Хакасия", "ЧЕРНОГ.КОПИ", "АО РАЗРЕЗ",
             "Московская", "ОЗЕРЫ", "ООО Озерское", 68.5],
        ])
        rows, errors = parse_rail_excel(buf)
        assert not errors
        assert rows[0]["data"]["volume"] == Decimal("68.5")
        assert isinstance(rows[0]["data"]["volume"], Decimal)

    def test_parse_rail_uses_actual_row_numbers_when_header_is_offset(self):
        from imports.excel_parser import parse_rail_excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "жд новый"
        ws.append(["служебная строка"])
        ws.append(["Дата отправления", "Номер вагона", "Груз", "Грузополучатель", "Объем"])
        ws.append([datetime.datetime(2025, 3, 1), 65579682, "УГОЛЬ Д", "ООО Озерское", 68])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = "rail_prefixed.xlsx"
        rows, errors = parse_rail_excel(buf)
        assert not errors
        assert rows[0]["row_num"] == 3
        assert not rows[0]["errors"]

    def test_detect_duplicates_auto(self):
        from imports.excel_parser import detect_duplicates
        from shipments_auto.models import AutoShipment
        from django.contrib.auth import get_user_model
        User = get_user_model()
        user = User.objects.create_user(username="dup_test_auto", password="x")
        AutoShipment.objects.create(
            shipment_date=datetime.date(2025, 1, 5),
            customer_object="ООО Тест",
            coal_grade="ДПК",
            quantity=15,
            created_by=user,
            updated_by=user,
        )
        rows = [{
            "row_num": 2,
            "data": {"shipment_date": datetime.date(2025, 1, 5), "customer_object": "ООО Тест", "coal_grade": "ДПК", "quantity": 15.0},
            "errors": [],
        }]
        result = detect_duplicates(rows, "auto")
        assert result[0]["is_duplicate"] is True

    def test_detect_duplicates_auto_chunks_large_lookup(self):
        from django.db import connection
        from django.test.utils import CaptureQueriesContext
        from imports.excel_parser import detect_duplicates

        rows = [
            {
                "row_num": i + 2,
                "data": {
                    "shipment_date": datetime.date(2025, 1, (i % 28) + 1),
                    "customer_object": f"ООО Массовый {i}",
                    "coal_grade": "ДПК",
                    "quantity": i + 1,
                },
                "errors": [],
            }
            for i in range(205)
        ]

        with CaptureQueriesContext(connection) as captured:
            result = detect_duplicates(rows, "auto")

        auto_selects = [
            query["sql"]
            for query in captured.captured_queries
            if "select" in query["sql"].lower() and "auto_shipments" in query["sql"].lower()
        ]
        assert len(auto_selects) == 3
        assert all(row["is_duplicate"] is False for row in result)

    def test_detect_duplicates_rail(self):
        from imports.excel_parser import detect_duplicates
        from shipments_rail.models import RailShipment
        from django.contrib.auth import get_user_model
        User = get_user_model()
        user = User.objects.create_user(username="dup_test_rail", password="x")
        RailShipment.objects.create(
            departure_date=datetime.date(2025, 3, 1),
            wagon_number="65579682",
            receiver="ООО Озерское",
            volume=68,
            cargo="УГОЛЬ Д",
            created_by=user,
            updated_by=user,
        )
        rows = [{
            "row_num": 2,
            "data": {"departure_date": datetime.date(2025, 3, 1), "wagon_number": "65579682", "receiver": "ООО Озерское", "volume": 68.0},
            "errors": [],
        }]
        result = detect_duplicates(rows, "rail")
        assert result[0]["is_duplicate"] is True

    def test_detect_duplicates_rail_chunks_large_lookup(self):
        from django.db import connection
        from django.test.utils import CaptureQueriesContext
        from imports.excel_parser import detect_duplicates

        rows = [
            {
                "row_num": i + 2,
                "data": {
                    "departure_date": datetime.date(2025, 3, (i % 28) + 1),
                    "wagon_number": f"{65000000 + i}",
                    "receiver": f"ООО Получатель {i}",
                    "volume": i + 1,
                },
                "errors": [],
            }
            for i in range(205)
        ]

        with CaptureQueriesContext(connection) as captured:
            result = detect_duplicates(rows, "rail")

        rail_selects = [
            query["sql"]
            for query in captured.captured_queries
            if "select" in query["sql"].lower() and "rail_shipments" in query["sql"].lower()
        ]
        assert len(rail_selects) == 3
        assert all(row["is_duplicate"] is False for row in result)

    def test_no_duplicate_when_different(self):
        from imports.excel_parser import detect_duplicates
        rows = [{
            "row_num": 2,
            "data": {"shipment_date": datetime.date(2025, 6, 15), "customer_object": "Нет в базе", "coal_grade": "ДПК", "quantity": 99.0},
            "errors": [],
        }]
        result = detect_duplicates(rows, "auto")
        assert result[0]["is_duplicate"] is False

    def test_normalize_dup_key(self):
        # V12-06: строковые компоненты нормализуются (strip + casefold),
        # не-строки (date/Decimal/None) проходят без изменений.
        from decimal import Decimal
        from imports.excel_parser import _normalize_dup_key
        assert _normalize_dup_key(("Восток ", "ДПК")) == ("восток", "дпк")
        d = datetime.date(2025, 1, 5)
        q = Decimal("15")
        assert _normalize_dup_key((d, "  X  ", q, None)) == (d, "x", q, None)

    def test_mark_duplicate_rows_normalizes_ci_collation_mismatch(self):
        # V12-06: при CI-коллации MariaDB SQL возвращает запись с иным
        # регистром/хвостовыми пробелами, чем импортный ключ. Нормализация
        # обеих сторон должна схлопнуть их в один ключ.
        # На SQLite (BINARY) сам SQL не воспроизводит CI-матч → БД-выборка
        # эмулируется фейком; полный прогон против MariaDB — задача V12-18.
        from decimal import Decimal
        from imports.excel_parser import _mark_duplicate_rows

        class _FakeQS:
            def __init__(self, rows):
                self._rows = rows

            def filter(self, *a, **k):
                return self

            def order_by(self, *a, **k):
                return self

            def values_list(self, *a, **k):
                return self._rows

        class _FakeManager:
            def __init__(self, rows):
                self._rows = rows

            def filter(self, *a, **k):
                return _FakeQS(self._rows)

        fields = ("shipment_date", "customer_object", "coal_grade", "quantity")
        # БД хранит "Восток" (заглавная, без пробела); CI-SQL вернёт эту строку
        db_row = (42, datetime.date(2025, 1, 5), "Восток", "ДПК", Decimal("15"))

        class _FakeModel:
            objects = _FakeManager([db_row])

        # импортный ключ отличается регистром и хвостовым пробелом
        import_key = (datetime.date(2025, 1, 5), "восток ", "ДПК", Decimal("15"))
        row = {"is_duplicate": False, "duplicate_ids": []}
        _mark_duplicate_rows(_FakeModel, [(import_key, row)], fields)

        assert row["is_duplicate"] is True
        assert row["duplicate_ids"] == [42]

    def test_auto_parser_required_field_error_uses_label(self):
        """Ошибка об обязательном поле должна содержать метку, а не внутреннее имя поля."""
        # Строка без объекта (customer_object пустой)
        file = _make_auto_excel([["январь", "15", "", "", "", "", "", "ДР", "100", "", "", "", ""]])
        from imports.excel_parser import parse_auto_excel
        rows, _ = parse_auto_excel(file, 2024)
        assert rows
        errors = rows[0]["errors"]
        assert any("Объект" in e for e in errors), f"Должно быть 'Объект' в ошибке: {errors}"
        assert not any("customer_object" in e for e in errors), (
            f"Не должно быть 'customer_object' в ошибке: {errors}"
        )

    def test_rail_parser_required_field_error_uses_label(self):
        """Rail-парсер: ошибка обязательного поля — метка вместо имени поля."""
        # Строка без грузополучателя (receiver пустой)
        file = _make_rail_excel([["2024-01-15", "11223344", "", "ДР", "", "", "", "", "", "", "100"]])
        from imports.excel_parser import parse_rail_excel
        rows, _ = parse_rail_excel(file)
        assert rows
        errors = rows[0]["errors"]
        assert any("Грузополучатель" in e for e in errors), f"Должно быть 'Грузополучатель' в ошибке: {errors}"
        assert not any("receiver" in e for e in errors), (
            f"Не должно быть 'receiver' в ошибке: {errors}"
        )

    def test_excel_parser_uses_defusedxml_elementtree(self):
        """V15-M5: excel_parser использует defusedxml.ElementTree, не xml.etree."""
        from imports import excel_parser
        import defusedxml.ElementTree as safe_et

        assert excel_parser.ElementTree is safe_et


@pytest.mark.django_db
class TestImportFlow:
    def test_upload_auto_creates_session(self, client, admin_user):
        client.login(username="import_admin", password="pass")
        buf = _make_auto_excel([
            ["январь", 5, "ООО Тест", None, None, None, None, "ДПК", 15, None, None, None, None],
        ])
        response = client.post(reverse("imports:upload"), {
            "shipment_type": "auto",
            "year": "2025",
            "excel_file": buf,
        })
        assert response.status_code == 302
        assert "preview" in response["Location"]
        token = _latest_token(client)
        assert token and client.session["import_jobs"][token]["tmp_path"]

    def test_upload_rail_creates_session(self, client, admin_user):
        client.login(username="import_admin", password="pass")
        buf = _make_rail_excel([
            [datetime.datetime(2025, 3, 1), 65579682, "ЭС519244", "УГОЛЬ Д", "", "", "", "", "ОЗЕРЫ", "ООО Озерское", 68],
        ])
        response = client.post(reverse("imports:upload"), {
            "shipment_type": "rail",
            "excel_file": buf,
        })
        assert response.status_code == 302
        token = _latest_token(client)
        assert token and client.session["import_jobs"][token]["tmp_path"]

    def test_upload_no_file_shows_error(self, client, admin_user):
        client.login(username="import_admin", password="pass")
        response = client.post(reverse("imports:upload"), {"shipment_type": "auto", "year": "2025"})
        assert response.status_code == 200

    def test_upload_auto_missing_year_shows_error(self, client, admin_user):
        client.login(username="import_admin", password="pass")
        buf = _make_auto_excel([["январь", 5, "ООО Тест", None, None, None, None, "ДПК", 15, None, None, None, None]])
        response = client.post(reverse("imports:upload"), {"shipment_type": "auto", "excel_file": buf})
        assert response.status_code == 200

    def test_import_rail_creates_log(self, client, admin_user):
        client.login(username="import_admin", password="pass")
        buf = _make_rail_excel([
            [datetime.datetime(2025, 4, 1), 11111111, "ЭС000001", "УГОЛЬ Т", "", "", "", "", "ТЕСТ", "ООО Тест", 50],
        ])
        client.post(reverse("imports:upload"), {"shipment_type": "rail", "excel_file": buf})
        response = _post_preview(client, {"selection_submitted": "1", "import_ids": ["2"]})
        assert response.status_code == 302
        log = ImportLog.objects.last()
        assert log is not None
        assert log.shipment_type == "rail"
        assert log.created_by == admin_user
        assert log.updated_rows == 0

    def test_import_creates_row_result_for_created_record(self, client, admin_user):
        client.login(username="import_admin", password="pass")
        buf = _make_rail_excel([
            [datetime.datetime(2025, 4, 1), 11111111, "ЭС000001", "УГОЛЬ Т", "", "", "", "", "ТЕСТ", "ООО Тест", 50],
        ])
        client.post(reverse("imports:upload"), {"shipment_type": "rail", "excel_file": buf})
        _post_preview(client, {"selection_submitted": "1", "import_ids": ["2"]})
        log = ImportLog.objects.last()
        row_result = log.row_results.get(row_num=2)
        assert row_result.status == ImportRowResult.STATUS_CREATED
        assert row_result.created_object_id is not None
        assert row_result.created_object_label
        assert row_result.source_data["wagon_number"] == "11111111"

    def test_import_auto_handles_more_than_200_valid_rows(self, client, admin_user):
        client.login(username="import_admin", password="pass")
        rows = [
            ["январь", 1, f"ООО Массовый {i}", None, f"А{i:03d}АА", "Иванов", f"ТТН-{i}", "ДПК", 10 + i, "т", "", "самовывоз", None]
            for i in range(1, 221)
        ]
        buf = _make_auto_excel(rows)

        upload_response = client.post(reverse("imports:upload"), {
            "shipment_type": "auto",
            "year": "2025",
            "excel_file": buf,
        })
        assert upload_response.status_code == 302
        assert "preview" in upload_response["Location"]

        import_ids = [str(i) for i in range(2, 222)]
        import_response = _post_preview(client, {
            "selection_submitted": "1",
            "import_ids": import_ids,
        })
        assert import_response.status_code == 302

        log = ImportLog.objects.last()
        assert log.total_rows == 220
        assert log.imported_rows == 220
        assert log.updated_rows == 0
        assert log.skipped_rows == 0
        assert log.duplicate_rows == 0
        assert log.error_rows == 0
        assert log.status == ImportLog.STATUS_SUCCESS
        assert log.row_results.filter(status=ImportRowResult.STATUS_CREATED).count() == 220

    def test_import_persists_error_duplicate_and_skipped_row_results(self, client, admin_user):
        from shipments_auto.models import AutoShipment

        AutoShipment.objects.create(
            shipment_date=datetime.date(2025, 1, 6),
            customer_object="ООО Дубль",
            coal_grade="ДПК",
            quantity=20,
            created_by=admin_user,
            updated_by=admin_user,
        )

        client.login(username="import_admin", password="pass")
        buf = _make_auto_excel([
            ["январь", 5, "ООО Создать", None, None, None, None, "ДПК", 15, None, None, None, None],
            ["январь", 6, "ООО Дубль", None, None, None, None, "ДПК", 20, None, None, None, None],
            [None, None, "ООО Ошибка", None, None, None, None, "ДПК", 10, None, None, None, None],
            ["январь", 7, "ООО Пропустить", None, None, None, None, "ДПК", 30, None, None, None, None],
        ])
        client.post(reverse("imports:upload"), {"shipment_type": "auto", "year": "2025", "excel_file": buf})
        _post_preview(client, {"selection_submitted": "1", "import_ids": ["2"]})

        log = ImportLog.objects.last()
        assert log.imported_rows == 1
        assert log.updated_rows == 0
        assert log.skipped_rows == 1
        assert log.duplicate_rows == 1
        assert log.error_rows == 1
        assert log.status == ImportLog.STATUS_PARTIAL

        statuses = {row.row_num: row.status for row in log.row_results.all()}
        assert statuses == {
            2: ImportRowResult.STATUS_CREATED,
            3: ImportRowResult.STATUS_DUPLICATE,
            4: ImportRowResult.STATUS_ERROR,
            5: ImportRowResult.STATUS_SKIPPED,
        }
        assert log.row_results.get(row_num=4).messages
        assert log.row_results.get(row_num=5).messages == ["Строка снята пользователем в предпросмотре."]

    def test_import_auto_creates_records(self, client, admin_user):
        client.login(username="import_admin", password="pass")
        buf = _make_auto_excel([
            ["февраль", 10, "ООО Авто Импорт", None, "Б456ВГ", "Петров", "2001", "ДМС", 25, "т", "", "самовывоз", None],
        ])
        client.post(reverse("imports:upload"), {"shipment_type": "auto", "year": "2025", "excel_file": buf})
        _post_preview(client, {"selection_submitted": "1", "import_ids": ["2"]})
        from shipments_auto.models import AutoShipment
        assert AutoShipment.objects.filter(customer_object="ООО Авто Импорт").exists()

    def test_import_auto_clears_session_state(self, client, admin_user):
        client.login(username="import_admin", password="pass")
        buf = _make_auto_excel([
            ["февраль", 10, "ООО Авто Импорт", None, "Б456ВГ", "Петров", "2001", "ДМС", 25, "т", "", "самовывоз", None],
        ])
        client.post(reverse("imports:upload"), {"shipment_type": "auto", "year": "2025", "excel_file": buf})
        token = _latest_token(client)
        response = _post_preview(client, {"selection_submitted": "1", "import_ids": ["2"]})
        assert response.status_code == 302
        assert token not in (client.session.get("import_jobs") or {})

    def test_repeated_confirm_post_with_same_token_does_not_duplicate(self, client, admin_user):
        """V19-LOW-3: конкурентный повторный POST с тем же токеном (double-click/
        вторая вкладка), поступивший, пока первый запрос ещё обрабатывает строки,
        не должен создавать дубли отгрузок и ImportLog."""
        from unittest.mock import patch
        from imports import views
        from shipments_auto.models import AutoShipment

        client.login(username="import_admin", password="pass")
        buf = _make_auto_excel([
            ["февраль", 11, "ООО Повтор", None, "Б111ВГ", "Сидоров", "3001", "ДМС", 12, "т", "", "самовывоз", None],
        ])
        client.post(reverse("imports:upload"), {"shipment_type": "auto", "year": "2025", "excel_file": buf})
        token = _latest_token(client)

        original_create = views._create_imported_shipment
        calls = {"count": 0}

        def racing_create(data, shipment_type, request, filename):
            calls["count"] += 1
            if calls["count"] == 1:
                # Симулируем вторую вкладку/двойной клик: тот же токен, тот же
                # клиент, запрос приходит, пока первая обработка строк ещё идёт
                # (job до правки снимается только в finally после неё).
                _post_preview(client, {"t": token, "selection_submitted": "1", "import_ids": ["2"]})
            return original_create(data, shipment_type, request, filename)

        with patch("imports.views._create_imported_shipment", side_effect=racing_create):
            response = _post_preview(client, {"t": token, "selection_submitted": "1", "import_ids": ["2"]})

        assert response.status_code == 302
        assert AutoShipment.objects.filter(customer_object="ООО Повтор").count() == 1
        assert ImportLog.objects.count() == 1

    def test_import_token_cleared_from_session_before_processing_rows(self, client, admin_user):
        """V19-LOW-3: токен должен быть снят из session ДО вызова обработки строк,
        а не только в finally после неё."""
        from unittest.mock import patch
        from imports import views

        client.login(username="import_admin", password="pass")
        buf = _make_rail_excel([
            [datetime.datetime(2025, 5, 5), 22222226, "ЭС000010", "УГОЛЬ Д", "", "", "", "", "АУДИТ", "ООО Токен", 30],
        ])
        client.post(reverse("imports:upload"), {"shipment_type": "rail", "excel_file": buf})
        token = _latest_token(client)

        captured = {}

        def fake_process(log, rows, import_ids, shipment_type, request, filename):
            jobs = request.session.get(views.IMPORT_JOBS_KEY) or {}
            captured["token_present"] = token in jobs
            views._apply_import_counters(log, views._empty_import_counters())
            views._save_final_import_log(log)

        with patch("imports.views._process_import_rows", side_effect=fake_process):
            response = _post_preview(client, {"t": token, "selection_submitted": "1", "import_ids": ["2"]})

        assert response.status_code == 302
        assert captured["token_present"] is False

    def test_import_creates_audit_log(self, client, admin_user):
        from audit.models import AuditLog
        client.login(username="import_admin", password="pass")
        buf = _make_rail_excel([
            [datetime.datetime(2025, 5, 1), 22222222, "ЭС000002", "УГОЛЬ Д", "", "", "", "", "АУДИТ", "ООО Аудит", 30],
        ])
        client.post(reverse("imports:upload"), {"shipment_type": "rail", "excel_file": buf})
        _post_preview(client, {"selection_submitted": "1", "import_ids": ["2"]})
        from shipments_rail.models import RailShipment
        obj = RailShipment.objects.filter(wagon_number="22222222").first()
        if obj:
            assert AuditLog.objects.filter(entity_type=AuditLog.ENTITY_RAIL, entity_id=obj.pk, action=AuditLog.ACTION_CREATE).exists()

    def test_confirm_bulk_creates_row_results_and_audit_logs(self, client, admin_user):
        from audit.models import AuditLog
        from django.db import connection
        from django.test.utils import CaptureQueriesContext

        client.login(username="import_admin", password="pass")
        buf = _make_rail_excel([
            [datetime.datetime(2025, 5, 3), 22222224, "ЭС000008", "УГОЛЬ Д", "", "", "", "", "АУДИТ", "ООО Bulk 1", 30],
            [datetime.datetime(2025, 5, 4), 22222225, "ЭС000009", "УГОЛЬ Д", "", "", "", "", "АУДИТ", "ООО Bulk 2", 31],
        ])
        client.post(reverse("imports:upload"), {"shipment_type": "rail", "excel_file": buf})

        with CaptureQueriesContext(connection) as captured:
            response = _post_preview(
                client,
                {"selection_submitted": "1", "import_ids": ["2", "3"]},
            )

        assert response.status_code == 302
        queries = [query["sql"].lower() for query in captured.captured_queries]
        row_result_inserts = [
            query for query in queries
            if query.lstrip().startswith("insert") and "import_row_results" in query
        ]
        audit_inserts = [
            query for query in queries
            if query.lstrip().startswith("insert") and "audit_log" in query
        ]
        import_log_updates = [
            query for query in queries
            if query.lstrip().startswith("update") and "import_logs" in query
        ]

        assert len(row_result_inserts) == 1
        assert len(audit_inserts) == 1
        assert len(import_log_updates) == 1

        log = ImportLog.objects.last()
        assert log.imported_rows == 2
        assert log.updated_rows == 0
        assert log.skipped_rows == 0
        assert log.duplicate_rows == 0
        assert log.error_rows == 0
        assert log.status == ImportLog.STATUS_SUCCESS
        assert log.row_results.filter(status=ImportRowResult.STATUS_CREATED).count() == 2
        assert AuditLog.objects.filter(source=AuditLog.SOURCE_IMPORT).count() == 2

    def test_confirm_creates_import_log_before_first_shipment(self, client, admin_user):
        from unittest.mock import patch
        from imports import views

        original_create = views._create_imported_shipment

        def assert_log_exists_before_create(data, shipment_type, request, filename):
            log = ImportLog.objects.get()
            assert log.status == ImportLog.STATUS_ERROR
            assert log.total_rows == 1
            assert log.imported_rows == 0
            return original_create(data, shipment_type, request, filename)

        client.login(username="import_admin", password="pass")
        buf = _make_rail_excel([
            [datetime.datetime(2025, 5, 2), 22222223, "ЭС000007", "УГОЛЬ Д", "", "", "", "", "АУДИТ", "ООО Ранний лог", 30],
        ])
        client.post(reverse("imports:upload"), {"shipment_type": "rail", "excel_file": buf})

        with patch("imports.views._create_imported_shipment", side_effect=assert_log_exists_before_create):
            response = _post_preview(
                client,
                {"selection_submitted": "1", "import_ids": ["2"]},
            )

        assert response.status_code == 302
        log = ImportLog.objects.get()
        assert log.status == ImportLog.STATUS_SUCCESS
        assert log.imported_rows == 1

    def test_confirm_rolls_back_created_shipment_when_row_result_fails(self, client, admin_user):
        from unittest.mock import patch
        from audit.models import AuditLog
        from shipments_auto.models import AutoShipment

        original_bulk_create = ImportRowResult.objects.bulk_create
        failed_once = False

        def flaky_row_result_bulk_create(*args, **kwargs):
            nonlocal failed_once
            if not failed_once:
                failed_once = True
                raise RuntimeError("row result insert failed")
            return original_bulk_create(*args, **kwargs)

        client.login(username="import_admin", password="pass")
        buf = _make_auto_excel([
            ["февраль", 11, "ООО Откат строки", None, "Б457ВГ", "Петров", "2002", "ДМС", 25, "т", "", "самовывоз", None],
        ])
        client.post(reverse("imports:upload"), {"shipment_type": "auto", "year": "2025", "excel_file": buf})

        with patch.object(ImportRowResult.objects, "bulk_create", side_effect=flaky_row_result_bulk_create):
            response = _post_preview(
                client,
                {"selection_submitted": "1", "import_ids": ["2"]},
            )

        assert response.status_code == 302
        assert not AutoShipment.objects.filter(customer_object="ООО Откат строки").exists()
        assert not AuditLog.objects.filter(source=AuditLog.SOURCE_IMPORT, new_values={"filename": "auto_test.xlsx"}).exists()
        log = ImportLog.objects.get()
        assert log.imported_rows == 0
        assert log.error_rows == 1
        assert log.status == ImportLog.STATUS_ERROR
        assert log.row_results.get(row_num=2).status == ImportRowResult.STATUS_ERROR

    def test_import_empty_selection_redirects_to_preview(self, client, admin_user):
        client.login(username="import_admin", password="pass")
        buf = _make_rail_excel([
            [datetime.datetime(2025, 6, 1), 33333333, "ЭС000003", "УГОЛЬ", "", "", "", "", "ТЕСТ", "ООО Тест", 10],
        ])
        client.post(reverse("imports:upload"), {"shipment_type": "rail", "excel_file": buf})
        count_before = ImportLog.objects.count()
        response = _post_preview(client, {"selection_submitted": "1"})
        assert response.status_code == 302
        assert ImportLog.objects.count() == count_before

    def test_import_no_selection_submitted_redirects_to_upload(self, client, admin_user):
        client.login(username="import_admin", password="pass")
        buf = _make_rail_excel([
            [datetime.datetime(2025, 7, 1), 44444444, "ЭС000004", "УГОЛЬ", "", "", "", "", "ТЕСТ", "ООО Тест", 5],
        ])
        client.post(reverse("imports:upload"), {"shipment_type": "rail", "excel_file": buf})
        response = _post_preview(client, {})
        assert response.status_code == 302
        assert "upload" in response["Location"]

    def test_detect_duplicates_not_called_without_selection_submitted(self, client, admin_user):
        """detect_duplicates is NOT called when selection_submitted is absent."""
        from unittest.mock import patch
        client.login(username="import_admin", password="pass")
        buf = _make_rail_excel([
            [datetime.datetime(2025, 7, 1), 44444444, "ЭС000004", "УГОЛЬ", "", "", "", "", "ТЕСТ", "ООО Тест", 5],
        ])
        client.post(reverse("imports:upload"), {"shipment_type": "rail", "excel_file": buf})
        with patch("imports.views.detect_duplicates") as mock_dup:
            _post_preview(client, {})  # no selection_submitted
        mock_dup.assert_not_called()

    def test_result_page_accessible(self, client, admin_user):
        log = ImportLog.objects.create(
            shipment_type="rail",
            filename="test.xlsx",
            status=ImportLog.STATUS_SUCCESS,
            total_rows=1,
            imported_rows=1,
            created_by=admin_user,
        )
        client.login(username="import_admin", password="pass")
        r = client.get(reverse("imports:result", kwargs={"pk": log.pk}))
        assert r.status_code == 200

    def test_result_page_explains_when_row_details_were_cleaned(self, client, admin_user):
        log = ImportLog.objects.create(
            shipment_type="auto",
            filename="cleaned.xlsx",
            status=ImportLog.STATUS_SUCCESS,
            total_rows=1,
            imported_rows=1,
            created_by=admin_user,
        )
        client.login(username="import_admin", password="pass")

        response = client.get(reverse("imports:result", kwargs={"pk": log.pk}))

        assert response.status_code == 200
        assert "Детали строк очищены" in response.content.decode("utf-8")

    def test_result_page_shows_row_details_link(self, client, admin_user):
        client.login(username="import_admin", password="pass")
        buf = _make_auto_excel([
            ["февраль", 10, "ООО Детали", None, "Б456ВГ", "Петров", "2001", "ДМС", 25, "т", "", "самовывоз", None],
        ])
        client.post(reverse("imports:upload"), {"shipment_type": "auto", "year": "2025", "excel_file": buf})
        _post_preview(client, {"selection_submitted": "1", "import_ids": ["2"]})
        log = ImportLog.objects.last()

        response = client.get(reverse("imports:result", kwargs={"pk": log.pk}))
        content = response.content.decode()
        assert response.status_code == 200
        assert "ООО Детали" in content
        assert "Создано" in content
        assert f"/auto/{log.row_results.get(row_num=2).created_object_id}/" in content

    def test_log_page_shows_extended_counters_and_details_link(self, client, admin_user):
        ImportLog.objects.create(
            shipment_type="rail",
            filename="history.xlsx",
            status=ImportLog.STATUS_PARTIAL,
            total_rows=4,
            imported_rows=1,
            updated_rows=0,
            skipped_rows=1,
            duplicate_rows=1,
            error_rows=1,
            created_by=admin_user,
        )
        client.login(username="import_admin", password="pass")
        response = client.get(reverse("imports:log"))
        content = response.content.decode()
        assert response.status_code == 200
        assert "Создано" in content
        assert "Обновлено" in content
        assert "Пропущено" in content
        assert "Открыть" in content

    def test_upload_second_time_keeps_both_tabs_isolated(self, client, admin_user):
        # V12-17: вторая загрузка (другая вкладка) НЕ затирает первую — у каждой
        # свой токен и свой tmp-файл.
        client.login(username="import_admin", password="pass")
        buf1 = _make_auto_excel([
            ["январь", 5, "ООО Первый", None, None, None, None, "ДПК", 15, None, None, None, None],
        ])
        client.post(reverse("imports:upload"), {"shipment_type": "auto", "year": "2025", "excel_file": buf1})
        token1 = _latest_token(client)
        first_path = client.session["import_jobs"][token1]["tmp_path"]
        assert os.path.isfile(first_path)

        buf2 = _make_auto_excel([
            ["февраль", 10, "ООО Второй", None, None, None, None, "ДПК", 20, None, None, None, None],
        ])
        client.post(reverse("imports:upload"), {"shipment_type": "auto", "year": "2025", "excel_file": buf2})
        token2 = _latest_token(client)
        second_path = client.session["import_jobs"][token2]["tmp_path"]

        assert token1 != token2
        assert os.path.isfile(first_path)  # первая вкладка не затёрта
        assert os.path.isfile(second_path)

    def test_upload_prunes_oldest_job_beyond_cap(self, client, admin_user):
        # V12-17: число параллельных заданий в session ограничено
        # IMPORT_JOBS_MAX; самое старое выселяется вместе с tmp-файлом.
        from imports.views import IMPORT_JOBS_MAX

        client.login(username="import_admin", password="pass")
        tokens = []
        paths = []
        for i in range(IMPORT_JOBS_MAX + 1):
            buf = _make_auto_excel([
                ["январь", (i % 28) + 1, f"ООО {i}", None, None, None, None, "ДПК", 15, None, None, None, None],
            ])
            client.post(reverse("imports:upload"), {"shipment_type": "auto", "year": "2025", "excel_file": buf})
            token = _latest_token(client)
            tokens.append(token)
            paths.append(client.session["import_jobs"][token]["tmp_path"])

        jobs = client.session["import_jobs"]
        assert len(jobs) == IMPORT_JOBS_MAX
        assert tokens[0] not in jobs            # самый старый выселен
        assert not os.path.isfile(paths[0])     # его tmp-файл удалён
        assert tokens[-1] in jobs
        assert os.path.isfile(paths[-1])

    def test_two_tabs_do_not_clobber_each_other(self, client, admin_user):
        # V12-17: предпросмотр по токену вкладки A видит данные A, не B.
        client.login(username="import_admin", password="pass")
        buf_a = _make_auto_excel([
            ["январь", 5, "ООО Вкладка A", None, None, None, None, "ДПК", 15, None, None, None, None],
        ])
        client.post(reverse("imports:upload"), {"shipment_type": "auto", "year": "2025", "excel_file": buf_a})
        token_a = _latest_token(client)

        buf_b = _make_auto_excel([
            ["февраль", 10, "ООО Вкладка B", None, None, None, None, "ДПК", 20, None, None, None, None],
        ])
        client.post(reverse("imports:upload"), {"shipment_type": "auto", "year": "2025", "excel_file": buf_b})
        token_b = _latest_token(client)

        assert token_a != token_b
        resp_a = _get_preview(client, token_a)
        assert resp_a.status_code == 200
        assert resp_a.context["preview_rows"]["valid"][0]["data"]["customer_object"] == "ООО Вкладка A"

        resp_b = _get_preview(client, token_b)
        assert resp_b.status_code == 200
        assert resp_b.context["preview_rows"]["valid"][0]["data"]["customer_object"] == "ООО Вкладка B"

    def test_preview_exposes_json_data_without_server_rendering_all_valid_rows(self, client, admin_user, settings, tmp_path):
        from imports.views import IMPORT_PREVIEW_PAGE_SIZE, _write_import_tmp

        settings.IMPORT_TMP_DIR = str(tmp_path / "import-previews")
        client.login(username="import_admin", password="pass")
        rows = [
            {
                "row_num": row_num,
                "data": {
                    "shipment_date": "2025-01-05",
                    "customer_object": f"ООО Preview {row_num}",
                    "coal_grade": "ДПК",
                    "quantity": str(row_num),
                    "base_code": "",
                    "ttn_number": "",
                },
                "errors": [],
                "is_duplicate": False,
                "duplicate_ids": [],
            }
            for row_num in range(2, IMPORT_PREVIEW_PAGE_SIZE + 7)
        ]
        path = _write_import_tmp(rows)
        token = _set_import_job(client, tmp_path=path, type="auto", filename="preview.xlsx")

        try:
            response = _get_preview(client, token)
        finally:
            if os.path.exists(path):
                os.unlink(path)

        content = response.content.decode("utf-8")
        assert response.status_code == 200
        assert 'id="import-preview-data"' in content
        assert f'data-page-size="{IMPORT_PREVIEW_PAGE_SIZE}"' in content
        assert content.count('name="import_ids"') == 0
        assert response.context["preview_rows"]["valid"][-1]["data"]["customer_object"] == (
            f"ООО Preview {IMPORT_PREVIEW_PAGE_SIZE + 6}"
        )

    def test_preview_post_can_import_row_outside_first_preview_page(self, client, admin_user, settings, tmp_path):
        from imports.views import IMPORT_PREVIEW_PAGE_SIZE

        settings.IMPORT_TMP_DIR = str(tmp_path / "import-previews")
        client.login(username="import_admin", password="pass")
        rows = [
            ["январь", (i % 28) + 1, f"ООО Hidden {i}", None, "", "", "", "ДПК", 10 + i, "", "", "", ""]
            for i in range(IMPORT_PREVIEW_PAGE_SIZE + 1)
        ]
        upload_response = client.post(reverse("imports:upload"), {
            "shipment_type": "auto",
            "year": "2025",
            "excel_file": _make_auto_excel(rows),
        })
        assert upload_response.status_code == 302
        hidden_row_num = IMPORT_PREVIEW_PAGE_SIZE + 2

        response = _post_preview(client, {
            "selection_submitted": "1",
            "import_ids": [str(hidden_row_num)],
        })

        assert response.status_code == 302
        log = ImportLog.objects.last()
        assert log.imported_rows == 1
        assert log.skipped_rows == IMPORT_PREVIEW_PAGE_SIZE
        assert log.status == ImportLog.STATUS_PARTIAL

    def test_write_import_tmp_uses_managed_tmp_dir(self, settings, tmp_path):
        from imports.views import _write_import_tmp

        managed_dir = tmp_path / "import-previews"
        settings.IMPORT_TMP_DIR = str(managed_dir)

        path = _write_import_tmp([])

        try:
            assert Path(path).parent == managed_dir
            assert Path(path).name.startswith("import_")
            assert Path(path).suffix == ".json"
            assert Path(path).is_file()
        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_read_import_tmp_rejects_path_outside_managed_dir(self, settings, tmp_path):
        # V12-17: session — недоверенный канал; _read_import_tmp не должен
        # читать файлы вне управляемой import-tmp-зоны, даже если они существуют.
        from imports.views import _read_import_tmp

        managed_dir = tmp_path / "import-previews"
        managed_dir.mkdir()
        settings.IMPORT_TMP_DIR = str(managed_dir)

        outside = tmp_path / "secret.json"
        outside.write_text('["leak"]', encoding="utf-8")

        assert _read_import_tmp(str(outside)) is None

    def test_read_import_tmp_reads_managed_file(self, settings, tmp_path):
        from imports.views import _read_import_tmp, _write_import_tmp

        managed_dir = tmp_path / "import-previews"
        settings.IMPORT_TMP_DIR = str(managed_dir)

        row = {"row_num": 2, "data": {}, "errors": [], "is_duplicate": False}
        path = _write_import_tmp([row])
        try:
            assert _read_import_tmp(path) == [row]
        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_read_import_tmp_rejects_missing_required_keys(self, settings, tmp_path):
        """V15-M4: синтаксически валидный JSON без обязательных ключей → None."""
        from imports.views import _read_import_tmp

        managed = tmp_path / "import-previews"
        managed.mkdir()
        settings.IMPORT_TMP_DIR = str(managed)

        bad = managed / "import_bad_schema.json"
        bad.write_text('[{"row_num": 2}]', encoding="utf-8")  # нет data/errors/is_duplicate

        assert _read_import_tmp(str(bad)) is None

    def test_read_import_tmp_accepts_valid_schema(self, settings, tmp_path):
        """V15-M4: строка с обязательными ключами возвращается без ошибок."""
        from imports.views import _read_import_tmp

        managed = tmp_path / "import-previews"
        managed.mkdir()
        settings.IMPORT_TMP_DIR = str(managed)

        good = managed / "import_good_schema.json"
        good.write_text(
            '[{"row_num": 2, "data": {}, "errors": [], "is_duplicate": false}]',
            encoding="utf-8",
        )

        rows = _read_import_tmp(str(good))
        assert rows is not None
        assert rows[0]["row_num"] == 2

    def test_upload_cleans_expired_tmp_files_only(self, client, admin_user, settings, tmp_path):
        managed_dir = tmp_path / "import-previews"
        managed_dir.mkdir()
        settings.IMPORT_TMP_DIR = str(managed_dir)
        settings.IMPORT_TMP_TTL_HOURS = 24

        expired = managed_dir / "import_expired.json"
        fresh = managed_dir / "import_fresh.json"
        unrelated = managed_dir / "other.json"
        for path in (expired, fresh, unrelated):
            path.write_text("[]", encoding="utf-8")

        now = datetime.datetime.now().timestamp()
        os.utime(expired, (now - 25 * 3600, now - 25 * 3600))
        os.utime(fresh, (now, now))
        os.utime(unrelated, (now - 25 * 3600, now - 25 * 3600))

        client.login(username="import_admin", password="pass")
        buf = _make_auto_excel([
            ["январь", 5, "ООО Новый", None, None, None, None, "ДПК", 15, None, None, None, None],
        ])
        response = client.post(reverse("imports:upload"), {
            "shipment_type": "auto",
            "year": "2025",
            "excel_file": buf,
        })

        assert response.status_code == 302
        assert not expired.exists()
        assert fresh.exists()
        assert unrelated.exists()
        token = _latest_token(client)
        assert Path(client.session["import_jobs"][token]["tmp_path"]).parent == managed_dir

    def test_cleanup_import_tmp_deletes_only_expired_import_json(self, settings, tmp_path):
        from imports.views import _cleanup_import_tmp

        managed_dir = tmp_path / "import-previews"
        managed_dir.mkdir()
        settings.IMPORT_TMP_DIR = str(managed_dir)

        expired = managed_dir / "import_expired.json"
        fresh = managed_dir / "import_fresh.json"
        unrelated = managed_dir / "import_notes.txt"
        for path in (expired, fresh, unrelated):
            path.write_text("[]", encoding="utf-8")

        now = datetime.datetime.now()
        old_ts = (now - datetime.timedelta(hours=2)).timestamp()
        os.utime(expired, (old_ts, old_ts))
        os.utime(unrelated, (old_ts, old_ts))

        result = _cleanup_import_tmp(ttl_hours=1, now=now)

        assert result == {"scanned": 2, "deleted": 1}
        assert not expired.exists()
        assert fresh.exists()
        assert unrelated.exists()

    def test_cleanup_import_tmp_command_dry_run_does_not_delete(self, settings, tmp_path):
        managed_dir = tmp_path / "import-previews"
        managed_dir.mkdir()
        settings.IMPORT_TMP_DIR = str(managed_dir)

        expired = managed_dir / "import_expired.json"
        expired.write_text("[]", encoding="utf-8")
        old_ts = (datetime.datetime.now() - datetime.timedelta(hours=2)).timestamp()
        os.utime(expired, (old_ts, old_ts))

        call_command("cleanup_import_tmp", "--older-than-hours", "1", "--dry-run")

        assert expired.exists()

    def test_cleanup_import_tmp_command_removes_expired_files(self, settings, tmp_path):
        managed_dir = tmp_path / "import-previews"
        managed_dir.mkdir()
        settings.IMPORT_TMP_DIR = str(managed_dir)

        expired = managed_dir / "import_expired.json"
        fresh = managed_dir / "import_fresh.json"
        expired.write_text("[]", encoding="utf-8")
        fresh.write_text("[]", encoding="utf-8")
        old_ts = (datetime.datetime.now() - datetime.timedelta(hours=2)).timestamp()
        os.utime(expired, (old_ts, old_ts))

        call_command("cleanup_import_tmp", "--older-than-hours", "1")

        assert not expired.exists()
        assert fresh.exists()

    def test_preview_with_missing_tmp_file_redirects(self, client, admin_user):
        client.login(username="import_admin", password="pass")
        buf = _make_auto_excel([
            ["январь", 5, "ООО Тест", None, None, None, None, "ДПК", 15, None, None, None, None],
        ])
        client.post(reverse("imports:upload"), {"shipment_type": "auto", "year": "2025", "excel_file": buf})
        path = _latest_path(client)
        os.unlink(path)

        response = _get_preview(client)
        assert response.status_code == 302
        assert "upload" in response["Location"]

    def test_tmp_file_deleted_after_successful_import(self, client, admin_user):
        client.login(username="import_admin", password="pass")
        buf = _make_rail_excel([
            [datetime.datetime(2025, 4, 1), 55555555, "ЭС000005", "УГОЛЬ Д", "", "", "", "", "ТЕСТ", "ООО Тест", 10],
        ])
        client.post(reverse("imports:upload"), {"shipment_type": "rail", "excel_file": buf})
        path = _latest_path(client)
        assert path and os.path.isfile(path)

        _post_preview(client, {"selection_submitted": "1", "import_ids": ["2"]})
        assert not os.path.isfile(path)

    def test_tmp_file_preserved_when_empty_selection(self, client, admin_user):
        client.login(username="import_admin", password="pass")
        buf = _make_rail_excel([
            [datetime.datetime(2025, 6, 1), 66666666, "ЭС000006", "УГОЛЬ", "", "", "", "", "ТЕСТ", "ООО Тест", 10],
        ])
        client.post(reverse("imports:upload"), {"shipment_type": "rail", "excel_file": buf})
        path = _latest_path(client)

        _post_preview(client, {"selection_submitted": "1"})
        assert os.path.isfile(path)

    def test_preview_get_with_invalid_session_type_returns_400_and_clears_session(self, client, admin_user):
        from imports.views import _write_import_tmp

        client.login(username="import_admin", password="pass")
        path = _write_import_tmp([])
        token = _set_import_job(client, tmp_path=path, type="bad", filename="bad.xlsx")

        try:
            response = _get_preview(client, token)
            assert response.status_code == 400
            assert token not in (client.session.get("import_jobs") or {})
        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_preview_post_with_invalid_session_type_returns_400_and_clears_session(self, client, admin_user):
        from imports.views import _write_import_tmp

        client.login(username="import_admin", password="pass")
        path = _write_import_tmp([])
        token = _set_import_job(client, tmp_path=path, type="bad", filename="bad.xlsx")

        try:
            response = _post_preview(client, {"selection_submitted": "1", "t": token})
            assert response.status_code == 400
            assert token not in (client.session.get("import_jobs") or {})
        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_preview_post_bad_serialized_date_records_row_error(self, client, admin_user):
        from imports.views import _write_import_tmp

        client.login(username="import_admin", password="pass")
        rows = [
            {
                "row_num": 2,
                "data": {
                    "shipment_date": "not-a-date",
                    "source_month_text": "январь",
                    "source_day_number": 5,
                    "customer_object": "ООО Ошибка даты",
                    "coal_grade": "ДПК",
                    "quantity": 15,
                },
                "errors": [],
                "is_duplicate": False,
            },
            {
                "row_num": 3,
                "data": {
                    "shipment_date": "2025-01-06",
                    "source_month_text": "январь",
                    "source_day_number": 6,
                    "customer_object": "ООО Валидная строка",
                    "coal_grade": "ДПК",
                    "quantity": 20,
                },
                "errors": [],
                "is_duplicate": False,
            },
        ]
        path = _write_import_tmp(rows)
        token = _set_import_job(client, tmp_path=path, type="auto", filename="bad_date.xlsx")

        response = _post_preview(
            client,
            {"selection_submitted": "1", "import_ids": ["2", "3"], "t": token},
        )

        assert response.status_code == 302
        log = ImportLog.objects.last()
        assert log.imported_rows == 1
        assert log.error_rows == 1
        assert log.status == ImportLog.STATUS_PARTIAL
        assert log.row_results.get(row_num=2).status == ImportRowResult.STATUS_ERROR
        assert log.row_results.get(row_num=3).status == ImportRowResult.STATUS_CREATED

    def test_confirm_rechecks_duplicates_toctou(self, client, admin_user):
        """Строка, добавленная между preview и confirm, должна быть поймана как дубль."""
        from imports.views import _write_import_tmp
        from shipments_auto.models import AutoShipment

        rows = [
            {
                "row_num": 2,
                "data": {
                    "shipment_date": "2025-07-01",
                    "customer_object": "ООО ТОКТОУ",
                    "coal_grade": "ДПК",
                    "quantity": 42.0,
                },
                "errors": [],
                "is_duplicate": False,  # при preview дубля не было
            },
        ]
        path = _write_import_tmp(rows)
        token = _set_import_job(client, tmp_path=path, type="auto", filename="toctou.xlsx")

        # Имитируем добавление записи между preview и confirm
        AutoShipment.objects.create(
            shipment_date=datetime.date(2025, 7, 1),
            customer_object="ООО ТОКТОУ",
            coal_grade="ДПК",
            quantity=42,
            created_by=admin_user,
            updated_by=admin_user,
        )

        client.login(username="import_admin", password="pass")
        response = _post_preview(
            client,
            {"selection_submitted": "1", "import_ids": ["2"], "t": token},
        )

        assert response.status_code == 302
        log = ImportLog.objects.last()
        assert log.imported_rows == 0
        assert log.duplicate_rows == 1
        assert log.row_results.get(row_num=2).status == ImportRowResult.STATUS_DUPLICATE
        assert not AutoShipment.objects.filter(
            shipment_date=datetime.date(2025, 7, 1),
            customer_object="ООО ТОКТОУ",
            coal_grade="ДПК",
            quantity=42,
        ).count() > 1


@pytest.mark.django_db
class TestImportFileValidation:
    def test_upload_get_invalid_shipment_type_returns_400(self, client, admin_user):
        client.login(username="import_admin", password="pass")
        response = client.get(reverse("imports:upload") + "?type=bad")
        assert response.status_code == 400

    def test_upload_post_invalid_shipment_type_returns_400_and_does_not_call_parser(
        self, client, admin_user
    ):
        from unittest.mock import patch

        client.login(username="import_admin", password="pass")
        buf = _make_auto_excel([
            ["январь", 5, "ООО Тест", None, None, None, None, "ДПК", 15, None, None, None, None],
        ])
        with (
            patch("imports.views.parse_auto_excel") as mock_parse_auto,
            patch("imports.views.parse_rail_excel") as mock_parse_rail,
        ):
            response = client.post(reverse("imports:upload"), {
                "shipment_type": "bad",
                "year": "2025",
                "excel_file": buf,
            })

        assert response.status_code == 400
        mock_parse_auto.assert_not_called()
        mock_parse_rail.assert_not_called()

    def test_valid_xlsx_accepted(self, client, admin_user):
        client.login(username="import_admin", password="pass")
        buf = _make_auto_excel([
            ["январь", 5, "ООО Тест", None, None, None, None, "ДПК", 15, None, None, None, None],
        ])
        response = client.post(reverse("imports:upload"), {
            "shipment_type": "auto",
            "year": "2025",
            "excel_file": buf,
        })
        assert response.status_code == 302
        assert "preview" in response["Location"]

    def test_xls_rejected(self, client, admin_user):
        client.login(username="import_admin", password="pass")
        buf = io.BytesIO(b"fake content")
        buf.name = "data.xls"
        response = client.post(reverse("imports:upload"), {
            "shipment_type": "auto",
            "year": "2025",
            "excel_file": buf,
        })
        assert response.status_code == 200
        assert "xlsx" in response.content.decode()

    def test_csv_rejected(self, client, admin_user):
        client.login(username="import_admin", password="pass")
        buf = io.BytesIO(b"col1,col2\nval1,val2")
        buf.name = "data.csv"
        response = client.post(reverse("imports:upload"), {
            "shipment_type": "auto",
            "year": "2025",
            "excel_file": buf,
        })
        assert response.status_code == 200
        assert "xlsx" in response.content.decode()

    def test_no_extension_rejected(self, client, admin_user):
        client.login(username="import_admin", password="pass")
        buf = io.BytesIO(b"some data")
        buf.name = "datafile"
        response = client.post(reverse("imports:upload"), {
            "shipment_type": "auto",
            "year": "2025",
            "excel_file": buf,
        })
        assert response.status_code == 200
        assert "xlsx" in response.content.decode()

    def test_oversized_file_rejected(self, client, admin_user, settings):
        settings.MAX_IMPORT_SIZE_MB = 1
        settings.MAX_IMPORT_SIZE_BYTES = 1 * 1024 * 1024
        client.login(username="import_admin", password="pass")
        buf = io.BytesIO(b"x" * (2 * 1024 * 1024))
        buf.name = "big.xlsx"
        response = client.post(reverse("imports:upload"), {
            "shipment_type": "auto",
            "year": "2025",
            "excel_file": buf,
        })
        assert response.status_code == 200
        assert "МБ" in response.content.decode()

    def test_invalid_file_does_not_call_parser(self, client, admin_user):
        from unittest.mock import patch
        client.login(username="import_admin", password="pass")
        buf = io.BytesIO(b"fake")
        buf.name = "data.csv"
        with patch("imports.views.parse_auto_excel") as mock_parse:
            client.post(reverse("imports:upload"), {
                "shipment_type": "auto",
                "year": "2025",
                "excel_file": buf,
            })
            mock_parse.assert_not_called()

    def test_corrupt_xlsx_rejected_before_openpyxl(self, client, admin_user):
        from unittest.mock import patch

        client.login(username="import_admin", password="pass")
        buf = io.BytesIO(b"not a zip")
        buf.name = "broken.xlsx"
        with patch("imports.excel_parser.openpyxl.load_workbook") as mock_load_workbook:
            response = client.post(reverse("imports:upload"), {
                "shipment_type": "auto",
                "year": "2025",
                "excel_file": buf,
            })

        assert response.status_code == 200
        assert "поврежд" in response.content.decode().lower()
        mock_load_workbook.assert_not_called()

    def test_xlsx_uncompressed_size_limit_rejected_before_openpyxl(
        self, client, admin_user, settings
    ):
        from unittest.mock import patch

        settings.MAX_IMPORT_UNCOMPRESSED_SIZE_MB = 1
        settings.MAX_IMPORT_UNCOMPRESSED_SIZE_BYTES = 10
        client.login(username="import_admin", password="pass")
        buf = _make_minimal_xlsx_zip({"xl/worksheets/sheet1.xml": "x" * 11})
        with patch("imports.excel_parser.openpyxl.load_workbook") as mock_load_workbook:
            response = client.post(reverse("imports:upload"), {
                "shipment_type": "auto",
                "year": "2025",
                "excel_file": buf,
            })

        assert response.status_code == 200
        assert "распакованный размер" in response.content.decode().lower()
        mock_load_workbook.assert_not_called()

    def test_xlsx_shared_strings_limit_rejected_before_openpyxl(
        self, client, admin_user, settings
    ):
        from unittest.mock import patch

        settings.MAX_IMPORT_SHARED_STRINGS = 1
        client.login(username="import_admin", password="pass")
        shared_strings = (
            '<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            "<si><t>one</t></si><si><t>two</t></si></sst>"
        )
        buf = _make_minimal_xlsx_zip({"xl/sharedStrings.xml": shared_strings})
        with patch("imports.excel_parser.openpyxl.load_workbook") as mock_load_workbook:
            response = client.post(reverse("imports:upload"), {
                "shipment_type": "auto",
                "year": "2025",
                "excel_file": buf,
            })

        assert response.status_code == 200
        assert "строковых значений" in response.content.decode().lower()
        mock_load_workbook.assert_not_called()


@pytest.mark.django_db
class TestResultPartialExplanation:
    def test_hidden_on_success(self, admin_user, client):
        log = ImportLog.objects.create(
            shipment_type="auto",
            filename="test.xlsx",
            status=ImportLog.STATUS_SUCCESS,
            total_rows=1,
            imported_rows=1,
            updated_rows=0,
            skipped_rows=0,
            error_rows=0,
            duplicate_rows=0,
            created_by=admin_user,
        )
        client.force_login(admin_user)
        response = client.get(reverse("imports:result", args=[log.pk]))
        assert response.status_code == 200
        assert "Импорт выполнен в частичном режиме" not in response.content.decode()

    def test_shown_on_partial(self, admin_user, client):
        log = ImportLog.objects.create(
            shipment_type="auto",
            filename="test.xlsx",
            status=ImportLog.STATUS_PARTIAL,
            total_rows=2,
            imported_rows=1,
            updated_rows=0,
            skipped_rows=1,
            error_rows=0,
            duplicate_rows=0,
            created_by=admin_user,
        )
        client.force_login(admin_user)
        response = client.get(reverse("imports:result", args=[log.pk]))
        assert response.status_code == 200
        assert "Импорт выполнен в частичном режиме" in response.content.decode()


def test_row_cap_error_does_not_imply_imported():
    """При >10000 строк сообщение должно говорить об ограничении, а не об успешном импорте."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ам"
    ws.append(["месяц", "число", "объект", "марка", "кол-во"])
    for i in range(10001):
        ws.append(["январь", str(1 + i % 28), f"Объект{i}", "ДР", "100"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    buf.name = "large.xlsx"
    from imports.excel_parser import parse_auto_excel
    _, errors = parse_auto_excel(buf, 2024)
    assert errors, "Должна быть parse_error при >10000 строк"
    assert "импортированы" not in errors[0].lower(), (
        f"Сообщение не должно говорить об импорте: {errors[0]}"
    )


def test_auto_parser_quantity_zero_gives_row_error():
    """quantity=0 в Excel даёт ошибку строки в парсере."""
    file = _make_auto_excel([["январь", "15", "ООО Тест", "", "", "", "", "ДР", "0", "", "", "", ""]])
    from imports.excel_parser import parse_auto_excel
    rows, parse_errors = parse_auto_excel(file, 2024)
    assert not parse_errors
    assert rows[0]["errors"], "Строка с quantity=0 должна иметь ошибку"
    assert any("Количество" in e for e in rows[0]["errors"])


def test_auto_parser_quantity_negative_gives_row_error():
    file = _make_auto_excel([["январь", "15", "ООО Тест", "", "", "", "", "ДР", "-5", "", "", "", ""]])
    from imports.excel_parser import parse_auto_excel
    rows, _ = parse_auto_excel(file, 2024)
    assert rows[0]["errors"]
    assert any("Количество" in e for e in rows[0]["errors"])


def test_rail_parser_volume_zero_gives_row_error():
    """volume=0 в Excel даёт ошибку строки в парсере."""
    file = _make_rail_excel([["2024-01-15", "11223344", "", "ДР", "", "", "", "", "", "Получатель", "0"]])
    from imports.excel_parser import parse_rail_excel
    rows, parse_errors = parse_rail_excel(file)
    assert not parse_errors
    assert rows[0]["errors"]
    assert any("Объём" in e for e in rows[0]["errors"])


def test_auto_parser_quantity_nan_gives_row_error():
    """V19-LOW-2: quantity=nan не должно валить весь файл, а даёт ошибку строки."""
    file = _make_auto_excel([
        ["январь", "15", "ООО Тест", "", "", "", "", "ДР", "nan", "", "", "", ""],
        ["январь", "16", "ООО Тест", "", "", "", "", "ДР", "10", "", "", "", ""],
    ])
    from imports.excel_parser import parse_auto_excel
    rows, parse_errors = parse_auto_excel(file, 2024)
    assert not parse_errors
    assert rows[0]["errors"], "Строка с quantity=nan должна иметь ошибку"
    assert not rows[1]["errors"], "Соседняя корректная строка должна остаться валидной"


def test_auto_parser_quantity_infinity_gives_row_error():
    """V19-LOW-2: quantity=inf не должно проходить в БД, а даёт ошибку строки."""
    file = _make_auto_excel([
        ["январь", "15", "ООО Тест", "", "", "", "", "ДР", "inf", "", "", "", ""],
        ["январь", "16", "ООО Тест", "", "", "", "", "ДР", "10", "", "", "", ""],
    ])
    from imports.excel_parser import parse_auto_excel
    rows, parse_errors = parse_auto_excel(file, 2024)
    assert not parse_errors
    assert rows[0]["errors"], "Строка с quantity=inf должна иметь ошибку"
    assert not rows[1]["errors"], "Соседняя корректная строка должна остаться валидной"


def test_auto_parser_quantity_negative_infinity_gives_row_error():
    """V19-LOW-2: quantity=-Infinity даёт ошибку строки."""
    file = _make_auto_excel([
        ["январь", "15", "ООО Тест", "", "", "", "", "ДР", "-Infinity", "", "", "", ""],
    ])
    from imports.excel_parser import parse_auto_excel
    rows, parse_errors = parse_auto_excel(file, 2024)
    assert not parse_errors
    assert rows[0]["errors"]


def test_rail_parser_volume_nan_gives_row_error():
    """V19-LOW-2: volume=nan не должно валить весь файл, а даёт ошибку строки."""
    file = _make_rail_excel([
        ["2024-01-15", "11223344", "", "ДР", "", "", "", "", "", "Получатель", "nan"],
        ["2024-01-16", "11223345", "", "ДР", "", "", "", "", "", "Получатель", "10"],
    ])
    from imports.excel_parser import parse_rail_excel
    rows, parse_errors = parse_rail_excel(file)
    assert not parse_errors
    assert rows[0]["errors"], "Строка с volume=nan должна иметь ошибку"
    assert not rows[1]["errors"], "Соседняя корректная строка должна остаться валидной"


def test_rail_parser_volume_infinity_gives_row_error():
    """V19-LOW-2: volume=Infinity не должно проходить в БД, а даёт ошибку строки."""
    file = _make_rail_excel([
        ["2024-01-15", "11223344", "", "ДР", "", "", "", "", "", "Получатель", "Infinity"],
    ])
    from imports.excel_parser import parse_rail_excel
    rows, parse_errors = parse_rail_excel(file)
    assert not parse_errors
    assert rows[0]["errors"]


@pytest.mark.django_db
class TestDuplicateIds:
    def test_detect_duplicates_stores_ids(self, django_user_model):
        """detect_duplicates должна заполнять duplicate_ids при нахождении дубля."""
        import datetime
        from decimal import Decimal
        from shipments_auto.models import AutoShipment
        from imports.excel_parser import detect_duplicates
        user = django_user_model.objects.create_user(username="dup_ids", password="pass")
        existing = AutoShipment.objects.create(
            shipment_date=datetime.date(2024, 1, 15),
            customer_object="Объект А",
            coal_grade="ДР",
            quantity=Decimal("100"),
            created_by=user,
            updated_by=user,
        )
        rows = [
            {
                "row_num": 2,
                "data": {
                    "shipment_date": "2024-01-15",
                    "customer_object": "Объект А",
                    "coal_grade": "ДР",
                    "quantity": "100",
                },
                "errors": [],
                "is_duplicate": False,
                "duplicate_ids": [],
            }
        ]
        result = detect_duplicates(rows, "auto")
        assert result[0]["is_duplicate"] is True
        assert existing.pk in result[0]["duplicate_ids"]

    def test_duplicate_message_includes_id(self, admin_user):
        """_process_import_row для дубля включает ID в сообщение."""
        import datetime
        from decimal import Decimal
        from shipments_auto.models import AutoShipment
        from imports.models import ImportLog, ImportRowResult
        from imports.views import _process_import_row, _empty_import_counters
        from unittest.mock import MagicMock
        existing = AutoShipment.objects.create(
            shipment_date=datetime.date(2024, 6, 1),
            customer_object="Объект Б",
            coal_grade="Г",
            quantity=Decimal("50"),
            created_by=admin_user,
            updated_by=admin_user,
        )
        log = ImportLog.objects.create(
            shipment_type="auto", filename="test.xlsx",
            status=ImportLog.STATUS_ERROR,
            total_rows=1, imported_rows=0, updated_rows=0,
            skipped_rows=0, error_rows=0, duplicate_rows=0,
            created_by=admin_user,
        )
        request = MagicMock()
        request.user = admin_user
        request.META = {"HTTP_USER_AGENT": "test"}
        batch = {
            "row_results": [],
            "audit_logs": [],
            "counters": _empty_import_counters(),
            "ip_address": "127.0.0.1",
            "user_agent": "test",
        }
        row = {
            "row_num": 2,
            "data": {},
            "errors": [],
            "is_duplicate": True,
            "duplicate_ids": [existing.pk],
        }
        _process_import_row(log, row, set(), "auto", request, "test.xlsx", batch)
        assert batch["row_results"]
        msg = batch["row_results"][0].messages[0]
        assert str(existing.pk) in msg, f"ID {existing.pk} должен быть в сообщении: {msg}"

