import datetime
import zipfile
from decimal import Decimal
import defusedxml.ElementTree as ElementTree

from django.conf import settings
import openpyxl

MONTHS_RU = {
    "январь": 1, "февраль": 2, "март": 3, "апрель": 4,
    "май": 5, "июнь": 6, "июль": 7, "август": 8,
    "сентябрь": 9, "октябрь": 10, "ноябрь": 11, "декабрь": 12,
}

AUTO_COLUMNS = {
    "месяц": "source_month_text",
    "число": "source_day_number",
    "объект": "customer_object",
    "подобъект": "sub_object",
    "машины": "vehicle_number",
    "водители": "driver_name",
    "номер ттн": "ttn_number",
    "марка": "coal_grade",
    "кол-во": "quantity",
    "база": "base_code",
    "упд": "upd_number",
    "перевозчик": "carrier",
    "остатки": "balance_note",
}

RAIL_COLUMNS = {
    "дата отправления": "departure_date",
    "номер вагона": "wagon_number",
    "номер документа": "document_number",
    "груз": "cargo",
    "область отправления": "origin_region",
    "станция отправления рф": "origin_station",
    "грузоотправитель": "sender",
    "область назначения": "destination_region",
    "станция назначения рф": "destination_station",
    "грузополучатель": "receiver",
    "объем": "volume",
    "объём": "volume",
}

AUTO_REQUIRED = ["customer_object", "coal_grade", "quantity"]
RAIL_REQUIRED = ["departure_date", "wagon_number", "receiver", "volume", "cargo"]
DUPLICATE_LOOKUP_CHUNK_SIZE = 100

AUTO_FIELD_LABEL = {
    "customer_object": "Объект",
    "coal_grade": "Марка угля",
    "quantity": "Количество",
}

RAIL_FIELD_LABEL = {
    "departure_date": "Дата отправления",
    "wagon_number": "Номер вагона",
    "receiver": "Грузополучатель",
    "volume": "Объём",
    "cargo": "Марка угля",
}


class UnsafeXlsxError(ValueError):
    pass


def _str(val):
    if val is None:
        return ""
    return str(val).strip()


def _detect_sheet(wb, shipment_type):
    if shipment_type == "auto":
        keywords = ["ам", "авт", "auto"]
    else:
        keywords = ["жд", "rail", "жд новый"]
    for name in wb.sheetnames:
        if any(kw in name.lower() for kw in keywords):
            return wb[name]
    return wb.active


def _map_headers(ws, column_map):
    header_row = None
    col_map = {}
    for row in ws.iter_rows(min_row=1, max_row=3):
        for cell in row:
            val = _str(cell.value).lower()
            if val in column_map:
                col_map[column_map[val]] = cell.column - 1
                header_row = cell.row
    return header_row, col_map


def _data_rows(ws, header_row):
    first_data_row = (header_row or 1) + 1
    for row_num, row in enumerate(
        ws.iter_rows(min_row=first_data_row, values_only=True),
        start=first_data_row,
    ):
        if all(v is None for v in row):
            continue
        yield row_num, row


def _parse_int(value):
    if value is None:
        return None
    try:
        return int(value)
    except (ValueError, TypeError):
        return None


def _parse_decimal(value):
    if value is None:
        return None
    try:
        d = Decimal(str(value))
    except Exception:
        return None
    if not d.is_finite():
        return None
    return d


def _rewind(file_obj):
    if hasattr(file_obj, "seek"):
        file_obj.seek(0)


def validate_xlsx_archive(file_obj):
    _rewind(file_obj)
    try:
        with zipfile.ZipFile(file_obj) as zf:
            infos = zf.infolist()
            names = {info.filename for info in infos}
            if "[Content_Types].xml" not in names or not any(name.startswith("xl/") for name in names):
                raise UnsafeXlsxError("Содержимое файла не соответствует формату .xlsx.")

            total_uncompressed = sum(info.file_size for info in infos)
            if total_uncompressed > settings.MAX_IMPORT_UNCOMPRESSED_SIZE_BYTES:
                raise UnsafeXlsxError(
                    "Распакованный размер файла превышает допустимый "
                    f"{settings.MAX_IMPORT_UNCOMPRESSED_SIZE_MB} МБ."
                )

            if "xl/sharedStrings.xml" in names:
                shared_strings = 0
                with zf.open("xl/sharedStrings.xml") as fh:
                    for _event, elem in ElementTree.iterparse(fh, events=("end",)):
                        if elem.tag.endswith("}si") or elem.tag == "si":
                            shared_strings += 1
                            if shared_strings > settings.MAX_IMPORT_SHARED_STRINGS:
                                raise UnsafeXlsxError(
                                    "Файл содержит слишком много строковых значений."
                                )
                            elem.clear()
    except zipfile.BadZipFile as exc:
        raise UnsafeXlsxError("Файл повреждён или не является корректным .xlsx.") from exc
    except ElementTree.ParseError as exc:
        raise UnsafeXlsxError("Файл .xlsx повреждён: не удалось прочитать строковые значения.") from exc
    finally:
        _rewind(file_obj)


def parse_auto_excel(file_obj, year):
    validate_xlsx_archive(file_obj)
    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    try:
        ws = _detect_sheet(wb, "auto")
        header_row, col_map = _map_headers(ws, AUTO_COLUMNS)
        if not col_map:
            return [], ["Не удалось определить заголовки столбцов."]

        rows = []
        parse_errors = []
        for data_idx, (i, row) in enumerate(_data_rows(ws, header_row), start=1):
            if data_idx > 10000:
                parse_errors.append("Файл содержит более 10 000 строк данных. Разбейте файл на части и загрузите повторно.")
                break
            if all(v is None for v in row):
                continue
            record = {}
            for field, col_idx in col_map.items():
                record[field] = row[col_idx] if col_idx < len(row) else None

            month_text = _str(record.get("source_month_text"))
            day_num = record.get("source_day_number")
            parsed_day_num = _parse_int(day_num)
            shipment_date = None
            date_error = None
            if month_text and day_num:
                month_num = MONTHS_RU.get(month_text.lower())
                if month_num and parsed_day_num is not None:
                    try:
                        shipment_date = datetime.date(year, month_num, parsed_day_num)
                    except (ValueError, TypeError):
                        date_error = f"Строка {i}: некорректная дата ({month_text} {day_num} {year})"
                elif month_num:
                    date_error = f"Строка {i}: некорректное число «{day_num}»"
                else:
                    date_error = f"Строка {i}: неизвестный месяц «{month_text}»"
            elif day_num:
                date_error = f"Строка {i}: нет месяца"

            record["shipment_date"] = shipment_date
            record["source_month_text"] = month_text
            record["source_day_number"] = parsed_day_num

            record["quantity"] = _parse_decimal(record.get("quantity"))

            record["coal_grade"] = _str(record.get("coal_grade"))
            record["customer_object"] = _str(record.get("customer_object"))

            errors = []
            if date_error:
                errors.append(date_error)
            if not record.get("shipment_date"):
                errors.append(f"Строка {i}: отсутствует дата отгрузки")
            for req in AUTO_REQUIRED:
                val = record.get(req)
                if val is None or val == "":
                    errors.append(f"Строка {i}: обязательное поле «{AUTO_FIELD_LABEL.get(req, req)}» пустое")
            qty = record.get("quantity")
            if qty is not None and qty <= Decimal("0"):
                errors.append(f"Строка {i}: поле «Количество» должно быть больше нуля")

            rows.append({"row_num": i, "data": record, "errors": errors})

        return rows, parse_errors
    finally:
        wb.close()


def parse_rail_excel(file_obj):
    validate_xlsx_archive(file_obj)
    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    try:
        ws = _detect_sheet(wb, "rail")
        header_row, col_map = _map_headers(ws, RAIL_COLUMNS)
        if not col_map:
            return [], ["Не удалось определить заголовки столбцов."]

        rows = []
        parse_errors = []
        for data_idx, (i, row) in enumerate(_data_rows(ws, header_row), start=1):
            if data_idx > 10000:
                parse_errors.append("Файл содержит более 10 000 строк данных. Разбейте файл на части и загрузите повторно.")
                break
            if all(v is None for v in row):
                continue
            record = {}
            for field, col_idx in col_map.items():
                record[field] = row[col_idx] if col_idx < len(row) else None

            dep_date = record.get("departure_date")
            if isinstance(dep_date, datetime.datetime):
                record["departure_date"] = dep_date.date()
            elif isinstance(dep_date, datetime.date):
                record["departure_date"] = dep_date
            elif dep_date:
                try:
                    record["departure_date"] = datetime.date.fromisoformat(str(dep_date)[:10])
                except ValueError:
                    record["departure_date"] = None

            record["volume"] = _parse_decimal(record.get("volume"))

            record["cargo"] = _str(record.get("cargo"))
            record["wagon_number"] = _str(record.get("wagon_number"))
            record["receiver"] = _str(record.get("receiver"))

            errors = []
            if not record.get("departure_date"):
                errors.append(f"Строка {i}: отсутствует или некорректная дата отправления")
            for req in RAIL_REQUIRED:
                val = record.get(req)
                if val is None or val == "":
                    errors.append(f"Строка {i}: обязательное поле «{RAIL_FIELD_LABEL.get(req, req)}» пустое")
            vol = record.get("volume")
            if vol is not None and vol <= Decimal("0"):
                errors.append(f"Строка {i}: поле «Объём» должно быть больше нуля")

            rows.append({"row_num": i, "data": record, "errors": errors})

        return rows, parse_errors
    finally:
        wb.close()


def _parse_date(v):
    """Нормализует строку ISO-даты или datetime.date к datetime.date. Бросает ValueError для невалидных значений."""
    import datetime
    if isinstance(v, datetime.date):
        return v
    return datetime.date.fromisoformat(v)


def _chunks(items, size):
    for start in range(0, len(items), size):
        yield items[start:start + size]


def _normalize_dup_key(key):
    # V12-06: MariaDB сравнивает строки по CI-коллации (регистронезависимо,
    # с обрезкой хвостовых пробелов). Чтобы Python-сопоставление совпадало с
    # SQL-выборкой, нормализуем строковые компоненты ключа с обеих сторон.
    # Не-строки (date/Decimal/None) остаются без изменений.
    return tuple(v.strip().casefold() if isinstance(v, str) else v for v in key)


def _find_existing_duplicate_keys(model, fields, keys):
    from functools import reduce
    import operator

    from django.db.models import Q

    unique_keys = list(dict.fromkeys(keys))
    existing = {}  # _normalize_dup_key(tuple) -> [pk, ...]
    for chunk in _chunks(unique_keys, DUPLICATE_LOOKUP_CHUNK_SIZE):
        conditions = [
            Q(**dict(zip(fields, key, strict=True)))
            for key in chunk
        ]
        for row_vals in (
            model.objects.filter(is_deleted=False)
            .filter(reduce(operator.or_, conditions))
            .order_by()
            .values_list("pk", *fields)
        ):
            pk = row_vals[0]
            key = _normalize_dup_key(row_vals[1:])
            existing.setdefault(key, []).append(pk)
    return existing


def _mark_duplicate_rows(model, keyed_rows, fields):
    existing = _find_existing_duplicate_keys(
        model,
        fields,
        [key for key, _row in keyed_rows],
    )
    for key, row in keyed_rows:
        ids = existing.get(_normalize_dup_key(key), [])
        row["is_duplicate"] = bool(ids)
        row["duplicate_ids"] = ids


def detect_duplicates(rows, shipment_type):
    from shipments_auto.models import AutoShipment
    from shipments_rail.models import RailShipment

    for row in rows:
        row["is_duplicate"] = False
        row["duplicate_ids"] = []

    valid_rows = [row for row in rows if not row["errors"]]
    if not valid_rows:
        return rows

    if shipment_type == "auto":
        keyed = []
        for row in valid_rows:
            d = row["data"]
            try:
                key = (
                    _parse_date(d.get("shipment_date")),
                    d.get("customer_object"),
                    d.get("coal_grade"),
                    _parse_decimal(d.get("quantity")),
                )
            except (ValueError, TypeError):
                continue
            keyed.append((key, row))
        if keyed:
            _mark_duplicate_rows(
                AutoShipment,
                keyed,
                ("shipment_date", "customer_object", "coal_grade", "quantity"),
            )
    else:
        keyed = []
        for row in valid_rows:
            d = row["data"]
            try:
                key = (
                    _parse_date(d.get("departure_date")),
                    d.get("wagon_number"),
                    d.get("receiver"),
                    _parse_decimal(d.get("volume")),
                )
            except (ValueError, TypeError):
                continue
            keyed.append((key, row))
        if keyed:
            _mark_duplicate_rows(
                RailShipment,
                keyed,
                ("departure_date", "wagon_number", "receiver", "volume"),
            )

    return rows
