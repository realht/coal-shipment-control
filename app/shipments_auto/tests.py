import logging
import pytest
from django.test import Client
from django.urls import reverse
from django.contrib.auth.models import Group, Permission
from urllib.parse import quote
from unittest.mock import patch

from .models import AutoShipment


def _xlsx_bytes(response):
    from django.http import StreamingHttpResponse
    if isinstance(response, StreamingHttpResponse):
        return b"".join(response.streaming_content)
    return response.content


class ListHandler(logging.Handler):
    def __init__(self, level=logging.NOTSET):
        super().__init__(level=level)
        self.records = []

    def emit(self, record):
        self.records.append(record)


@pytest.fixture
def client():
    return Client()


@pytest.fixture
def viewer_user(django_user_model):
    user = django_user_model.objects.create_user(username="viewer_auto", password="pass")
    group, _ = Group.objects.get_or_create(name="viewer")
    perm = Permission.objects.get(codename="view_autoshipment", content_type__app_label="shipments_auto")
    group.permissions.add(perm)
    user.groups.add(group)
    return user


@pytest.fixture
def operator_user(django_user_model):
    user = django_user_model.objects.create_user(username="operator_auto", password="pass")
    group, _ = Group.objects.get_or_create(name="operator")
    for codename in ("view_autoshipment", "add_autoshipment", "change_autoshipment", "delete_autoshipment"):
        perm = Permission.objects.get(codename=codename, content_type__app_label="shipments_auto")
        group.permissions.add(perm)
    user.groups.add(group)
    return user


@pytest.fixture
def shipment(operator_user):
    return AutoShipment.objects.create(
        shipment_date="2026-01-15",
        customer_object="Объект А",
        coal_grade="ДГ",
        quantity="1234.567",
        created_by=operator_user,
        updated_by=operator_user,
    )


@pytest.mark.django_db
class TestAutoShipmentList:
    def test_anonymous_redirects(self, client):
        response = client.get(reverse("auto:list"))
        assert response.status_code == 302
        assert "/accounts/login/" in response["Location"]

    def test_viewer_can_see_list(self, client, viewer_user):
        client.login(username="viewer_auto", password="pass")
        response = client.get(reverse("auto:list"))
        assert response.status_code == 200

    def test_list_table_presets_use_json_script(self, client, viewer_user):
        client.login(username="viewer_auto", password="pass")
        response = client.get(reverse("auto:list"))
        html = response.content.decode()
        assert 'data-presets-script-id="auto-table-presets"' in html
        assert 'id="auto-table-presets"' in html
        assert 'data-presets="' not in html

    def test_list_shows_shipment(self, client, viewer_user, shipment):
        client.login(username="viewer_auto", password="pass")
        response = client.get(reverse("auto:list"))
        assert "Объект А" in response.content.decode()

    def test_viewer_has_no_add_button(self, client, viewer_user):
        client.login(username="viewer_auto", password="pass")
        response = client.get(reverse("auto:list"))
        assert "Добавить автоотгрузку" not in response.content.decode()

    def test_operator_has_add_button(self, client, operator_user):
        client.login(username="operator_auto", password="pass")
        response = client.get(reverse("auto:list"))
        assert "/auto/new/" in response.content.decode()

    def test_search_filters_results(self, client, viewer_user, shipment):
        AutoShipment.objects.create(
            shipment_date="2026-01-16",
            customer_object="Объект Б",
            coal_grade="Т",
            quantity="100",
        )
        client.login(username="viewer_auto", password="pass")
        response = client.get(reverse("auto:list") + "?q=Объект+А")
        objects = [s.customer_object for s in response.context["shipments"]]
        assert "Объект А" in objects
        assert "Объект Б" not in objects

    def test_search_no_duplicate_when_multiple_fields_match(self, client, viewer_user):
        AutoShipment.objects.create(
            shipment_date="2026-01-15",
            customer_object="Уголь-объект",
            coal_grade="Уголь-марка",
            quantity="50",
        )
        client.login(username="viewer_auto", password="pass")
        response = client.get(reverse("auto:list") + "?q=Уголь")
        shipments = list(response.context["shipments"])
        ids = [s.pk for s in shipments]
        assert len(ids) == len(set(ids)), "Дубли в queryset при поиске по нескольким полям"

    def test_date_filter(self, client, viewer_user, shipment):
        AutoShipment.objects.create(
            shipment_date="2026-03-01",
            customer_object="Мартовский",
            coal_grade="Т",
            quantity="50",
        )
        client.login(username="viewer_auto", password="pass")
        response = client.get(reverse("auto:list") + "?date_from=2026-02-01&date_to=2026-03-31")
        objects = [s.customer_object for s in response.context["shipments"]]
        assert "Мартовский" in objects
        assert "Объект А" not in objects

    def test_invalid_top_level_date_filter_is_ignored_without_500(self, client, viewer_user, shipment):
        client.login(username="viewer_auto", password="pass")
        response = client.get(reverse("auto:list") + "?date_from=zzzz&date_to=abc")

        assert response.status_code == 200
        assert shipment in list(response.context["shipments"])

    def test_pagination(self, client, viewer_user):
        for i in range(30):
            AutoShipment.objects.create(
                shipment_date="2026-01-01",
                customer_object=f"Объект {i}",
                coal_grade="ДГ",
                quantity="10",
            )
        client.login(username="viewer_auto", password="pass")
        response = client.get(reverse("auto:list"))
        assert response.status_code == 200
        assert response.context["is_paginated"]
        assert len(response.context["shipments"]) == 25


@pytest.mark.django_db
class TestAutoShipmentDetail:
    def test_viewer_can_see_detail(self, client, viewer_user, shipment):
        client.login(username="viewer_auto", password="pass")
        response = client.get(reverse("auto:detail", kwargs={"pk": shipment.pk}))
        assert response.status_code == 200
        assert "Объект А" in response.content.decode()

    def test_viewer_has_no_edit_button(self, client, viewer_user, shipment):
        client.login(username="viewer_auto", password="pass")
        response = client.get(reverse("auto:detail", kwargs={"pk": shipment.pk}))
        assert "Редактировать" not in response.content.decode()

    def test_operator_has_edit_button(self, client, operator_user, shipment):
        client.login(username="operator_auto", password="pass")
        response = client.get(reverse("auto:detail", kwargs={"pk": shipment.pk}))
        assert "Редактировать" in response.content.decode()


@pytest.mark.django_db
class TestAutoShipmentCreate:
    def test_viewer_cannot_create(self, client, viewer_user):
        client.login(username="viewer_auto", password="pass")
        response = client.get(reverse("auto:create"))
        assert response.status_code == 403

    def test_operator_can_get_form(self, client, operator_user):
        client.login(username="operator_auto", password="pass")
        response = client.get(reverse("auto:create"))
        assert response.status_code == 200

    def test_operator_can_create(self, client, operator_user):
        client.login(username="operator_auto", password="pass")
        response = client.post(reverse("auto:create"), {
            "shipment_date": "2026-02-10",
            "customer_object": "Новый объект",
            "coal_grade_select": "__other__",
            "coal_grade_other": "ДГ",
            "quantity": "500.000",
            "vehicle_number": "",
            "driver_name": "",
            "ttn_number": "",
            "carrier": "",
            "comment": "",
            "sub_object": "",
            "base_code_select": "",
            "upd_number": "",
            "balance_note": "",
        })
        assert response.status_code == 302
        assert AutoShipment.objects.filter(customer_object="Новый объект").exists()

    def test_decimal_precision_preserved(self, client, operator_user):
        client.login(username="operator_auto", password="pass")
        client.post(reverse("auto:create"), {
            "shipment_date": "2026-02-11",
            "customer_object": "Дробный тест",
            "coal_grade_select": "__other__",
            "coal_grade_other": "Т",
            "quantity": "1234.567",
            "vehicle_number": "", "driver_name": "", "ttn_number": "",
            "carrier": "", "comment": "", "sub_object": "",
            "base_code_select": "", "upd_number": "", "balance_note": "",
        })
        obj = AutoShipment.objects.get(customer_object="Дробный тест")
        from decimal import Decimal
        assert obj.quantity == Decimal("1234.567")

    def test_create_writes_audit_log(self, client, operator_user):
        from audit.models import AuditLog
        client.login(username="operator_auto", password="pass")
        client.post(reverse("auto:create"), {
            "shipment_date": "2026-02-12",
            "customer_object": "Аудит объект",
            "coal_grade_select": "__other__",
            "coal_grade_other": "ДГ",
            "quantity": "10",
            "vehicle_number": "", "driver_name": "", "ttn_number": "",
            "carrier": "", "comment": "", "sub_object": "",
            "base_code_select": "", "upd_number": "", "balance_note": "",
        })
        obj = AutoShipment.objects.get(customer_object="Аудит объект")
        assert AuditLog.objects.filter(
            entity_type=AuditLog.ENTITY_AUTO,
            entity_id=obj.pk,
            action=AuditLog.ACTION_CREATE,
        ).exists()


@pytest.mark.django_db
class TestAutoShipmentUpdate:
    def test_viewer_cannot_update(self, client, viewer_user, shipment):
        client.login(username="viewer_auto", password="pass")
        response = client.get(reverse("auto:update", kwargs={"pk": shipment.pk}))
        assert response.status_code == 403

    def test_operator_can_update(self, client, operator_user, shipment):
        client.login(username="operator_auto", password="pass")
        response = client.post(reverse("auto:update", kwargs={"pk": shipment.pk}), {
            "shipment_date": "2026-01-15",
            "customer_object": "Изменённый объект",
            "coal_grade_select": "__other__",
            "coal_grade_other": "ДГ",
            "quantity": "999.000",
            "vehicle_number": "", "driver_name": "", "ttn_number": "",
            "carrier": "", "comment": "", "sub_object": "",
            "base_code_select": "", "upd_number": "", "balance_note": "",
            "updated_at_token": shipment.updated_at.isoformat(),
        })
        assert response.status_code == 302
        shipment.refresh_from_db()
        assert shipment.customer_object == "Изменённый объект"

    def test_update_writes_audit_log(self, client, operator_user, shipment):
        from audit.models import AuditLog
        client.login(username="operator_auto", password="pass")
        client.post(reverse("auto:update", kwargs={"pk": shipment.pk}), {
            "shipment_date": "2026-01-15",
            "customer_object": "После изменения",
            "coal_grade_select": "__other__",
            "coal_grade_other": "Т",
            "quantity": "100",
            "vehicle_number": "", "driver_name": "", "ttn_number": "",
            "carrier": "", "comment": "", "sub_object": "",
            "base_code_select": "", "upd_number": "", "balance_note": "",
            "updated_at_token": shipment.updated_at.isoformat(),
        })
        assert AuditLog.objects.filter(
            entity_type=AuditLog.ENTITY_AUTO,
            entity_id=shipment.pk,
            action=AuditLog.ACTION_UPDATE,
        ).exists()


@pytest.mark.django_db
class TestAutoShipmentOptimisticLock:
    _base_post = {
        "shipment_date": "2026-01-15",
        "customer_object": "Объект А",
        "coal_grade_select": "__other__",
        "coal_grade_other": "ДГ",
        "quantity": "999.000",
        "vehicle_number": "", "driver_name": "", "ttn_number": "",
        "carrier": "", "comment": "", "sub_object": "",
        "base_code_select": "", "upd_number": "", "balance_note": "",
    }

    def test_missing_token_rejected(self, client, operator_user, shipment):
        client.login(username="operator_auto", password="pass")
        response = client.post(
            reverse("auto:update", kwargs={"pk": shipment.pk}),
            {**self._base_post},
        )
        assert response.status_code == 200
        shipment.refresh_from_db()
        assert shipment.customer_object == "Объект А"

    def test_stale_token_rejected_with_who_and_when(self, client, operator_user, shipment):
        stale_token = shipment.updated_at.isoformat()
        # имитируем изменение другим пользователем
        other = type(operator_user).objects.create_user(username="other_auto", password="p")
        shipment.customer_object = "Чужое изменение"
        shipment.updated_by = other
        shipment.save()

        client.login(username="operator_auto", password="pass")
        response = client.post(
            reverse("auto:update", kwargs={"pk": shipment.pk}),
            {**self._base_post, "updated_at_token": stale_token},
        )
        assert response.status_code == 200
        content = response.content.decode()
        assert "other_auto" in content
        shipment.refresh_from_db()
        assert shipment.customer_object == "Чужое изменение"

    def test_fresh_token_after_conflict_saves_ok(self, client, operator_user, shipment):
        other = type(operator_user).objects.create_user(username="other_auto2", password="p")
        shipment.customer_object = "Промежуточное"
        shipment.updated_by = other
        shipment.save()

        client.login(username="operator_auto", password="pass")
        response = client.post(
            reverse("auto:update", kwargs={"pk": shipment.pk}),
            {**self._base_post, "customer_object": "Финальное", "updated_at_token": shipment.updated_at.isoformat()},
        )
        assert response.status_code == 302
        shipment.refresh_from_db()
        assert shipment.customer_object == "Финальное"

    def test_parallel_post_with_same_token_does_not_overwrite_first_save(self, client, operator_user, shipment):
        stale_snapshot = AutoShipment.objects.get(pk=shipment.pk)
        token = stale_snapshot.updated_at.isoformat()

        client.login(username="operator_auto", password="pass")
        first_response = client.post(
            reverse("auto:update", kwargs={"pk": shipment.pk}),
            {**self._base_post, "customer_object": "Первое изменение", "updated_at_token": token},
        )
        assert first_response.status_code == 302

        with patch("shipments_auto.views.AutoShipmentUpdateView.get_object", return_value=stale_snapshot):
            second_response = client.post(
                reverse("auto:update", kwargs={"pk": shipment.pk}),
                {**self._base_post, "customer_object": "Второе изменение", "updated_at_token": token},
            )

        assert second_response.status_code == 200
        assert "Ваши изменения не сохранены" in second_response.content.decode()
        shipment.refresh_from_db()
        assert shipment.customer_object == "Первое изменение"


@pytest.mark.django_db
class TestAutoShipmentSoftDelete:
    def test_viewer_cannot_delete(self, client, viewer_user, shipment):
        client.login(username="viewer_auto", password="pass")
        response = client.post(reverse("auto:delete", kwargs={"pk": shipment.pk}))
        assert response.status_code == 403

    def test_operator_can_soft_delete(self, client, operator_user, shipment):
        client.login(username="operator_auto", password="pass")
        response = client.post(reverse("auto:delete", kwargs={"pk": shipment.pk}))
        assert response.status_code == 302
        shipment.refresh_from_db()
        assert shipment.is_deleted is True

    def test_delete_redirect_preserves_list_filters(self, client, operator_user, shipment):
        client.login(username="operator_auto", password="pass")
        next_url = reverse("auto:list") + "?f_customer_object=%D0%9E%D0%B1%D1%8A%D0%B5%D0%BA%D1%82+%D0%90&sort=customer_object&dir=asc"
        response = client.post(
            reverse("auto:delete", kwargs={"pk": shipment.pk}) + f"?next={quote(next_url, safe='')}"
        )
        assert response.status_code == 302
        assert response["Location"] == next_url

    def test_delete_rejects_external_next_url(self, client, operator_user, shipment):
        client.login(username="operator_auto", password="pass")
        response = client.post(reverse("auto:delete", kwargs={"pk": shipment.pk}) + "?next=https%3A%2F%2Fevil.test%2F")
        assert response.status_code == 302
        assert response["Location"] == reverse("auto:list")

    def test_list_delete_link_includes_current_filters(self, client, operator_user, shipment):
        client.login(username="operator_auto", password="pass")
        response = client.get(reverse("auto:list") + "?q=%D0%9E%D0%B1%D1%8A%D0%B5%D0%BA%D1%82&sort=customer_object&dir=asc")
        content = response.content.decode()
        expected_next = quote("/auto/?q=%D0%9E%D0%B1%D1%8A%D0%B5%D0%BA%D1%82&sort=customer_object&dir=asc", safe="/")
        assert f'{reverse("auto:delete", kwargs={"pk": shipment.pk})}?next={expected_next}' in content

    def test_deleted_not_in_list(self, client, viewer_user, shipment):
        shipment.is_deleted = True
        shipment.save()
        client.login(username="viewer_auto", password="pass")
        response = client.get(reverse("auto:list"))
        assert "Объект А" not in response.content.decode()

    def test_soft_delete_writes_audit_log(self, client, operator_user, shipment):
        from audit.models import AuditLog
        client.login(username="operator_auto", password="pass")
        client.post(reverse("auto:delete", kwargs={"pk": shipment.pk}))
        assert AuditLog.objects.filter(
            entity_type=AuditLog.ENTITY_AUTO,
            entity_id=shipment.pk,
            action=AuditLog.ACTION_DELETE,
        ).exists()

    def test_hard_delete_not_called(self, client, operator_user, shipment):
        pk = shipment.pk
        client.login(username="operator_auto", password="pass")
        client.post(reverse("auto:delete", kwargs={"pk": pk}))
        assert AutoShipment.all_objects.filter(pk=pk).exists()

    def test_soft_delete_and_audit_log_atomic(self, client, operator_user, shipment):
        """V13-L6: soft-delete + AuditLog выполняются в одной транзакции."""
        from audit.models import AuditLog
        client.login(username="operator_auto", password="pass")
        client.post(reverse("auto:delete", kwargs={"pk": shipment.pk}))
        shipment.refresh_from_db()
        assert shipment.is_deleted is True
        assert AuditLog.objects.filter(
            action=AuditLog.ACTION_DELETE,
            entity_type=AuditLog.ENTITY_AUTO,
            entity_id=shipment.pk,
        ).exists()


@pytest.mark.django_db
class TestAutoShipmentExport:
    @pytest.fixture
    def exporter_user(self, django_user_model):
        user = django_user_model.objects.create_user(username="exporter_auto", password="pass")
        group, _ = Group.objects.get_or_create(name="exporter_auto_group")
        for codename in ("view_autoshipment", "export_excel"):
            perm = Permission.objects.get(codename=codename, content_type__app_label="shipments_auto")
            group.permissions.add(perm)
        user.groups.add(group)
        return user

    def test_viewer_without_perm_gets_403(self, client, viewer_user):
        client.login(username="viewer_auto", password="pass")
        response = client.get(reverse("auto:export"))
        assert response.status_code == 403

    def test_anonymous_redirects(self, client):
        response = client.get(reverse("auto:export"))
        assert response.status_code == 302
        assert "/accounts/login/" in response["Location"]

    def test_export_returns_xlsx(self, client, exporter_user, shipment):
        client.login(username="exporter_auto", password="pass")
        response = client.get(reverse("auto:export"))
        assert response.status_code == 200
        assert response["Content-Type"] == (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert "attachment" in response["Content-Disposition"]
        assert _xlsx_bytes(response)[:4] == b"PK\x03\x04"

    def test_export_contains_data(self, client, exporter_user, shipment):
        import openpyxl, io
        client.login(username="exporter_auto", password="pass")
        response = client.get(reverse("auto:export"))
        wb = openpyxl.load_workbook(io.BytesIO(_xlsx_bytes(response)))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 2
        assert rows[1][3] == "Объект А"

    def test_export_excludes_soft_deleted_rows_via_default_manager(self, client, exporter_user, shipment):
        import openpyxl, io

        AutoShipment.all_objects.create(
            shipment_date="2026-01-16",
            customer_object="Удалённый экспорт",
            coal_grade="ДГ",
            quantity="100",
            is_deleted=True,
        )
        client.login(username="exporter_auto", password="pass")

        response = client.get(reverse("auto:export"))

        wb = openpyxl.load_workbook(io.BytesIO(_xlsx_bytes(response)))
        rows = list(wb.active.iter_rows(values_only=True))
        objects = [row[3] for row in rows[1:]]
        assert shipment.customer_object in objects
        assert "Удалённый экспорт" not in objects

    def test_export_empty_queryset(self, client, exporter_user):
        import openpyxl, io
        client.login(username="exporter_auto", password="pass")
        response = client.get(reverse("auto:export"))
        wb = openpyxl.load_workbook(io.BytesIO(_xlsx_bytes(response)))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 1

    def test_export_respects_search_filter(self, client, exporter_user, shipment):
        import openpyxl, io
        AutoShipment.objects.create(
            shipment_date="2026-02-01",
            customer_object="Другой объект",
            coal_grade="Т",
            quantity="100",
        )
        client.login(username="exporter_auto", password="pass")
        response = client.get(reverse("auto:export") + "?q=Объект+А")
        wb = openpyxl.load_workbook(io.BytesIO(_xlsx_bytes(response)))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 2
        assert rows[1][3] == "Объект А"

    def test_export_respects_date_filter(self, client, exporter_user, shipment):
        import openpyxl, io
        AutoShipment.objects.create(
            shipment_date="2026-03-01",
            customer_object="Мартовский",
            coal_grade="ДГ",
            quantity="50",
        )
        client.login(username="exporter_auto", password="pass")
        response = client.get(reverse("auto:export") + "?date_from=2026-03-01&date_to=2026-03-31")
        wb = openpyxl.load_workbook(io.BytesIO(_xlsx_bytes(response)))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 2
        assert rows[1][3] == "Мартовский"

    def test_export_invalid_top_level_date_filter_is_ignored_without_500(self, client, exporter_user, shipment):
        client.login(username="exporter_auto", password="pass")
        response = client.get(reverse("auto:export") + "?date_from=zzzz&date_to=abc")

        assert response.status_code == 200

    def test_export_respects_column_filter(self, client, exporter_user, shipment):
        import openpyxl, io
        AutoShipment.objects.create(
            shipment_date="2026-02-01",
            customer_object="Другой объект",
            coal_grade="Т",
            quantity="100",
        )
        client.login(username="exporter_auto", password="pass")
        response = client.get(reverse("auto:export") + "?f_coal_grade=ДГ")
        wb = openpyxl.load_workbook(io.BytesIO(_xlsx_bytes(response)))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 2
        assert rows[1][8] == "ДГ"

    def test_export_respects_text_column_filter(self, client, exporter_user):
        import openpyxl, io
        from core.field_config import invalidate_entity_config
        from core.models import FieldSettings

        FieldSettings.objects.filter(entity="auto_shipment", field_name="comment").update(
            allow_filter=True,
            filter_type="text",
            visible=True,
            show_in_list=True,
        )
        invalidate_entity_config("auto_shipment")
        AutoShipment.objects.create(
            shipment_date="2026-02-01",
            customer_object="Нужный",
            coal_grade="ДГ",
            quantity="100",
            comment="точный маркер",
        )
        AutoShipment.objects.create(
            shipment_date="2026-02-02",
            customer_object="Лишний",
            coal_grade="ДГ",
            quantity="100",
            comment="другой текст",
        )
        client.login(username="exporter_auto", password="pass")
        response = client.get(reverse("auto:export") + "?f_comment=маркер")
        wb = openpyxl.load_workbook(io.BytesIO(_xlsx_bytes(response)))
        rows = list(wb.active.iter_rows(values_only=True))
        assert len(rows) == 2
        assert rows[1][3] == "Нужный"

    def test_export_rejects_too_many_rows(self, client, exporter_user, settings):
        settings.FULL_EXPORT_MAX_ROWS = 1
        AutoShipment.objects.create(shipment_date="2026-01-01", customer_object="A", coal_grade="ДГ", quantity="10")
        AutoShipment.objects.create(shipment_date="2026-01-02", customer_object="B", coal_grade="ДГ", quantity="10")
        client.force_login(exporter_user)
        response = client.get(reverse("auto:export"), follow=True)
        assert response.status_code == 200
        msgs = list(response.context["messages"])
        assert any("Слишком много" in str(m) for m in msgs)

    def test_export_reject_redirects_to_known_list_with_filters(self, client, exporter_user, settings):
        settings.FULL_EXPORT_MAX_ROWS = 1
        AutoShipment.objects.create(shipment_date="2026-01-01", customer_object="A", coal_grade="ДГ", quantity="10")
        AutoShipment.objects.create(shipment_date="2026-01-02", customer_object="B", coal_grade="ДГ", quantity="10")
        client.force_login(exporter_user)
        query = "?date_from=2026-01-01&date_to=2026-12-31"

        response = client.get(
            reverse("auto:export") + query,
            HTTP_REFERER="https://example.invalid/not-used/",
        )

        assert response.status_code == 302
        assert response["Location"] == reverse("auto:list") + query

    def test_export_large_export_is_logged(self, client, exporter_user, settings):
        settings.FULL_EXPORT_MAX_ROWS = 10000
        settings.PARTIAL_EXPORT_MAX_IDS = 0
        AutoShipment.objects.create(shipment_date="2026-01-01", customer_object="A", coal_grade="ДГ", quantity="10")
        client.force_login(exporter_user)
        logger = logging.getLogger("core.shipment_views")
        handler = ListHandler(level=logging.INFO)
        old_level = logger.level
        logger.addHandler(handler)
        logger.setLevel(logging.INFO)
        try:
            response = client.get(reverse("auto:export"))
        finally:
            logger.removeHandler(handler)
            logger.setLevel(old_level)
        assert response.status_code == 200
        assert any("Large export" in r.getMessage() for r in handler.records)


@pytest.mark.django_db
class TestAutoShipmentFilterSort:
    def test_column_filter_coal_grade(self, client, viewer_user):
        AutoShipment.objects.create(
            shipment_date="2026-01-01", customer_object="Объект-ДГ", coal_grade="ДГ", quantity="10",
        )
        AutoShipment.objects.create(
            shipment_date="2026-01-02", customer_object="Объект-Т", coal_grade="Т", quantity="20",
        )
        client.login(username="viewer_auto", password="pass")
        response = client.get(reverse("auto:list") + "?f_coal_grade=ДГ")
        grades = [s.coal_grade for s in response.context["shipments"]]
        assert "ДГ" in grades
        assert "Т" not in grades

    def test_column_filter_multiple_values(self, client, viewer_user):
        AutoShipment.objects.create(
            shipment_date="2026-01-01", customer_object="Объект-ДГ", coal_grade="ДГ", quantity="10",
        )
        AutoShipment.objects.create(
            shipment_date="2026-01-02", customer_object="Объект-Т", coal_grade="Т", quantity="20",
        )
        AutoShipment.objects.create(
            shipment_date="2026-01-03", customer_object="Объект-АК", coal_grade="АК", quantity="30",
        )
        client.login(username="viewer_auto", password="pass")
        response = client.get(reverse("auto:list") + "?f_coal_grade=ДГ&f_coal_grade=Т")
        grades = [s.coal_grade for s in response.context["shipments"]]
        assert "ДГ" in grades
        assert "Т" in grades
        assert "АК" not in grades

    def test_sort_by_quantity_asc(self, client, viewer_user):
        AutoShipment.objects.create(
            shipment_date="2026-01-01", customer_object="Малый", coal_grade="ДГ", quantity="10",
        )
        AutoShipment.objects.create(
            shipment_date="2026-01-02", customer_object="Большой", coal_grade="Т", quantity="999",
        )
        client.login(username="viewer_auto", password="pass")
        response = client.get(reverse("auto:list") + "?sort=quantity&dir=asc")
        assert response.status_code == 200
        shipments = list(response.context["shipments"])
        assert shipments[0].customer_object == "Малый"
        assert shipments[1].customer_object == "Большой"

    def test_sort_by_quantity_desc(self, client, viewer_user):
        AutoShipment.objects.create(
            shipment_date="2026-01-01", customer_object="Малый", coal_grade="ДГ", quantity="10",
        )
        AutoShipment.objects.create(
            shipment_date="2026-01-02", customer_object="Большой", coal_grade="Т", quantity="999",
        )
        client.login(username="viewer_auto", password="pass")
        response = client.get(reverse("auto:list") + "?sort=quantity&dir=desc")
        assert response.status_code == 200
        shipments = list(response.context["shipments"])
        assert shipments[0].customer_object == "Большой"

    def test_unknown_sort_field_ignored(self, client, viewer_user):
        client.login(username="viewer_auto", password="pass")
        response = client.get(reverse("auto:list") + "?sort=__evil__&dir=asc")
        assert response.status_code == 200

    def test_unknown_filter_field_ignored(self, client, viewer_user):
        client.login(username="viewer_auto", password="pass")
        response = client.get(reverse("auto:list") + "?f___evil__=x")
        assert response.status_code == 200

    def test_filter_urls_in_context(self, client, viewer_user):
        client.login(username="viewer_auto", password="pass")
        response = client.get(reverse("auto:list"))
        assert "filter_urls" in response.context
        assert "coal_grade" in response.context["filter_urls"]
        assert "filter-values/coal_grade" in response.context["filter_urls"]["coal_grade"]
        assert response.context["filter_query_safe_limit"] == 3694

    def test_active_filters_in_context(self, client, viewer_user):
        client.login(username="viewer_auto", password="pass")
        response = client.get(reverse("auto:list") + "?f_coal_grade=ДГ")
        assert response.context["active_filters"].get("coal_grade") == ["ДГ"]

    def test_text_column_filter_comment(self, client, viewer_user):
        from core.field_config import invalidate_entity_config
        from core.models import FieldSettings

        FieldSettings.objects.filter(entity="auto_shipment", field_name="comment").update(
            allow_filter=True,
            filter_type="text",
            visible=True,
            show_in_list=True,
        )
        invalidate_entity_config("auto_shipment")
        AutoShipment.objects.create(
            shipment_date="2026-01-01",
            customer_object="Нужный",
            coal_grade="ДГ",
            quantity="10",
            comment="особый маркер",
        )
        AutoShipment.objects.create(
            shipment_date="2026-01-02",
            customer_object="Лишний",
            coal_grade="ДГ",
            quantity="20",
            comment="без совпадения",
        )
        client.login(username="viewer_auto", password="pass")
        response = client.get(reverse("auto:list") + "?f_comment=маркер")

        objects = [s.customer_object for s in response.context["shipments"]]
        assert objects == ["Нужный"]
        assert response.context["active_text_filters"].get("comment") == "маркер"

    def test_column_filters_are_preserved_in_toolbar_and_sort_links(self, client, viewer_user):
        from core.field_config import invalidate_entity_config
        from core.models import FieldSettings

        FieldSettings.objects.filter(entity="auto_shipment", field_name="comment").update(
            allow_filter=True,
            filter_type="text",
            visible=True,
            show_in_list=True,
        )
        invalidate_entity_config("auto_shipment")
        AutoShipment.objects.create(
            shipment_date="2026-01-01",
            customer_object="А",
            coal_grade="ДГ",
            quantity="10",
            comment="маркер",
        )
        client.login(username="viewer_auto", password="pass")
        response = client.get(reverse("auto:list") + "?f_comment=маркер&f_quantity_from=5")
        html = response.content.decode()

        assert 'name="f_comment" value="маркер"' in html
        assert 'name="f_quantity_from" value="5"' in html
        assert "f_comment=%D0%BC%D0%B0%D1%80%D0%BA%D0%B5%D1%80" in html
        assert "f_quantity_from=5" in html


@pytest.mark.django_db
class TestAutoShipmentExportSelected:
    @pytest.fixture
    def exporter_user(self, django_user_model):
        user = django_user_model.objects.create_user(username="exporter_sel_auto", password="pass")
        group, _ = Group.objects.get_or_create(name="exporter_sel_auto_group")
        for codename in ("view_autoshipment", "export_excel"):
            perm = Permission.objects.get(codename=codename, content_type__app_label="shipments_auto")
            group.permissions.add(perm)
        user.groups.add(group)
        return user

    def test_export_selected_returns_xlsx(self, client, exporter_user, shipment):
        client.login(username="exporter_sel_auto", password="pass")
        response = client.post(
            reverse("auto:export_selected"),
            {"ids": [str(shipment.pk)]},
        )
        assert response.status_code == 200
        assert response["Content-Type"] == (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert _xlsx_bytes(response)[:4] == b"PK\x03\x04"

    def test_export_selected_only_chosen_rows(self, client, exporter_user, shipment):
        import openpyxl, io
        other = AutoShipment.objects.create(
            shipment_date="2026-02-01",
            customer_object="Другой объект",
            coal_grade="Т",
            quantity="100",
        )
        client.login(username="exporter_sel_auto", password="pass")
        response = client.post(
            reverse("auto:export_selected"),
            {"ids": [str(shipment.pk)]},
        )
        wb = openpyxl.load_workbook(io.BytesIO(_xlsx_bytes(response)))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 2
        assert rows[1][3] == "Объект А"
        objects_in_file = [r[3] for r in rows[1:]]
        assert other.customer_object not in objects_in_file

    def test_export_selected_empty_ids_redirects_with_message(self, client, exporter_user):
        client.login(username="exporter_sel_auto", password="pass")
        response = client.post(reverse("auto:export_selected"), {"ids": []}, follow=True)
        assert response.status_code == 200
        assert response.redirect_chain == [(reverse("auto:list"), 302)]
        assert any("Не выбраны записи для экспорта." in str(m) for m in response.context["messages"])

    def test_export_selected_requires_permission(self, client, viewer_user):
        client.login(username="viewer_auto", password="pass")
        response = client.post(reverse("auto:export_selected"), {"ids": ["1"]})
        assert response.status_code == 403

    def test_export_selected_rejects_invalid_ids_with_redirect(self, client, exporter_user):
        client.login(username="exporter_sel_auto", password="pass")
        response = client.post(reverse("auto:export_selected"), {"ids": ["abc"]}, follow=True)
        assert response.status_code == 200
        assert response.redirect_chain == [(reverse("auto:list"), 302)]
        assert any("Некорректный список ID для экспорта." in str(m) for m in response.context["messages"])

    def test_export_selected_rejects_zero_and_negative_ids_with_redirect(self, client, exporter_user):
        client.login(username="exporter_sel_auto", password="pass")
        for value in ("0", "-1"):
            response = client.post(reverse("auto:export_selected"), {"ids": [value]}, follow=True)
            assert response.status_code == 200
            assert response.redirect_chain == [(reverse("auto:list"), 302)]
            assert any("Некорректный список ID для экспорта." in str(m) for m in response.context["messages"])

    def test_export_selected_rejects_too_many_ids_with_redirect(self, client, exporter_user, settings):
        settings.PARTIAL_EXPORT_MAX_IDS = 1
        client.login(username="exporter_sel_auto", password="pass")
        response = client.post(reverse("auto:export_selected"), {"ids": ["1", "2"]}, follow=True)
        assert response.status_code == 200
        assert response.redirect_chain == [(reverse("auto:list"), 302)]
        assert any("Можно экспортировать не более 1 записей за раз." in str(m) for m in response.context["messages"])

    def test_export_selected_rejects_missing_or_deleted_ids_with_redirect(self, client, exporter_user, shipment):
        client.login(username="exporter_sel_auto", password="pass")
        missing_response = client.post(reverse("auto:export_selected"), {"ids": ["999999"]}, follow=True)
        assert missing_response.status_code == 200
        assert missing_response.redirect_chain == [(reverse("auto:list"), 302)]
        assert any("Часть выбранных записей не найдена или недоступна." in str(m) for m in missing_response.context["messages"])
        shipment.delete()
        response = client.post(reverse("auto:export_selected"), {"ids": [str(shipment.pk)]}, follow=True)
        assert response.status_code == 200
        assert response.redirect_chain == [(reverse("auto:list"), 302)]
        assert any("Часть выбранных записей не найдена или недоступна." in str(m) for m in response.context["messages"])

    def test_export_selected_rejection_preserves_query_string(self, client, exporter_user):
        client.login(username="exporter_sel_auto", password="pass")
        query = "?date_from=2026-01-01&sort=customer_object&dir=asc"
        response = client.post(reverse("auto:export_selected") + query, {"ids": ["abc"]})
        assert response.status_code == 302
        assert response["Location"] == reverse("auto:list") + query

    def test_export_selected_form_action_preserves_filters_without_page(self, client, exporter_user):
        client.login(username="exporter_sel_auto", password="pass")
        response = client.get(reverse("auto:list") + "?q=%D0%A2%D0%B5%D1%81%D1%82&page=1&sort=customer_object")
        html = response.content.decode()
        assert f'action="{reverse("auto:export_selected")}?q=%D0%A2%D0%B5%D1%81%D1%82&amp;sort=customer_object"' in html

    def test_export_selected_deduplicates_ids(self, client, exporter_user, shipment):
        import openpyxl, io
        client.login(username="exporter_sel_auto", password="pass")
        response = client.post(
            reverse("auto:export_selected"),
            {"ids": [str(shipment.pk), str(shipment.pk)]},
        )
        assert response.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(_xlsx_bytes(response)))
        rows = list(wb.active.iter_rows(values_only=True))
        assert len(rows) == 2


@pytest.mark.django_db
class TestAutoShipmentDeletedAndRestore:
    @pytest.fixture
    def deleted_shipment(self, operator_user):
        s = AutoShipment.all_objects.create(
            shipment_date="2026-03-10",
            customer_object="Удалённый объект",
            coal_grade="Т",
            quantity="50",
            is_deleted=True,
            created_by=operator_user,
            updated_by=operator_user,
        )
        return s

    def test_viewer_cannot_access_deleted_list(self, client, viewer_user):
        client.login(username="viewer_auto", password="pass")
        response = client.get(reverse("auto:deleted"))
        assert response.status_code == 403

    def test_operator_can_access_deleted_list(self, client, operator_user, deleted_shipment):
        client.login(username="operator_auto", password="pass")
        response = client.get(reverse("auto:deleted"))
        assert response.status_code == 200

    def test_deleted_list_shows_only_deleted(self, client, operator_user, shipment, deleted_shipment):
        client.login(username="operator_auto", password="pass")
        response = client.get(reverse("auto:deleted"))
        content = response.content.decode()
        assert "Удалённый объект" in content
        assert "Объект А" not in content

    def test_restore_success(self, client, operator_user, deleted_shipment):
        client.login(username="operator_auto", password="pass")
        response = client.post(reverse("auto:restore", kwargs={"pk": deleted_shipment.pk}))
        assert response.status_code == 302
        deleted_shipment.refresh_from_db()
        assert deleted_shipment.is_deleted is False

    def test_restore_writes_audit_log(self, client, operator_user, deleted_shipment):
        from audit.models import AuditLog
        client.login(username="operator_auto", password="pass")
        client.post(reverse("auto:restore", kwargs={"pk": deleted_shipment.pk}))
        assert AuditLog.objects.filter(
            entity_type=AuditLog.ENTITY_AUTO,
            entity_id=deleted_shipment.pk,
            action=AuditLog.ACTION_RESTORE,
        ).exists()

    def test_restore_record_leaves_deleted_list(self, client, operator_user, deleted_shipment):
        client.login(username="operator_auto", password="pass")
        client.post(reverse("auto:restore", kwargs={"pk": deleted_shipment.pk}))
        assert not AutoShipment.all_objects.filter(pk=deleted_shipment.pk, is_deleted=True).exists()

    def test_restore_record_in_main_list(self, client, operator_user, deleted_shipment):
        client.login(username="operator_auto", password="pass")
        client.post(reverse("auto:restore", kwargs={"pk": deleted_shipment.pk}))
        assert AutoShipment.objects.filter(pk=deleted_shipment.pk).exists()

    def test_restore_nonexistent_shows_error(self, client, operator_user):
        client.login(username="operator_auto", password="pass")
        response = client.post(reverse("auto:restore", kwargs={"pk": 999999}))
        assert response.status_code == 302
        assert response["Location"] == reverse("auto:deleted")

    def test_viewer_cannot_restore(self, client, viewer_user, deleted_shipment):
        client.login(username="viewer_auto", password="pass")
        response = client.post(reverse("auto:restore", kwargs={"pk": deleted_shipment.pk}))
        assert response.status_code == 403

    def test_restore_rolls_back_on_audit_failure(self, client, operator_user, deleted_shipment):
        """If log_shipment raises, obj.save() is also rolled back (atomic)."""
        from unittest.mock import patch
        import pytest
        client.login(username="operator_auto", password="pass")
        with patch("core.shipment_views.log_shipment", side_effect=Exception("audit fail")):
            with pytest.raises(Exception, match="audit fail"):
                client.post(reverse("auto:restore", kwargs={"pk": deleted_shipment.pk}))
        deleted_shipment.refresh_from_db()
        assert deleted_shipment.is_deleted is True


@pytest.mark.django_db
class TestAutoShipmentFormCatalogRequired:
    """Регресс B-920: clean() проверяет required для каталожных полей."""

    _BASE_DATA = {
        "shipment_date": "2026-01-15",
        "coal_grade": "ДГ",
        "quantity": "100.000",
    }

    def _cfg(self, required):
        return {
            "carrier": {
                "visible": True,
                "required": required,
                "use_catalog": True,
                "section": "main",
                "is_system": False,
                "show_in_list": True,
                "sort_order": 10,
                "label": "Перевозчик",
                "allow_filter": False,
                "allow_sort": False,
                "filter_type": "none",
                "sticky_col": False,
                "preset_membership": "",
            }
        }

    def test_required_catalog_field_empty_select_raises_error(self):
        from unittest.mock import patch
        from .forms import AutoShipmentForm

        data = {**self._BASE_DATA, "carrier_select": "", "carrier_other": ""}
        with patch("core.forms.get_entity_config", return_value=self._cfg(required=True)):
            form = AutoShipmentForm(data=data)
            assert not form.is_valid()
            assert "carrier_select" in form.errors

    def test_optional_catalog_field_empty_select_is_valid(self):
        from unittest.mock import patch
        from .forms import AutoShipmentForm

        data = {**self._BASE_DATA, "carrier_select": "", "carrier_other": ""}
        with patch("core.forms.get_entity_config", return_value=self._cfg(required=False)):
            form = AutoShipmentForm(data=data)
            assert form.is_valid()
            assert form.cleaned_data.get("carrier") == ""


@pytest.mark.django_db
class TestAutoFilterValuesView:
    def test_anonymous_redirects(self, client):
        response = client.get(reverse("auto:filter_values", kwargs={"field": "coal_grade"}))
        assert response.status_code == 302

    def test_returns_json_values(self, client, viewer_user):
        AutoShipment.objects.create(
            shipment_date="2026-01-01", customer_object="А", coal_grade="ДГ", quantity="10",
        )
        client.login(username="viewer_auto", password="pass")
        response = client.get(reverse("auto:filter_values", kwargs={"field": "coal_grade"}))
        assert response.status_code == 200
        data = response.json()
        assert "ДГ" in data["values"]
        assert data["has_more"] is False

    def test_search_filters_values(self, client, viewer_user):
        AutoShipment.objects.create(
            shipment_date="2026-01-01", customer_object="А", coal_grade="ДГ", quantity="10",
        )
        AutoShipment.objects.create(
            shipment_date="2026-01-02", customer_object="Б", coal_grade="СС", quantity="20",
        )
        client.login(username="viewer_auto", password="pass")
        response = client.get(
            reverse("auto:filter_values", kwargs={"field": "coal_grade"}) + "?q=дг"
        )
        assert response.status_code == 200
        data = response.json()
        assert "ДГ" in data["values"]
        assert "СС" not in data["values"]

    def test_non_sqlite_search_uses_single_icontains_lookup(self, client, viewer_user, monkeypatch):
        from django.db import connection
        from django.test.utils import CaptureQueriesContext

        AutoShipment.objects.create(
            shipment_date="2026-01-01",
            customer_object="А",
            coal_grade="target-grade",
            quantity="10",
        )
        AutoShipment.objects.create(
            shipment_date="2026-01-02",
            customer_object="Б",
            coal_grade="other-grade",
            quantity="20",
        )
        monkeypatch.setattr(connection, "vendor", "mysql")
        client.login(username="viewer_auto", password="pass")

        with CaptureQueriesContext(connection) as captured:
            response = client.get(reverse("auto:filter_values", kwargs={"field": "coal_grade"}) + "?q=target")

        assert response.status_code == 200
        assert response.json()["values"] == ["target-grade"]
        auto_selects = [
            query["sql"]
            for query in captured.captured_queries
            if "select" in query["sql"].lower() and "auto_shipments" in query["sql"].lower()
        ]
        value_query = next(sql for sql in auto_selects if "coal_grade" in sql)
        assert value_query.upper().count("LIKE") == 1

    def test_limits_distinct_values_in_sql_without_cache(self, client, viewer_user):
        from django.core.cache import cache
        from django.db import connection
        from django.test.utils import CaptureQueriesContext

        cache_key = "filter_values:auto_shipment:coal_grade"
        cache.delete(cache_key)
        for i in range(205):
            AutoShipment.objects.create(
                shipment_date="2026-01-01",
                customer_object=f"Объект {i}",
                coal_grade=f"GRADE-{i:03d}",
                quantity="10",
            )
        client.login(username="viewer_auto", password="pass")

        with CaptureQueriesContext(connection) as captured:
            response = client.get(reverse("auto:filter_values", kwargs={"field": "coal_grade"}))

        assert response.status_code == 200
        data = response.json()
        assert len(data["values"]) == 200
        assert data["has_more"] is True
        auto_selects = [
            query["sql"]
            for query in captured.captured_queries
            if "select" in query["sql"].lower() and "auto_shipments" in query["sql"].lower()
        ]
        assert any("LIMIT 201" in sql.upper() for sql in auto_selects)
        assert cache.get(cache_key) is None

    def test_invalid_field_returns_404(self, client, viewer_user):
        client.login(username="viewer_auto", password="pass")
        response = client.get(reverse("auto:filter_values", kwargs={"field": "nonexistent_xyz"}))
        assert response.status_code == 404

    def test_query_filter_values_unknown_field_returns_tuple(self):
        """V17-LOW-1: FieldDoesNotExist должен вернуть кортеж (values, has_more), а не []."""
        from core.shipment_views import _query_filter_values

        result = _query_filter_values(AutoShipment, "nonexistent_xyz")
        assert result == ([], False)
        values, has_more = result  # распаковка не должна падать ValueError
        assert values == []
        assert has_more is False


class TestV9Med1DeadSignalRemoved:
    """V9-MED-1: Мёртвые сигналы инвалидации фильтр-кэша удалены."""

    def test_invalidate_filter_values_cache_removed(self):
        from core import shipment_views
        assert not hasattr(shipment_views, "invalidate_filter_values_cache")

    def test_no_cache_import_in_shipment_views(self):
        import inspect
        from core import shipment_views
        src = inspect.getsourcefile(shipment_views)
        with open(src) as f:
            content = f.read()
        assert "from django.core.cache import cache" not in content


class TestV9Low3FullExportUsesSoftDeleteManager:
    """V9-LOW-3: full export relies on SoftDeleteManager for active rows."""

    def test_full_export_does_not_duplicate_soft_delete_filter(self):
        import inspect
        from core import shipment_views

        source = inspect.getsource(shipment_views.ShipmentExportMixin.get)

        assert "objects.filter(is_deleted=False)" not in source
        assert "objects.all()" in source


class TestV9Low4FullExportRedirect:
    """V9-LOW-4: full export rejection redirects to known list views."""

    def test_full_export_rejection_does_not_use_http_referer(self):
        import inspect
        from core import shipment_views

        source = inspect.getsource(shipment_views.ShipmentExportMixin.get)

        assert "HTTP_REFERER" not in source


@pytest.mark.django_db
class TestAutoShipmentListFilterChipLabels:
    """Фильтр-чипы показывают человекочитаемые метки вместо имён полей."""

    def test_range_filter_chip_shows_label(self, client, django_user_model):
        from core.field_config import invalidate_entity_config
        from core.models import FieldSettings
        user = django_user_model.objects.create_user(username="lbl_test_range", password="pass")
        user.user_permissions.add(*Permission.objects.filter(
            content_type__app_label="shipments_auto"
        ))
        FieldSettings.objects.filter(entity="auto_shipment", field_name="shipment_date").update(
            allow_filter=True, filter_type="date", label="Дата отгрузки"
        )
        invalidate_entity_config("auto_shipment")
        client.force_login(user)
        response = client.get(reverse("auto:list") + "?f_shipment_date_from=2024-01-01")
        content = response.content.decode()
        assert "Дата отгрузки" in content, "Чип должен показывать 'Дата отгрузки'"
        assert "shipment_date: от" not in content, "Чип не должен показывать raw имя поля"

    def test_text_filter_chip_shows_label(self, client, django_user_model):
        from core.field_config import invalidate_entity_config
        from core.models import FieldSettings
        user = django_user_model.objects.create_user(username="lbl_test_text", password="pass")
        user.user_permissions.add(*Permission.objects.filter(
            content_type__app_label="shipments_auto"
        ))
        FieldSettings.objects.filter(entity="auto_shipment", field_name="customer_object").update(
            allow_filter=True, filter_type="text", label="Объект"
        )
        invalidate_entity_config("auto_shipment")
        client.force_login(user)
        response = client.get(reverse("auto:list") + "?f_customer_object=Тест")
        content = response.content.decode()
        assert "Объект: Тест" in content, "Чип должен показывать 'Объект: Тест'"
        assert "customer_object: Тест" not in content


@pytest.mark.django_db
class TestAutoShipmentQuantityValidator:
    def test_quantity_zero_fails_full_clean(self, django_user_model):
        import datetime
        from decimal import Decimal
        from django.core.exceptions import ValidationError
        user = django_user_model.objects.create_user(username="qval_auto", password="pass")
        obj = AutoShipment(
            shipment_date=datetime.date(2024, 1, 15),
            customer_object="Объект",
            coal_grade="ДР",
            quantity=Decimal("0"),
            created_by=user,
            updated_by=user,
        )
        with pytest.raises(ValidationError):
            obj.full_clean()

    def test_quantity_negative_fails_full_clean(self, django_user_model):
        import datetime
        from decimal import Decimal
        from django.core.exceptions import ValidationError
        user = django_user_model.objects.create_user(username="qval_auto2", password="pass")
        obj = AutoShipment(
            shipment_date=datetime.date(2024, 1, 15),
            customer_object="Объект",
            coal_grade="ДР",
            quantity=Decimal("-10"),
            created_by=user,
            updated_by=user,
        )
        with pytest.raises(ValidationError):
            obj.full_clean()

    def test_quantity_positive_passes_full_clean(self, django_user_model):
        import datetime
        from decimal import Decimal
        user = django_user_model.objects.create_user(username="qval_auto3", password="pass")
        obj = AutoShipment(
            shipment_date=datetime.date(2024, 1, 15),
            customer_object="Объект",
            coal_grade="ДР",
            quantity=Decimal("42.5"),
            created_by=user,
            updated_by=user,
        )
        obj.full_clean()  # должно пройти без исключения
