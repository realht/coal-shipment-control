import logging
import pytest
from django.test import Client
from django.urls import reverse
from django.contrib.auth.models import Group, Permission
from urllib.parse import quote
from unittest.mock import patch

from .models import RailShipment


def _xlsx_bytes(response):
    from django.http import StreamingHttpResponse
    if isinstance(response, StreamingHttpResponse):
        return b"".join(response.streaming_content)
    return response.content


class ListHandler(logging.Handler):
    def __init__(self, level=logging.NOTSET):
        super().__init__(level=level)
        self.records = []

    def emit(self, record):
        self.records.append(record)


@pytest.fixture
def client():
    return Client()


@pytest.fixture
def viewer_user(django_user_model):
    user = django_user_model.objects.create_user(username="viewer_rail", password="pass")
    group, _ = Group.objects.get_or_create(name="viewer")
    perm = Permission.objects.get(codename="view_railshipment", content_type__app_label="shipments_rail")
    group.permissions.add(perm)
    user.groups.add(group)
    return user


@pytest.fixture
def operator_user(django_user_model):
    user = django_user_model.objects.create_user(username="operator_rail", password="pass")
    group, _ = Group.objects.get_or_create(name="operator_rail")
    for codename in ("view_railshipment", "add_railshipment", "change_railshipment", "delete_railshipment"):
        perm = Permission.objects.get(codename=codename, content_type__app_label="shipments_rail")
        group.permissions.add(perm)
    user.groups.add(group)
    return user


@pytest.fixture
def shipment(operator_user):
    return RailShipment.objects.create(
        departure_date="2026-01-20",
        wagon_number="12345678",
        cargo="Уголь ДГ",
        receiver="ООО Получатель",
        volume="2500.000",
        created_by=operator_user,
        updated_by=operator_user,
    )


@pytest.mark.django_db
class TestRailShipmentList:
    def test_anonymous_redirects(self, client):
        response = client.get(reverse("rail:list"))
        assert response.status_code == 302
        assert "/accounts/login/" in response["Location"]

    def test_viewer_can_see_list(self, client, viewer_user):
        client.login(username="viewer_rail", password="pass")
        response = client.get(reverse("rail:list"))
        assert response.status_code == 200

    def test_list_table_presets_use_json_script(self, client, viewer_user):
        client.login(username="viewer_rail", password="pass")
        response = client.get(reverse("rail:list"))
        html = response.content.decode()
        assert 'data-presets-script-id="rail-table-presets"' in html
        assert 'id="rail-table-presets"' in html
        assert 'data-presets="' not in html

    def test_list_shows_shipment(self, client, viewer_user, shipment):
        client.login(username="viewer_rail", password="pass")
        response = client.get(reverse("rail:list"))
        assert "12345678" in response.content.decode()

    def test_viewer_has_no_add_button(self, client, viewer_user):
        client.login(username="viewer_rail", password="pass")
        response = client.get(reverse("rail:list"))
        assert "/rail/new/" not in response.content.decode()

    def test_operator_has_add_button(self, client, operator_user):
        client.login(username="operator_rail", password="pass")
        response = client.get(reverse("rail:list"))
        assert "/rail/new/" in response.content.decode()

    def test_search_filters_results(self, client, viewer_user, shipment):
        RailShipment.objects.create(
            departure_date="2026-01-21",
            wagon_number="99999999",
            cargo="Уголь Т",
            receiver="ООО Другой",
            volume="1000",
        )
        client.login(username="viewer_rail", password="pass")
        response = client.get(reverse("rail:list") + "?q=12345678")
        wagons = [s.wagon_number for s in response.context["shipments"]]
        assert "12345678" in wagons
        assert "99999999" not in wagons

    def test_search_no_duplicate_when_multiple_fields_match(self, client, viewer_user):
        RailShipment.objects.create(
            departure_date="2026-01-15",
            wagon_number="Уголь-вагон",
            cargo="Уголь-груз",
            volume="100",
        )
        client.login(username="viewer_rail", password="pass")
        response = client.get(reverse("rail:list") + "?q=Уголь")
        shipments = list(response.context["shipments"])
        ids = [s.pk for s in shipments]
        assert len(ids) == len(set(ids)), "Дубли в queryset при поиске по нескольким полям"

    def test_date_filter(self, client, viewer_user, shipment):
        RailShipment.objects.create(
            departure_date="2026-04-01",
            wagon_number="88888888",
            cargo="Уголь Ж",
            receiver="ООО Апрель",
            volume="500",
        )
        client.login(username="viewer_rail", password="pass")
        response = client.get(reverse("rail:list") + "?date_from=2026-03-01&date_to=2026-04-30")
        wagons = [s.wagon_number for s in response.context["shipments"]]
        assert "88888888" in wagons
        assert "12345678" not in wagons

    def test_invalid_top_level_date_filter_is_ignored_without_500(self, client, viewer_user, shipment):
        client.login(username="viewer_rail", password="pass")
        response = client.get(reverse("rail:list") + "?date_from=zzzz&date_to=abc")

        assert response.status_code == 200
        assert shipment in list(response.context["shipments"])

    def test_pagination(self, client, viewer_user):
        for i in range(30):
            RailShipment.objects.create(
                departure_date="2026-01-01",
                wagon_number=f"{i:08d}",
                cargo="Уголь",
                receiver=f"Получатель {i}",
                volume="100",
            )
        client.login(username="viewer_rail", password="pass")
        response = client.get(reverse("rail:list"))
        assert response.status_code == 200
        assert response.context["is_paginated"]
        assert len(response.context["shipments"]) == 25


@pytest.mark.django_db
class TestRailShipmentDetail:
    def test_viewer_can_see_detail(self, client, viewer_user, shipment):
        client.login(username="viewer_rail", password="pass")
        response = client.get(reverse("rail:detail", kwargs={"pk": shipment.pk}))
        assert response.status_code == 200
        assert "12345678" in response.content.decode()

    def test_viewer_has_no_edit_button(self, client, viewer_user, shipment):
        client.login(username="viewer_rail", password="pass")
        response = client.get(reverse("rail:detail", kwargs={"pk": shipment.pk}))
        assert "Редактировать" not in response.content.decode()

    def test_operator_has_edit_button(self, client, operator_user, shipment):
        client.login(username="operator_rail", password="pass")
        response = client.get(reverse("rail:detail", kwargs={"pk": shipment.pk}))
        assert "Редактировать" in response.content.decode()


@pytest.mark.django_db
class TestRailShipmentCreate:
    def test_viewer_cannot_create(self, client, viewer_user):
        client.login(username="viewer_rail", password="pass")
        response = client.get(reverse("rail:create"))
        assert response.status_code == 403

    def test_operator_can_get_form(self, client, operator_user):
        client.login(username="operator_rail", password="pass")
        response = client.get(reverse("rail:create"))
        assert response.status_code == 200

    def test_operator_can_create(self, client, operator_user):
        client.login(username="operator_rail", password="pass")
        response = client.post(reverse("rail:create"), {
            "departure_date": "2026-03-05",
            "wagon_number": "77777777",
            "document_number": "",
            "cargo_select": "__other__",
            "cargo_other": "Уголь ДГ",
            "destination_station": "",
            "receiver": "Новый получатель",
            "volume": "3000.000",
            "comment": "",
            "origin_region": "",
            "origin_station": "",
            "sender": "",
            "destination_region": "",
        })
        assert response.status_code == 302
        assert RailShipment.objects.filter(wagon_number="77777777").exists()

    def test_decimal_precision_preserved(self, client, operator_user):
        client.login(username="operator_rail", password="pass")
        client.post(reverse("rail:create"), {
            "departure_date": "2026-03-06",
            "wagon_number": "66666666",
            "document_number": "",
            "cargo_select": "__other__",
            "cargo_other": "Уголь Т",
            "destination_station": "",
            "receiver": "Дробный получатель",
            "volume": "1234.567",
            "comment": "",
            "origin_region": "",
            "origin_station": "",
            "sender": "",
            "destination_region": "",
        })
        from decimal import Decimal
        obj = RailShipment.objects.get(wagon_number="66666666")
        assert obj.volume == Decimal("1234.567")

    def test_create_writes_audit_log(self, client, operator_user):
        from audit.models import AuditLog
        client.login(username="operator_rail", password="pass")
        client.post(reverse("rail:create"), {
            "departure_date": "2026-03-07",
            "wagon_number": "55555555",
            "document_number": "",
            "cargo_select": "__other__",
            "cargo_other": "Уголь",
            "destination_station": "",
            "receiver": "Аудит получатель",
            "volume": "100",
            "comment": "",
            "origin_region": "",
            "origin_station": "",
            "sender": "",
            "destination_region": "",
        })
        obj = RailShipment.objects.get(wagon_number="55555555")
        assert AuditLog.objects.filter(
            entity_type=AuditLog.ENTITY_RAIL,
            entity_id=obj.pk,
            action=AuditLog.ACTION_CREATE,
        ).exists()


@pytest.mark.django_db
class TestRailShipmentUpdate:
    def test_viewer_cannot_update(self, client, viewer_user, shipment):
        client.login(username="viewer_rail", password="pass")
        response = client.get(reverse("rail:update", kwargs={"pk": shipment.pk}))
        assert response.status_code == 403

    def test_operator_can_update(self, client, operator_user, shipment):
        client.login(username="operator_rail", password="pass")
        response = client.post(reverse("rail:update", kwargs={"pk": shipment.pk}), {
            "departure_date": "2026-01-20",
            "wagon_number": "12345678",
            "document_number": "",
            "cargo_select": "__other__",
            "cargo_other": "Уголь Т",
            "destination_station": "",
            "receiver": "Изменённый получатель",
            "volume": "999.000",
            "comment": "",
            "origin_region": "",
            "origin_station": "",
            "sender": "",
            "destination_region": "",
            "updated_at_token": shipment.updated_at.isoformat(),
        })
        assert response.status_code == 302
        shipment.refresh_from_db()
        assert shipment.receiver == "Изменённый получатель"

    def test_update_writes_audit_log(self, client, operator_user, shipment):
        from audit.models import AuditLog
        client.login(username="operator_rail", password="pass")
        client.post(reverse("rail:update", kwargs={"pk": shipment.pk}), {
            "departure_date": "2026-01-20",
            "wagon_number": "12345678",
            "document_number": "",
            "cargo_select": "__other__",
            "cargo_other": "Уголь",
            "destination_station": "",
            "receiver": "После изменения",
            "volume": "100",
            "comment": "",
            "origin_region": "",
            "origin_station": "",
            "sender": "",
            "destination_region": "",
            "updated_at_token": shipment.updated_at.isoformat(),
        })
        assert AuditLog.objects.filter(
            entity_type=AuditLog.ENTITY_RAIL,
            entity_id=shipment.pk,
            action=AuditLog.ACTION_UPDATE,
        ).exists()


@pytest.mark.django_db
class TestRailShipmentOptimisticLock:
    _base_post = {
        "departure_date": "2026-01-20",
        "wagon_number": "12345678",
        "document_number": "",
        "cargo_select": "__other__",
        "cargo_other": "Уголь ДГ",
        "destination_station": "",
        "receiver": "ООО Получатель",
        "volume": "2500.000",
        "comment": "",
        "origin_region": "",
        "origin_station": "",
        "sender": "",
        "destination_region": "",
    }

    def test_missing_token_rejected(self, client, operator_user, shipment):
        client.login(username="operator_rail", password="pass")
        response = client.post(
            reverse("rail:update", kwargs={"pk": shipment.pk}),
            {**self._base_post},
        )
        assert response.status_code == 200
        shipment.refresh_from_db()
        assert shipment.receiver == "ООО Получатель"

    def test_stale_token_rejected_with_who_and_when(self, client, operator_user, shipment):
        stale_token = shipment.updated_at.isoformat()
        other = type(operator_user).objects.create_user(username="other_rail", password="p")
        shipment.receiver = "Чужой получатель"
        shipment.updated_by = other
        shipment.save()

        client.login(username="operator_rail", password="pass")
        response = client.post(
            reverse("rail:update", kwargs={"pk": shipment.pk}),
            {**self._base_post, "updated_at_token": stale_token},
        )
        assert response.status_code == 200
        content = response.content.decode()
        assert "other_rail" in content
        shipment.refresh_from_db()
        assert shipment.receiver == "Чужой получатель"

    def test_fresh_token_after_conflict_saves_ok(self, client, operator_user, shipment):
        other = type(operator_user).objects.create_user(username="other_rail2", password="p")
        shipment.receiver = "Промежуточный"
        shipment.updated_by = other
        shipment.save()

        client.login(username="operator_rail", password="pass")
        response = client.post(
            reverse("rail:update", kwargs={"pk": shipment.pk}),
            {**self._base_post, "receiver": "Финальный", "updated_at_token": shipment.updated_at.isoformat()},
        )
        assert response.status_code == 302
        shipment.refresh_from_db()
        assert shipment.receiver == "Финальный"

    def test_parallel_post_with_same_token_does_not_overwrite_first_save(self, client, operator_user, shipment):
        stale_snapshot = RailShipment.objects.get(pk=shipment.pk)
        token = stale_snapshot.updated_at.isoformat()

        client.login(username="operator_rail", password="pass")
        first_response = client.post(
            reverse("rail:update", kwargs={"pk": shipment.pk}),
            {**self._base_post, "receiver": "Первый получатель", "updated_at_token": token},
        )
        assert first_response.status_code == 302

        with patch("shipments_rail.views.RailShipmentUpdateView.get_object", return_value=stale_snapshot):
            second_response = client.post(
                reverse("rail:update", kwargs={"pk": shipment.pk}),
                {**self._base_post, "receiver": "Второй получатель", "updated_at_token": token},
            )

        assert second_response.status_code == 200
        assert "Ваши изменения не сохранены" in second_response.content.decode()
        shipment.refresh_from_db()
        assert shipment.receiver == "Первый получатель"


@pytest.mark.django_db
class TestRailShipmentSoftDelete:
    def test_viewer_cannot_delete(self, client, viewer_user, shipment):
        client.login(username="viewer_rail", password="pass")
        response = client.post(reverse("rail:delete", kwargs={"pk": shipment.pk}))
        assert response.status_code == 403

    def test_operator_can_soft_delete(self, client, operator_user, shipment):
        client.login(username="operator_rail", password="pass")
        response = client.post(reverse("rail:delete", kwargs={"pk": shipment.pk}))
        assert response.status_code == 302
        shipment.refresh_from_db()
        assert shipment.is_deleted is True

    def test_delete_redirect_preserves_list_filters(self, client, operator_user, shipment):
        client.login(username="operator_rail", password="pass")
        next_url = reverse("rail:list") + "?f_wagon_number=12345678&sort=wagon_number&dir=asc"
        response = client.post(
            reverse("rail:delete", kwargs={"pk": shipment.pk}) + f"?next={quote(next_url, safe='')}"
        )
        assert response.status_code == 302
        assert response["Location"] == next_url

    def test_delete_rejects_external_next_url(self, client, operator_user, shipment):
        client.login(username="operator_rail", password="pass")
        response = client.post(reverse("rail:delete", kwargs={"pk": shipment.pk}) + "?next=https%3A%2F%2Fevil.test%2F")
        assert response.status_code == 302
        assert response["Location"] == reverse("rail:list")

    def test_list_delete_link_includes_current_filters(self, client, operator_user, shipment):
        client.login(username="operator_rail", password="pass")
        response = client.get(reverse("rail:list") + "?q=12345678&sort=wagon_number&dir=asc")
        content = response.content.decode()
        expected_next = quote("/rail/?q=12345678&sort=wagon_number&dir=asc", safe="/")
        assert f'{reverse("rail:delete", kwargs={"pk": shipment.pk})}?next={expected_next}' in content

    def test_deleted_not_in_list(self, client, viewer_user, shipment):
        shipment.is_deleted = True
        shipment.save()
        client.login(username="viewer_rail", password="pass")
        response = client.get(reverse("rail:list"))
        assert "12345678" not in response.content.decode()

    def test_soft_delete_writes_audit_log(self, client, operator_user, shipment):
        from audit.models import AuditLog
        client.login(username="operator_rail", password="pass")
        client.post(reverse("rail:delete", kwargs={"pk": shipment.pk}))
        assert AuditLog.objects.filter(
            entity_type=AuditLog.ENTITY_RAIL,
            entity_id=shipment.pk,
            action=AuditLog.ACTION_DELETE,
        ).exists()

    def test_hard_delete_not_called(self, client, operator_user, shipment):
        pk = shipment.pk
        client.login(username="operator_rail", password="pass")
        client.post(reverse("rail:delete", kwargs={"pk": pk}))
        assert RailShipment.all_objects.filter(pk=pk).exists()


@pytest.mark.django_db
class TestRailShipmentExport:
    @pytest.fixture
    def exporter_user(self, django_user_model):
        user = django_user_model.objects.create_user(username="exporter_rail", password="pass")
        group, _ = Group.objects.get_or_create(name="exporter_rail_group")
        for codename in ("view_railshipment", "export_excel"):
            perm = Permission.objects.get(codename=codename, content_type__app_label="shipments_rail")
            group.permissions.add(perm)
        user.groups.add(group)
        return user

    def test_viewer_without_perm_gets_403(self, client, viewer_user):
        client.login(username="viewer_rail", password="pass")
        response = client.get(reverse("rail:export"))
        assert response.status_code == 403

    def test_anonymous_redirects(self, client):
        response = client.get(reverse("rail:export"))
        assert response.status_code == 302
        assert "/accounts/login/" in response["Location"]

    def test_export_returns_xlsx(self, client, exporter_user, shipment):
        client.login(username="exporter_rail", password="pass")
        response = client.get(reverse("rail:export"))
        assert response.status_code == 200
        assert response["Content-Type"] == (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert "attachment" in response["Content-Disposition"]
        assert _xlsx_bytes(response)[:4] == b"PK\x03\x04"

    def test_export_contains_data(self, client, exporter_user, shipment):
        import openpyxl, io
        client.login(username="exporter_rail", password="pass")
        response = client.get(reverse("rail:export"))
        wb = openpyxl.load_workbook(io.BytesIO(_xlsx_bytes(response)))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 2
        assert rows[1][9] == "ООО Получатель"

    def test_export_excludes_soft_deleted_rows_via_default_manager(self, client, exporter_user, shipment):
        import openpyxl, io

        RailShipment.all_objects.create(
            departure_date="2026-01-21",
            wagon_number="87654321",
            cargo="Уголь ДГ",
            receiver="Удалённый получатель",
            volume="100",
            is_deleted=True,
        )
        client.login(username="exporter_rail", password="pass")

        response = client.get(reverse("rail:export"))

        wb = openpyxl.load_workbook(io.BytesIO(_xlsx_bytes(response)))
        rows = list(wb.active.iter_rows(values_only=True))
        receivers = [row[9] for row in rows[1:]]
        assert shipment.receiver in receivers
        assert "Удалённый получатель" not in receivers

    def test_export_empty_queryset(self, client, exporter_user):
        import openpyxl, io
        client.login(username="exporter_rail", password="pass")
        response = client.get(reverse("rail:export"))
        wb = openpyxl.load_workbook(io.BytesIO(_xlsx_bytes(response)))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 1

    def test_export_respects_search_filter(self, client, exporter_user, shipment):
        import openpyxl, io
        RailShipment.objects.create(
            departure_date="2026-02-01",
            wagon_number="99999999",
            cargo="Уголь Т",
            receiver="Другой получатель",
            volume="100",
        )
        client.login(username="exporter_rail", password="pass")
        response = client.get(reverse("rail:export") + "?q=ООО+Получатель")
        wb = openpyxl.load_workbook(io.BytesIO(_xlsx_bytes(response)))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 2
        assert rows[1][9] == "ООО Получатель"

    def test_export_respects_date_filter(self, client, exporter_user, shipment):
        import openpyxl, io
        RailShipment.objects.create(
            departure_date="2026-03-01",
            wagon_number="88888888",
            cargo="Уголь ДГ",
            receiver="Мартовский получатель",
            volume="50",
        )
        client.login(username="exporter_rail", password="pass")
        response = client.get(reverse("rail:export") + "?date_from=2026-03-01&date_to=2026-03-31")
        wb = openpyxl.load_workbook(io.BytesIO(_xlsx_bytes(response)))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 2
        assert rows[1][9] == "Мартовский получатель"

    def test_export_invalid_top_level_date_filter_is_ignored_without_500(self, client, exporter_user, shipment):
        client.login(username="exporter_rail", password="pass")
        response = client.get(reverse("rail:export") + "?date_from=zzzz&date_to=abc")

        assert response.status_code == 200

    def test_export_rejects_too_many_rows(self, client, exporter_user, settings):
        settings.FULL_EXPORT_MAX_ROWS = 1
        RailShipment.objects.create(departure_date="2026-01-01", wagon_number="11111111", cargo="Уголь ДГ", volume="50")
        RailShipment.objects.create(departure_date="2026-01-02", wagon_number="22222222", cargo="Уголь ДГ", volume="50")
        client.force_login(exporter_user)
        response = client.get(reverse("rail:export"), follow=True)
        assert response.status_code == 200
        msgs = list(response.context["messages"])
        assert any("Слишком много" in str(m) for m in msgs)

    def test_export_reject_redirects_to_known_list_with_filters(self, client, exporter_user, settings):
        settings.FULL_EXPORT_MAX_ROWS = 1
        RailShipment.objects.create(departure_date="2026-01-01", wagon_number="11111111", cargo="Уголь ДГ", volume="50")
        RailShipment.objects.create(departure_date="2026-01-02", wagon_number="22222222", cargo="Уголь ДГ", volume="50")
        client.force_login(exporter_user)
        query = "?date_from=2026-01-01&date_to=2026-12-31"

        response = client.get(
            reverse("rail:export") + query,
            HTTP_REFERER="https://example.invalid/not-used/",
        )

        assert response.status_code == 302
        assert response["Location"] == reverse("rail:list") + query

    def test_export_large_export_is_logged(self, client, exporter_user, settings):
        settings.FULL_EXPORT_MAX_ROWS = 10000
        settings.PARTIAL_EXPORT_MAX_IDS = 0
        RailShipment.objects.create(departure_date="2026-01-01", wagon_number="11111111", cargo="Уголь ДГ", volume="50")
        client.force_login(exporter_user)
        logger = logging.getLogger("core.shipment_views")
        handler = ListHandler(level=logging.INFO)
        old_level = logger.level
        logger.addHandler(handler)
        logger.setLevel(logging.INFO)
        try:
            response = client.get(reverse("rail:export"))
        finally:
            logger.removeHandler(handler)
            logger.setLevel(old_level)
        assert response.status_code == 200
        assert any("Large export" in r.getMessage() for r in handler.records)


@pytest.mark.django_db
class TestRailShipmentExportSelected:
    @pytest.fixture
    def exporter_user(self, django_user_model):
        user = django_user_model.objects.create_user(username="exporter_sel_rail", password="pass")
        group, _ = Group.objects.get_or_create(name="exporter_sel_rail_group")
        for codename in ("view_railshipment", "export_excel"):
            perm = Permission.objects.get(codename=codename, content_type__app_label="shipments_rail")
            group.permissions.add(perm)
        user.groups.add(group)
        return user

    def test_export_selected_returns_xlsx(self, client, exporter_user, shipment):
        client.login(username="exporter_sel_rail", password="pass")
        response = client.post(
            reverse("rail:export_selected"),
            {"ids": [str(shipment.pk)]},
        )
        assert response.status_code == 200
        assert response["Content-Type"] == (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert _xlsx_bytes(response)[:4] == b"PK\x03\x04"

    def test_export_selected_only_chosen_rows(self, client, exporter_user, shipment):
        import openpyxl, io
        other = RailShipment.objects.create(
            departure_date="2026-02-01",
            wagon_number="99999999",
            cargo="Уголь Т",
            receiver="Другой получатель",
            volume="100",
        )
        client.login(username="exporter_sel_rail", password="pass")
        response = client.post(
            reverse("rail:export_selected"),
            {"ids": [str(shipment.pk)]},
        )
        wb = openpyxl.load_workbook(io.BytesIO(_xlsx_bytes(response)))
        rows = list(wb.active.iter_rows(values_only=True))
        assert len(rows) == 2
        assert rows[1][9] == "ООО Получатель"
        assert other.receiver not in [r[9] for r in rows[1:]]

    def test_export_selected_empty_ids_redirects_with_message(self, client, exporter_user):
        client.login(username="exporter_sel_rail", password="pass")
        response = client.post(reverse("rail:export_selected"), {"ids": []}, follow=True)
        assert response.status_code == 200
        assert response.redirect_chain == [(reverse("rail:list"), 302)]
        assert any("Не выбраны записи для экспорта." in str(m) for m in response.context["messages"])

    def test_export_selected_requires_permission(self, client, viewer_user):
        client.login(username="viewer_rail", password="pass")
        response = client.post(reverse("rail:export_selected"), {"ids": ["1"]})
        assert response.status_code == 403

    def test_export_selected_rejects_invalid_ids_with_redirect(self, client, exporter_user):
        client.login(username="exporter_sel_rail", password="pass")
        response = client.post(reverse("rail:export_selected"), {"ids": ["abc"]}, follow=True)
        assert response.status_code == 200
        assert response.redirect_chain == [(reverse("rail:list"), 302)]
        assert any("Некорректный список ID для экспорта." in str(m) for m in response.context["messages"])

    def test_export_selected_rejects_zero_and_negative_ids_with_redirect(self, client, exporter_user):
        client.login(username="exporter_sel_rail", password="pass")
        for value in ("0", "-1"):
            response = client.post(reverse("rail:export_selected"), {"ids": [value]}, follow=True)
            assert response.status_code == 200
            assert response.redirect_chain == [(reverse("rail:list"), 302)]
            assert any("Некорректный список ID для экспорта." in str(m) for m in response.context["messages"])

    def test_export_selected_rejects_too_many_ids_with_redirect(self, client, exporter_user, settings):
        settings.PARTIAL_EXPORT_MAX_IDS = 1
        client.login(username="exporter_sel_rail", password="pass")
        response = client.post(reverse("rail:export_selected"), {"ids": ["1", "2"]}, follow=True)
        assert response.status_code == 200
        assert response.redirect_chain == [(reverse("rail:list"), 302)]
        assert any("Можно экспортировать не более 1 записей за раз." in str(m) for m in response.context["messages"])

    def test_export_selected_rejects_missing_or_deleted_ids_with_redirect(self, client, exporter_user, shipment):
        client.login(username="exporter_sel_rail", password="pass")
        missing_response = client.post(reverse("rail:export_selected"), {"ids": ["999999"]}, follow=True)
        assert missing_response.status_code == 200
        assert missing_response.redirect_chain == [(reverse("rail:list"), 302)]
        assert any("Часть выбранных записей не найдена или недоступна." in str(m) for m in missing_response.context["messages"])
        shipment.delete()
        response = client.post(reverse("rail:export_selected"), {"ids": [str(shipment.pk)]}, follow=True)
        assert response.status_code == 200
        assert response.redirect_chain == [(reverse("rail:list"), 302)]
        assert any("Часть выбранных записей не найдена или недоступна." in str(m) for m in response.context["messages"])

    def test_export_selected_rejection_preserves_query_string(self, client, exporter_user):
        client.login(username="exporter_sel_rail", password="pass")
        query = "?date_from=2026-01-01&sort=wagon_number&dir=asc"
        response = client.post(reverse("rail:export_selected") + query, {"ids": ["abc"]})
        assert response.status_code == 302
        assert response["Location"] == reverse("rail:list") + query

    def test_export_selected_form_action_preserves_filters_without_page(self, client, exporter_user):
        client.login(username="exporter_sel_rail", password="pass")
        response = client.get(reverse("rail:list") + "?q=123&page=1&sort=wagon_number")
        html = response.content.decode()
        assert f'action="{reverse("rail:export_selected")}?q=123&amp;sort=wagon_number"' in html

    def test_export_selected_deduplicates_ids(self, client, exporter_user, shipment):
        import openpyxl, io
        client.login(username="exporter_sel_rail", password="pass")
        response = client.post(
            reverse("rail:export_selected"),
            {"ids": [str(shipment.pk), str(shipment.pk)]},
        )
        assert response.status_code == 200
        rows = list(openpyxl.load_workbook(io.BytesIO(_xlsx_bytes(response))).active.iter_rows(values_only=True))
        assert len(rows) == 2


@pytest.mark.django_db
class TestRailShipmentDeletedAndRestore:
    @pytest.fixture
    def deleted_shipment(self, operator_user):
        s = RailShipment.all_objects.create(
            departure_date="2026-03-10",
            wagon_number="99999999",
            cargo="Уголь Т",
            volume="100",
            is_deleted=True,
            created_by=operator_user,
            updated_by=operator_user,
        )
        return s

    def test_viewer_cannot_access_deleted_list(self, client, viewer_user):
        client.login(username="viewer_rail", password="pass")
        response = client.get(reverse("rail:deleted"))
        assert response.status_code == 403

    def test_operator_can_access_deleted_list(self, client, operator_user, deleted_shipment):
        client.login(username="operator_rail", password="pass")
        response = client.get(reverse("rail:deleted"))
        assert response.status_code == 200

    def test_deleted_list_shows_only_deleted(self, client, operator_user, shipment, deleted_shipment):
        client.login(username="operator_rail", password="pass")
        response = client.get(reverse("rail:deleted"))
        content = response.content.decode()
        assert "99999999" in content
        assert "12345678" not in content

    def test_restore_success(self, client, operator_user, deleted_shipment):
        client.login(username="operator_rail", password="pass")
        response = client.post(reverse("rail:restore", kwargs={"pk": deleted_shipment.pk}))
        assert response.status_code == 302
        deleted_shipment.refresh_from_db()
        assert deleted_shipment.is_deleted is False

    def test_restore_writes_audit_log(self, client, operator_user, deleted_shipment):
        from audit.models import AuditLog
        client.login(username="operator_rail", password="pass")
        client.post(reverse("rail:restore", kwargs={"pk": deleted_shipment.pk}))
        assert AuditLog.objects.filter(
            entity_type=AuditLog.ENTITY_RAIL,
            entity_id=deleted_shipment.pk,
            action=AuditLog.ACTION_RESTORE,
        ).exists()

    def test_restore_record_leaves_deleted_list(self, client, operator_user, deleted_shipment):
        client.login(username="operator_rail", password="pass")
        client.post(reverse("rail:restore", kwargs={"pk": deleted_shipment.pk}))
        assert not RailShipment.all_objects.filter(pk=deleted_shipment.pk, is_deleted=True).exists()

    def test_restore_record_in_main_list(self, client, operator_user, deleted_shipment):
        client.login(username="operator_rail", password="pass")
        client.post(reverse("rail:restore", kwargs={"pk": deleted_shipment.pk}))
        assert RailShipment.objects.filter(pk=deleted_shipment.pk).exists()

    def test_restore_nonexistent_shows_error(self, client, operator_user):
        client.login(username="operator_rail", password="pass")
        response = client.post(reverse("rail:restore", kwargs={"pk": 999999}))
        assert response.status_code == 302
        assert response["Location"] == reverse("rail:deleted")

    def test_viewer_cannot_restore(self, client, viewer_user, deleted_shipment):
        client.login(username="viewer_rail", password="pass")
        response = client.post(reverse("rail:restore", kwargs={"pk": deleted_shipment.pk}))
        assert response.status_code == 403


@pytest.mark.django_db
class TestRailFilterValuesView:
    def test_anonymous_redirects(self, client):
        response = client.get(reverse("rail:filter_values", kwargs={"field": "cargo"}))
        assert response.status_code == 302

    def test_returns_json_values(self, client, viewer_user):
        RailShipment.objects.create(
            departure_date="2026-01-01", wagon_number="№1", cargo="ДГ", volume="100",
        )
        client.login(username="viewer_rail", password="pass")
        response = client.get(reverse("rail:filter_values", kwargs={"field": "cargo"}))
        assert response.status_code == 200
        data = response.json()
        assert "ДГ" in data["values"]
        assert data["has_more"] is False

    def test_search_filters_values(self, client, viewer_user):
        RailShipment.objects.create(
            departure_date="2026-01-01", wagon_number="№1", cargo="ДГ", volume="100",
        )
        RailShipment.objects.create(
            departure_date="2026-01-02", wagon_number="№2", cargo="СС", volume="200",
        )
        client.login(username="viewer_rail", password="pass")
        response = client.get(
            reverse("rail:filter_values", kwargs={"field": "cargo"}) + "?q=дг"
        )
        assert response.status_code == 200
        data = response.json()
        assert "ДГ" in data["values"]
        assert "СС" not in data["values"]

    def test_search_filters_values_in_sql_with_limit(self, client, viewer_user):
        from django.db import connection
        from django.test.utils import CaptureQueriesContext

        for i in range(205):
            RailShipment.objects.create(
                departure_date="2026-01-01",
                wagon_number=f"{65000000 + i}",
                cargo=f"OTHER-{i:03d}",
                volume="100",
            )
        RailShipment.objects.create(
            departure_date="2026-01-02",
            wagon_number="66000000",
            cargo="target-cargo",
            volume="100",
        )
        client.login(username="viewer_rail", password="pass")

        with CaptureQueriesContext(connection) as captured:
            response = client.get(reverse("rail:filter_values", kwargs={"field": "cargo"}) + "?q=target")

        assert response.status_code == 200
        assert response.json()["values"] == ["target-cargo"]
        rail_selects = [
            query["sql"]
            for query in captured.captured_queries
            if "select" in query["sql"].lower() and "rail_shipments" in query["sql"].lower()
        ]
        assert any("LIKE" in sql.upper() and "LIMIT 201" in sql.upper() for sql in rail_selects)

    def test_list_context_contains_filter_query_safe_limit(self, client, viewer_user):
        client.login(username="viewer_rail", password="pass")
        response = client.get(reverse("rail:list"))

        assert response.status_code == 200
        assert response.context["filter_query_safe_limit"] == 3694

    def test_has_more_true_when_distinct_values_exceed_limit(self, client, viewer_user):
        for i in range(205):
            RailShipment.objects.create(
                departure_date="2026-01-01",
                wagon_number=f"{67000000 + i}",
                cargo=f"CARGO-{i:03d}",
                volume="100",
            )
        client.login(username="viewer_rail", password="pass")

        response = client.get(reverse("rail:filter_values", kwargs={"field": "cargo"}))

        assert response.status_code == 200
        data = response.json()
        assert len(data["values"]) == 200
        assert data["has_more"] is True

    def test_invalid_field_returns_404(self, client, viewer_user):
        client.login(username="viewer_rail", password="pass")
        response = client.get(reverse("rail:filter_values", kwargs={"field": "nonexistent_xyz"}))
        assert response.status_code == 404


@pytest.mark.django_db
class TestRailShipmentVolumeValidator:
    def test_volume_zero_fails_full_clean(self, django_user_model):
        import datetime
        from decimal import Decimal
        from django.core.exceptions import ValidationError
        user = django_user_model.objects.create_user(username="vval_rail", password="pass")
        from shipments_rail.models import RailShipment
        obj = RailShipment(
            departure_date=datetime.date(2024, 1, 15),
            wagon_number="11223344",
            cargo="ДР",
            receiver="Получатель",
            volume=Decimal("0"),
            created_by=user,
            updated_by=user,
        )
        with pytest.raises(ValidationError):
            obj.full_clean()

    def test_volume_positive_passes(self, django_user_model):
        import datetime
        from decimal import Decimal
        user = django_user_model.objects.create_user(username="vval_rail2", password="pass")
        from shipments_rail.models import RailShipment
        obj = RailShipment(
            departure_date=datetime.date(2024, 1, 15),
            wagon_number="11223344",
            cargo="ДР",
            receiver="Получатель",
            volume=Decimal("55.0"),
            created_by=user,
            updated_by=user,
        )
        obj.full_clean()  # должно пройти без исключения
